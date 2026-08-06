from datetime import date, datetime, timezone
from io import BytesIO
from unittest.mock import AsyncMock, patch

import pytest
from openpyxl import load_workbook
from sqlalchemy import func, select

from app.models import (
    ClientNameMapping,
    OwnedCourseOpening,
    OwnedSettlementCompareResultRow,
    SchedulerJobQueue,
    Settlement,
)
from app.routes.settlements import _refresh_settlements_consolidated


async def _seed_and_refresh_mv(db_session, *rows):
    for row in rows:
        db_session.add(row)
    await db_session.commit()
    await _refresh_settlements_consolidated(db_session)



@pytest.mark.asyncio
async def test_get_compare_reads_stored_results_only(
    test_client, authenticated_user, db_session
):
    headers = authenticated_user["headers"]
    extracted_at = datetime.now(timezone.utc)
    await _seed_and_refresh_mv(
        db_session,
        ClientNameMapping(institution_name="기관A", client_name="고객A"),
        OwnedCourseOpening(
            year=2024,
            institution_name="기관A",
            course_name="과정매칭",
            tra_start_date=date(2024, 5, 1),
            tra_end_date=date(2024, 5, 10),
            reg_course_man="2",
            extracted_at=extracted_at,
        ),
        OwnedCourseOpening(
            year=2024,
            institution_name="기관A",
            course_name="과정미정산",
            tra_start_date=date(2024, 6, 1),
            extracted_at=extracted_at,
        ),
        Settlement(
            purchase_ym="202405",
            purchase_year=2024,
            client_name="고객A",
            course_name="과정매칭",
            education_period_date=date(2024, 5, 1),
            headcount=2,
        ),
    )

    empty = await test_client.get(
        "/settlements/compare-owned?year=2024",
        headers=headers,
    )
    assert empty.status_code == 200, empty.text
    assert empty.json()["has_result"] is False
    assert empty.json()["total"] == 0

    export_before = await test_client.get(
        "/settlements/compare-owned/export?year=2024",
        headers=headers,
    )
    assert export_before.status_code == 404

    run = await test_client.post(
        "/settlements/compare-owned?year=2024",
        headers=headers,
    )
    assert run.status_code == 200, run.text
    body = run.json()
    assert body["has_result"] is True
    assert body["matched"] == 1
    assert body["partial"] == 0
    assert body["separate"] == 0
    assert body["unsettled"] == 1
    assert body["unmapped"] == 0
    assert body["compared_at"]

    stored = (
        await db_session.execute(
            select(OwnedSettlementCompareResultRow).where(
                OwnedSettlementCompareResultRow.year == 2024
            )
        )
    ).scalars().all()
    assert len(stored) == 2

    # GET does not rewrite rows
    again = await test_client.get(
        "/settlements/compare-owned?year=2024",
        headers=headers,
    )
    assert again.status_code == 200
    assert again.json()["matched"] == 1
    count_after_get = await db_session.scalar(
        select(func.count()).select_from(OwnedSettlementCompareResultRow).where(
            OwnedSettlementCompareResultRow.year == 2024
        )
    )
    assert count_after_get == 2

    matched = await test_client.get(
        "/settlements/compare-owned/items?year=2024&status=matched&page=1&size=50",
        headers=headers,
    )
    assert matched.status_code == 200, matched.text
    matched_body = matched.json()
    assert matched_body["total"] == 1
    assert matched_body["items"][0]["course_name"] == "과정매칭"
    assert matched_body["items"][0]["client_name"] == "고객A"

    export_res = await test_client.get(
        "/settlements/compare-owned/export?year=2024",
        headers=headers,
    )
    assert export_res.status_code == 200, export_res.text
    wb = load_workbook(BytesIO(export_res.content))
    assert wb.sheetnames == ["요약", "미정산", "일부정산", "맵핑없음", "정산됨"]
    assert wb["요약"]["B2"].value == 2024
    assert wb["요약"]["B3"].value == 2
    assert wb["미정산"]["C2"].value == "과정미정산"
    assert wb["정산됨"]["C2"].value == "과정매칭"
    wb.close()


@pytest.mark.asyncio
async def test_post_compare_replaces_year_rows(
    test_client, authenticated_user, db_session
):
    headers = authenticated_user["headers"]
    extracted_at = datetime.now(timezone.utc)
    await _seed_and_refresh_mv(
        db_session,
        ClientNameMapping(institution_name="기관A", client_name="고객A"),
        OwnedCourseOpening(
            year=2024,
            institution_name="기관A",
            course_name="과정1",
            tra_start_date=date(2024, 1, 1),
            extracted_at=extracted_at,
        ),
    )

    first = await test_client.post(
        "/settlements/compare-owned?year=2024", headers=headers
    )
    assert first.status_code == 200
    assert first.json()["unsettled"] == 1

    db_session.add(
        OwnedCourseOpening(
            year=2024,
            institution_name="기관A",
            course_name="과정2",
            tra_start_date=date(2024, 2, 1),
            extracted_at=extracted_at,
        )
    )
    await db_session.commit()

    second = await test_client.post(
        "/settlements/compare-owned?year=2024", headers=headers
    )
    assert second.status_code == 200
    assert second.json()["total"] == 2

    rows = (
        await db_session.execute(
            select(OwnedSettlementCompareResultRow).where(
                OwnedSettlementCompareResultRow.year == 2024
            )
        )
    ).scalars().all()
    assert len(rows) == 2
    assert {r.course_name for r in rows} == {"과정1", "과정2"}


@pytest.mark.asyncio
async def test_compare_auto_registers_identity_mapping(
    test_client, authenticated_user, db_session
):
    headers = authenticated_user["headers"]
    extracted_at = datetime.now(timezone.utc)
    await _seed_and_refresh_mv(
        db_session,
        OwnedCourseOpening(
            year=2024,
            institution_name="동일고객사",
            course_name="과정자동",
            tra_start_date=date(2024, 4, 1),
            reg_course_man="1",
            extracted_at=extracted_at,
        ),
        Settlement(
            purchase_ym="202404",
            purchase_year=2024,
            client_name="동일고객사",
            course_name="과정자동",
            education_period_date=date(2024, 4, 1),
            headcount=1,
        ),
    )

    res = await test_client.post(
        "/settlements/compare-owned?year=2024",
        headers=headers,
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["matched"] == 1
    assert body["unmapped"] == 0

    mapping = (
        await db_session.execute(
            select(ClientNameMapping).where(
                ClientNameMapping.institution_name == "동일고객사",
                ClientNameMapping.is_delete == False,  # noqa: E712
            )
        )
    ).scalars().one()
    assert mapping.client_name == "동일고객사"


@pytest.mark.asyncio
async def test_compare_auto_registers_mapping_ignoring_internal_spaces(
    test_client, authenticated_user, db_session
):
    headers = authenticated_user["headers"]
    extracted_at = datetime.now(timezone.utc)
    await _seed_and_refresh_mv(
        db_session,
        OwnedCourseOpening(
            year=2024,
            institution_name="에이 비씨",
            course_name="과정공백",
            tra_start_date=date(2024, 7, 1),
            reg_course_man="1",
            extracted_at=extracted_at,
        ),
        Settlement(
            purchase_ym="202407",
            purchase_year=2024,
            client_name="에이비씨",
            course_name="과정공백",
            education_period_date=date(2024, 7, 1),
            headcount=1,
        ),
    )

    res = await test_client.post(
        "/settlements/compare-owned?year=2024",
        headers=headers,
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["matched"] == 1
    assert body["unmapped"] == 0

    mapping = (
        await db_session.execute(
            select(ClientNameMapping).where(
                ClientNameMapping.institution_name == "에이 비씨",
                ClientNameMapping.is_delete == False,  # noqa: E712
            )
        )
    ).scalars().one()
    assert mapping.client_name == "에이비씨"


def test_normalize_name_for_mapping_removes_all_whitespace():
    from app.routes.settlements import _normalize_name_for_mapping

    assert _normalize_name_for_mapping(" 에이 비씨 ") == "에이비씨"
    assert _normalize_name_for_mapping("에이\t비씨\n") == "에이비씨"
    assert _normalize_name_for_mapping(None) == ""


def test_aggregate_owned_opening_rows_sums_headcount_and_skips_non_numeric():
    from scheduler.jobs.owned_course_opening_extract import (
        _aggregate_owned_opening_rows,
    )

    extracted_at = datetime(2024, 6, 1, tzinfo=timezone.utc)
    rows = _aggregate_owned_opening_rows(
        [
            {
                "institution_name": "A",
                "course_name": "과정",
                "tra_start_date": "2024-06-30",
                "tra_end_date": "2024-07-30",
                "reg_course_man": "88",
            },
            {
                "institution_name": "A",
                "course_name": "과정",
                "tra_start_date": "2024-06-30",
                "tra_end_date": "2024-07-30",
                "reg_course_man": "115",
            },
            {
                "institution_name": "A",
                "course_name": "과정",
                "tra_start_date": "2024-06-30",
                "tra_end_date": "2024-07-30",
                "reg_course_man": "N/A",
            },
            {
                "institution_name": "B",
                "course_name": "다른과정",
                "tra_start_date": "2024-01-01",
                "tra_end_date": None,
                "reg_course_man": "x",
            },
        ],
        year=2024,
        extracted_at=extracted_at,
    )
    assert len(rows) == 2
    by_inst = {r.institution_name: r for r in rows}
    assert by_inst["A"].reg_course_man == "203"
    assert by_inst["A"].tra_start_date == date(2024, 6, 30)
    assert by_inst["B"].reg_course_man is None


@pytest.mark.asyncio
async def test_compare_does_not_revive_soft_deleted_mapping(
    test_client, authenticated_user, db_session
):
    headers = authenticated_user["headers"]
    extracted_at = datetime.now(timezone.utc)
    await _seed_and_refresh_mv(
        db_session,
        ClientNameMapping(
            institution_name="삭제된기관",
            client_name="삭제된기관",
            is_delete=True,
        ),
        OwnedCourseOpening(
            year=2024,
            institution_name="삭제된기관",
            course_name="과정",
            tra_start_date=date(2024, 1, 1),
            extracted_at=extracted_at,
        ),
        Settlement(
            purchase_ym="202401",
            purchase_year=2024,
            client_name="삭제된기관",
            course_name="과정",
            education_period_date=date(2024, 1, 1),
        ),
    )

    res = await test_client.post(
        "/settlements/compare-owned?year=2024",
        headers=headers,
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["unmapped"] == 1
    assert body["matched"] == 0

    mapping = (
        await db_session.execute(
            select(ClientNameMapping).where(
                ClientNameMapping.institution_name == "삭제된기관"
            )
        )
    ).scalars().one()
    assert mapping.is_delete is True


@pytest.mark.asyncio
async def test_refresh_enqueues_scheduler_queue(
    test_client, authenticated_user, db_session
):
    headers = authenticated_user["headers"]

    res = await test_client.post(
        "/settlements/compare-owned/refresh?year=2024",
        headers=headers,
    )
    assert res.status_code == 202, res.text
    body = res.json()
    assert body["year"] == 2024
    assert body["status"] == "PENDING"
    queue_id = body["id"]

    queue = await db_session.get(SchedulerJobQueue, queue_id)
    assert queue is not None
    assert queue.job_key == "owned_course_opening_extract"
    assert queue.payload["year"] == 2024
    assert queue.payload["queue_id"] == queue_id

    again = await test_client.post(
        "/settlements/compare-owned/refresh?year=2024",
        headers=headers,
    )
    assert again.status_code == 202
    assert again.json()["id"] == queue_id


@pytest.mark.asyncio
async def test_extract_replaces_year_cache(db_session):
    from scheduler.jobs.owned_course_opening_extract import _run_extract

    old_at = datetime(2024, 1, 1, tzinfo=timezone.utc)
    db_session.add(
        OwnedCourseOpening(
            year=2024,
            institution_name="구기관",
            course_name="구과정",
            tra_start_date=date(2024, 1, 1),
            extracted_at=old_at,
        )
    )
    q = SchedulerJobQueue(
        job_key="owned_course_opening_extract",
        action="RUN_NOW",
        status="PROCESSING",
        payload={"year": 2024, "min_score": 0},
    )
    db_session.add(q)
    await db_session.commit()
    await db_session.refresh(q)
    q.payload = {**(q.payload or {}), "queue_id": q.id}
    await db_session.commit()

    mock_rows = [
        {
            "institution_name": "신기관",
            "course_name": "신과정",
            "tra_start_date": "2024-03-01",
            "tra_end_date": "2024-03-31",
            "reg_course_man": "5",
        },
        {
            "institution_name": "신기관",
            "course_name": "신과정",
            "tra_start_date": "2024-03-01",
            "tra_end_date": "2024-03-31",
            "reg_course_man": "7",
        },
        {
            "institution_name": "타기관",
            "course_name": "타과정",
            "tra_start_date": "2024-04-01",
            "tra_end_date": "2024-04-30",
            "reg_course_man": "3",
        },
    ]

    with (
        patch(
            "scheduler.jobs.owned_course_opening_extract._load_active_owned_names",
            new=AsyncMock(return_value=["신과정"]),
        ) as load_names,
        patch(
            "scheduler.jobs.owned_course_opening_extract._build_owned_scroll_body",
            return_value={"query": {"match_all": {}}},
        ) as build_body,
        patch(
            "scheduler.jobs.owned_course_opening_extract._scroll_owned_courses",
            new=AsyncMock(return_value=mock_rows),
        ),
        patch(
            "scheduler.jobs.owned_course_opening_extract.AsyncElasticsearch",
        ) as es_cls,
    ):
        es_cls.return_value.close = AsyncMock()
        result = await _run_extract(
            {"year": 2024, "min_score": 0, "queue_id": q.id}
        )

    load_names.assert_awaited_once()
    assert load_names.await_args.kwargs.get("usable_in_year") == 2024
    build_body.assert_called_once_with(
        ["신과정"], 2024, 0.0, has_reg_course_man=True
    )
    assert result["row_count"] == 2
    assert result["source_row_count"] == 3
    await db_session.refresh(q)
    assert q.payload["row_count"] == 2
    assert q.payload["source_row_count"] == 3
    assert q.payload.get("extracted_at")

    rows_2024 = (
        await db_session.execute(
            select(OwnedCourseOpening).where(OwnedCourseOpening.year == 2024)
        )
    ).scalars().all()
    assert len(rows_2024) == 2
    by_course = {r.course_name: r for r in rows_2024}
    assert by_course["신과정"].institution_name == "신기관"
    assert by_course["신과정"].reg_course_man == "12"
    assert by_course["타과정"].reg_course_man == "3"
