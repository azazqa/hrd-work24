"""업체(company) 스코프 분리 테스트."""

from __future__ import annotations

import io
from datetime import date, datetime, timezone
from unittest.mock import AsyncMock, patch

import pytest
from openpyxl import Workbook
from sqlalchemy import func, select

from app.models import (
    ClientNameMapping,
    OwnedCourse,
    OwnedCourseOpening,
    OwnedSettlementCompareResultRow,
    SeparateSettlement,
    Settlement,
)
from app.routes.courses import _load_active_owned_names
from app.routes.settlements import _refresh_settlements_consolidated
from scheduler.jobs.owned_course_opening_extract import _aggregate_owned_opening_rows
from tests.conftest import create_company


def _owned_xlsx(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["개발년도", "과정명", "구분"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _settlement_xlsx(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["매입년월", "고객사", "과정명", "인원"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _separate_xlsx(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["고객사", "과정명", "계약기간"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.asyncio
async def test_companies_crud(test_client, authenticated_user):
    headers = authenticated_user["headers"]

    created = await test_client.post(
        "/companies",
        headers=headers,
        json={"name": "알파교육", "is_active": True},
    )
    assert created.status_code == 201, created.text
    company_id = created.json()["id"]

    listed = await test_client.get("/companies?is_active=true", headers=headers)
    assert listed.status_code == 200
    assert any(i["id"] == company_id for i in listed.json()["items"])

    updated = await test_client.put(
        f"/companies/{company_id}",
        headers=headers,
        json={"name": "알파교육(수정)", "is_active": False},
    )
    assert updated.status_code == 200
    assert updated.json()["name"] == "알파교육(수정)"
    assert updated.json()["is_active"] is False

    deleted = await test_client.delete(f"/companies/{company_id}", headers=headers)
    assert deleted.status_code == 204

    missing = await test_client.get(f"/companies/{company_id}", headers=headers)
    assert missing.status_code == 404


@pytest.mark.asyncio
async def test_owned_course_import_upsert_scoped_by_company(
    test_client, authenticated_user, db_session
):
    headers = authenticated_user["headers"]
    company_a = await create_company(db_session, "업체A")
    company_b = await create_company(db_session, "업체B")

    content = _owned_xlsx([[2024, "동일과정", "구분1"]])

    res_a = await test_client.post(
        "/owned-courses/import",
        headers=headers,
        files={
            "file": (
                "a.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"company_id": str(company_a.id)},
    )
    assert res_a.status_code == 200, res_a.text
    assert res_a.json()["created"] == 1

    res_b = await test_client.post(
        "/owned-courses/import",
        headers=headers,
        files={
            "file": (
                "b.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"company_id": str(company_b.id)},
    )
    assert res_b.status_code == 200, res_b.text
    assert res_b.json()["created"] == 1

    total = await db_session.scalar(
        select(func.count())
        .select_from(OwnedCourse)
        .where(
            OwnedCourse.is_delete == False,  # noqa: E712
            OwnedCourse.course_name == "동일과정",
        )
    )
    assert total == 2

    again_a = await test_client.post(
        "/owned-courses/import",
        headers=headers,
        files={
            "file": (
                "a2.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"company_id": str(company_a.id)},
    )
    assert again_a.status_code == 200
    assert again_a.json()["updated"] == 1
    assert again_a.json()["created"] == 0


@pytest.mark.asyncio
async def test_settlement_import_does_not_delete_other_company(
    test_client, authenticated_user, db_session
):
    headers = authenticated_user["headers"]
    company_a = await create_company(db_session, "정산A")
    company_b = await create_company(db_session, "정산B")

    db_session.add(
        Settlement(
            company_id=company_b.id,
            purchase_ym="202401",
            purchase_year=2024,
            client_name="고객B",
            course_name="과정B",
            headcount=1,
        )
    )
    await db_session.commit()

    content = _settlement_xlsx([["202401", "고객A", "과정A", 2]])
    res = await test_client.post(
        "/settlements/import",
        headers=headers,
        files={
            "file": (
                "s.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"year": "2024", "company_id": str(company_a.id)},
    )
    assert res.status_code == 200, res.text
    assert res.json()["created"] == 1

    b_count = await db_session.scalar(
        select(func.count())
        .select_from(Settlement)
        .where(
            Settlement.is_delete == False,  # noqa: E712
            Settlement.company_id == company_b.id,
        )
    )
    assert b_count == 1

    listed_b = await test_client.get(
        f"/settlements?company_id={company_b.id}",
        headers=headers,
    )
    assert listed_b.status_code == 200
    assert listed_b.json()["total"] == 1


@pytest.mark.asyncio
async def test_separate_import_scoped_hard_delete(
    test_client, authenticated_user, db_session
):
    headers = authenticated_user["headers"]
    company_a = await create_company(db_session, "별도A")
    company_b = await create_company(db_session, "별도B")

    db_session.add(
        SeparateSettlement(
            company_id=company_b.id,
            client_name="고객B",
            course_name="과정B",
            contract_period="2024.01.01~2024.12.31",
        )
    )
    await db_session.commit()

    content = _separate_xlsx([["고객A", "과정A", "2024.01.01~2024.12.31"]])
    res = await test_client.post(
        "/settlements/separate/import",
        headers=headers,
        files={
            "file": (
                "sep.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"company_id": str(company_a.id)},
    )
    assert res.status_code == 200, res.text
    assert res.json()["created"] == 1

    b_count = await db_session.scalar(
        select(func.count()).select_from(SeparateSettlement).where(
            SeparateSettlement.company_id == company_b.id
        )
    )
    assert b_count == 1


@pytest.mark.asyncio
async def test_load_active_owned_names_filters_company_and_year(db_session):
    company_a = await create_company(db_session, "이름A")
    company_b = await create_company(db_session, "이름B")
    db_session.add_all(
        [
            OwnedCourse(
                company_id=company_a.id,
                course_name="A-2024",
                dev_year=2024,
                is_active=True,
            ),
            OwnedCourse(
                company_id=company_a.id,
                course_name="A-2020",
                dev_year=2020,
                is_active=True,
            ),
            OwnedCourse(
                company_id=company_b.id,
                course_name="B-2024",
                dev_year=2024,
                is_active=True,
            ),
        ]
    )
    await db_session.commit()

    names = await _load_active_owned_names(
        db_session, usable_in_year=2024, company_id=company_a.id
    )
    assert names == ["A-2024"]


def test_extract_aggregate_sets_company_id():
    extracted_at = datetime.now(timezone.utc)
    rows = _aggregate_owned_opening_rows(
        [
            {
                "institution_name": "기관",
                "course_name": "과정",
                "tra_start_date": "2024-01-01",
                "tra_end_date": "2024-01-10",
                "reg_course_man": "1",
            }
        ],
        year=2024,
        company_id=99,
        extracted_at=extracted_at,
    )
    assert len(rows) == 1
    assert rows[0].company_id == 99
    assert rows[0].year == 2024


@pytest.mark.asyncio
async def test_compare_independent_by_company(
    test_client, authenticated_user, db_session
):
    headers = authenticated_user["headers"]
    company_a = await create_company(db_session, "비교A")
    company_b = await create_company(db_session, "비교B")
    extracted_at = datetime.now(timezone.utc)

    db_session.add(ClientNameMapping(institution_name="기관X", client_name="고객X"))
    db_session.add_all(
        [
            OwnedCourseOpening(
                company_id=company_a.id,
                year=2024,
                institution_name="기관X",
                course_name="과정공용",
                tra_start_date=date(2024, 3, 1),
                reg_course_man="1",
                extracted_at=extracted_at,
            ),
            OwnedCourseOpening(
                company_id=company_b.id,
                year=2024,
                institution_name="기관X",
                course_name="과정공용",
                tra_start_date=date(2024, 3, 1),
                reg_course_man="1",
                extracted_at=extracted_at,
            ),
            Settlement(
                company_id=company_a.id,
                purchase_ym="202403",
                purchase_year=2024,
                client_name="고객X",
                course_name="과정공용",
                education_period_date=date(2024, 3, 1),
                headcount=1,
            ),
        ]
    )
    await db_session.commit()
    await _refresh_settlements_consolidated(db_session)

    run_a = await test_client.post(
        f"/settlements/compare-owned?year=2024&company_id={company_a.id}",
        headers=headers,
    )
    assert run_a.status_code == 200, run_a.text
    assert run_a.json()["matched"] == 1
    assert run_a.json()["unsettled"] == 0
    assert run_a.json()["company_id"] == company_a.id

    run_b = await test_client.post(
        f"/settlements/compare-owned?year=2024&company_id={company_b.id}",
        headers=headers,
    )
    assert run_b.status_code == 200, run_b.text
    assert run_b.json()["matched"] == 0
    assert run_b.json()["unsettled"] == 1
    assert run_b.json()["company_id"] == company_b.id

    get_a = await test_client.get(
        f"/settlements/compare-owned?year=2024&company_id={company_a.id}",
        headers=headers,
    )
    assert get_a.json()["matched"] == 1

    a_rows = (
        await db_session.execute(
            select(OwnedSettlementCompareResultRow).where(
                OwnedSettlementCompareResultRow.company_id == company_a.id,
                OwnedSettlementCompareResultRow.year == 2024,
            )
        )
    ).scalars().all()
    b_rows = (
        await db_session.execute(
            select(OwnedSettlementCompareResultRow).where(
                OwnedSettlementCompareResultRow.company_id == company_b.id,
                OwnedSettlementCompareResultRow.year == 2024,
            )
        )
    ).scalars().all()
    assert len(a_rows) == 1
    assert a_rows[0].status == "matched"
    assert len(b_rows) == 1
    assert b_rows[0].status == "unsettled"


@pytest.mark.asyncio
async def test_refresh_dedupes_by_year_and_company(
    test_client, authenticated_user, db_session
):
    headers = authenticated_user["headers"]
    company = await create_company(db_session, "추출업체")

    with patch(
        "scheduler.jobs.owned_course_opening_extract._load_active_owned_names",
        new_callable=AsyncMock,
        return_value=[],
    ):
        first = await test_client.post(
            f"/settlements/compare-owned/refresh?year=2024&company_id={company.id}",
            headers=headers,
        )
        assert first.status_code == 202, first.text
        second = await test_client.post(
            f"/settlements/compare-owned/refresh?year=2024&company_id={company.id}",
            headers=headers,
        )
        assert second.status_code == 202
        assert first.json()["id"] == second.json()["id"]
        assert first.json()["company_id"] == company.id
