from datetime import date, datetime, timezone
from decimal import Decimal

import pytest
from sqlalchemy import func, select

from app.models import (
    ClientNameMapping,
    OwnedCourseOpening,
    Settlement,
    SettlementConsolidated,
)
from app.routes.settlements import _refresh_settlements_consolidated


def _same_course_kwargs(**overrides):
    base = dict(
        purchase_ym="202405",
        purchase_year=2024,
        sales_ym="202404",
        client_name="고객합산",
        course_name="분할과정",
        education_period="2024.05.01~2024.05.31",
        education_period_date=date(2024, 5, 1),
        headcount=1,
        base_tuition=Decimal("1000"),
        textbook_fee=Decimal("0"),
        exclude_amount=Decimal("0"),
        share_rate=Decimal("0.3"),
        net_sales=Decimal("700"),
        settlement_rate=Decimal("0.5"),
        settlement_amount=Decimal("350"),
        note="동일",
        sales_rep="홍길동",
    )
    base.update(overrides)
    return base


@pytest.mark.asyncio
async def test_consolidated_mv_sums_identical_rows(db_session):
    for _ in range(10):
        db_session.add(Settlement(**_same_course_kwargs()))
    await db_session.commit()
    await _refresh_settlements_consolidated(db_session)

    rows = (
        await db_session.execute(
            select(SettlementConsolidated).where(
                SettlementConsolidated.client_name == "고객합산",
                SettlementConsolidated.course_name == "분할과정",
            )
        )
    ).scalars().all()
    assert len(rows) == 1
    assert rows[0].headcount == 10
    assert rows[0].base_tuition == Decimal("1000")
    assert rows[0].share_rate == Decimal("0.3")
    assert rows[0].settlement_rate == Decimal("0.5")
    assert rows[0].net_sales == Decimal("7000")
    assert rows[0].settlement_amount == Decimal("3500")


@pytest.mark.asyncio
async def test_consolidated_mv_sums_when_amounts_differ_but_rates_same(db_session):
    """순매출·정산액만 달라도 요율 키가 같으면 SUM 한다."""
    db_session.add(
        Settlement(
            **_same_course_kwargs(
                net_sales=Decimal("700"),
                settlement_amount=Decimal("350"),
            )
        )
    )
    db_session.add(
        Settlement(
            **_same_course_kwargs(
                net_sales=Decimal("800"),
                settlement_amount=Decimal("351"),
            )
        )
    )
    await db_session.commit()
    await _refresh_settlements_consolidated(db_session)

    rows = (
        await db_session.execute(
            select(SettlementConsolidated).where(
                SettlementConsolidated.client_name == "고객합산"
            )
        )
    ).scalars().all()
    assert len(rows) == 1
    assert rows[0].headcount == 2
    assert rows[0].base_tuition == Decimal("1000")
    assert rows[0].net_sales == Decimal("1500")
    assert rows[0].settlement_amount == Decimal("701")


@pytest.mark.asyncio
async def test_consolidated_mv_keeps_note_diff_separate(db_session):
    """비고가 다르면 다른 과정으로 유지한다."""
    db_session.add(Settlement(**_same_course_kwargs(note="비고A")))
    db_session.add(Settlement(**_same_course_kwargs(note="비고B")))
    await db_session.commit()
    await _refresh_settlements_consolidated(db_session)

    count = await db_session.scalar(
        select(func.count()).select_from(SettlementConsolidated).where(
            SettlementConsolidated.client_name == "고객합산"
        )
    )
    assert count == 2


@pytest.mark.asyncio
async def test_compare_uses_consolidated_mv(
    test_client, authenticated_user, db_session
):
    headers = authenticated_user["headers"]
    extracted_at = datetime.now(timezone.utc)
    db_session.add(
        ClientNameMapping(institution_name="고객합산", client_name="고객합산")
    )
    db_session.add(
        OwnedCourseOpening(
            year=2024,
            institution_name="고객합산",
            course_name="분할과정",
            tra_start_date=date(2024, 5, 1),
            reg_course_man="10",
            extracted_at=extracted_at,
        )
    )
    for _ in range(10):
        db_session.add(Settlement(**_same_course_kwargs()))
    await db_session.commit()
    await _refresh_settlements_consolidated(db_session)

    res = await test_client.post(
        "/settlements/compare-owned?year=2024",
        headers=headers,
    )
    assert res.status_code == 200, res.text
    assert res.json()["matched"] == 1
    assert res.json()["partial"] == 0
    assert res.json()["unsettled"] == 0


@pytest.mark.asyncio
async def test_compare_partial_when_headcount_differs(
    test_client, authenticated_user, db_session
):
    headers = authenticated_user["headers"]
    extracted_at = datetime.now(timezone.utc)
    db_session.add(
        ClientNameMapping(institution_name="고객합산", client_name="고객합산")
    )
    db_session.add(
        OwnedCourseOpening(
            year=2024,
            institution_name="고객합산",
            course_name="분할과정",
            tra_start_date=date(2024, 5, 1),
            reg_course_man="10",
            extracted_at=extracted_at,
        )
    )
    for _ in range(3):
        db_session.add(Settlement(**_same_course_kwargs()))
    await db_session.commit()
    await _refresh_settlements_consolidated(db_session)

    res = await test_client.post(
        "/settlements/compare-owned?year=2024",
        headers=headers,
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["matched"] == 0
    assert body["partial"] == 1
    assert body["unsettled"] == 0

    items = await test_client.get(
        "/settlements/compare-owned/items?year=2024&status=partial&page=1&size=50",
        headers=headers,
    )
    assert items.status_code == 200
    assert items.json()["total"] == 1
    assert items.json()["items"][0]["status"] == "partial"
    assert items.json()["items"][0]["reg_course_man"] == "10"


@pytest.mark.asyncio
async def test_compare_matches_ignoring_spaces_in_course_and_client(
    test_client, authenticated_user, db_session
):
    """과정명·고객사명 내부 공백 차이만 있어도 정산 키로 매칭한다."""
    headers = authenticated_user["headers"]
    extracted_at = datetime.now(timezone.utc)
    db_session.add(
        ClientNameMapping(
            institution_name="(SE)에스이사이버평생교육원",
            client_name="(주)에스이 스페셜에듀",
        )
    )
    db_session.add(
        OwnedCourseOpening(
            year=2024,
            institution_name="(SE)에스이사이버평생교육원",
            course_name="간호사가 꼭 알아야 할 현장 실무",
            tra_start_date=date(2024, 6, 30),
            reg_course_man="88",
            extracted_at=extracted_at,
        )
    )
    db_session.add(
        Settlement(
            **_same_course_kwargs(
                purchase_ym="202406",
                purchase_year=2024,
                client_name="(주)에스이스페셜에듀",
                course_name="간호사가꼭알아야할현장실무",
                education_period="2024.06.30",
                education_period_date=date(2024, 6, 30),
                headcount=161,
            )
        )
    )
    await db_session.commit()
    await _refresh_settlements_consolidated(db_session)

    res = await test_client.post(
        "/settlements/compare-owned?year=2024",
        headers=headers,
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["matched"] == 0
    assert body["partial"] == 1
    assert body["unsettled"] == 0


def test_normalize_course_name_for_compare():
    from app.routes.settlements import _normalize_course_name_for_compare

    assert (
        _normalize_course_name_for_compare(
            "[2023기업직업훈련카드]마음을 움직이는 카피라이팅과 글쓰기"
        )
        == "마음을움직이는카피라이팅과글쓰기"
    )
    assert (
        _normalize_course_name_for_compare(
            "평범한 회사와 직장인을 위한 정보 보안 기본서 (기업직업훈련카드)"
        )
        == "평범한회사와직장인을위한정보보안기본서"
    )
    assert (
        _normalize_course_name_for_compare(
            "[2023기업직업훈련카드]평범한 회사와 직장인을 위한 정보 보안 기본서_23.12"
        )
        == "평범한회사와직장인을위한정보보안기본서"
    )
    assert (
        _normalize_course_name_for_compare("[A][B]카피 라이팅")
        == "카피라이팅"
    )
    assert _normalize_course_name_for_compare("   ") == ""
    assert _normalize_course_name_for_compare(None) == ""
    assert (
        _normalize_course_name_for_compare("간호사가 꼭 알아야 할 현장 실무")
        == _normalize_course_name_for_compare(
            "[태그]간호사가꼭알아야할현장실무_24.06"
        )
    )
    # 한글·영문·숫자 외 특수문자 제거
    assert (
        _normalize_course_name_for_compare("정보-보안 / 기본서! (카드)")
        == "정보보안기본서"
    )
    assert (
        _normalize_course_name_for_compare("AI·머신러닝·101")
        == "AI머신러닝101"
    )
    assert (
        _normalize_course_name_for_compare("실무&이론")
        == _normalize_course_name_for_compare("실무 이론")
    )


@pytest.mark.asyncio
async def test_compare_matches_normalized_course_name_tags(
    test_client, authenticated_user, db_session
):
    """선행[]·후행()·_접미·특수문자 정규화 후 과정명이 같으면 매칭한다."""
    headers = authenticated_user["headers"]
    extracted_at = datetime.now(timezone.utc)
    db_session.add(
        ClientNameMapping(institution_name="정규화기관", client_name="정규화고객")
    )
    db_session.add(
        OwnedCourseOpening(
            year=2024,
            institution_name="정규화기관",
            course_name="간호사가 꼭 알아야 할 현장 실무",
            tra_start_date=date(2024, 6, 30),
            reg_course_man="88",
            extracted_at=extracted_at,
        )
    )
    db_session.add(
        Settlement(
            **_same_course_kwargs(
                purchase_ym="202406",
                purchase_year=2024,
                client_name="정규화고객",
                course_name="[태그]간호사·가 꼭 알아야 할 현장-실무!_24.06",
                education_period="2024.06.30",
                education_period_date=date(2024, 6, 30),
                headcount=161,
            )
        )
    )
    await db_session.commit()
    await _refresh_settlements_consolidated(db_session)

    res = await test_client.post(
        "/settlements/compare-owned?year=2024",
        headers=headers,
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["matched"] == 0
    assert body["partial"] == 1
    assert body["unsettled"] == 0


@pytest.mark.asyncio
async def test_export_settlements_consolidated_xlsx(
    test_client, authenticated_user, db_session
):
    headers = authenticated_user["headers"]
    for _ in range(3):
        db_session.add(Settlement(**_same_course_kwargs()))
    db_session.add(
        Settlement(
            **_same_course_kwargs(
                client_name="다른고객",
                course_name="다른과정",
                settlement_amount=Decimal("100"),
            )
        )
    )
    await db_session.commit()
    await _refresh_settlements_consolidated(db_session)

    res = await test_client.get("/settlements/export", headers=headers)
    assert res.status_code == 200, res.text
    assert "spreadsheetml" in (res.headers.get("content-type") or "")

    from io import BytesIO

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(res.content))
    ws = wb.active
    assert ws.title == "정산"
    headers_row = [c.value for c in ws[1]]
    assert headers_row[0] == "매입년월"
    assert headers_row[4] == "교육기간"
    assert headers_row[5] == "교육기간(변환)"
    assert headers_row[6] == "인원"
    # 3 identical -> 1 row (headcount 3) + 1 other = 2 data rows
    assert ws.max_row == 3
    # find 분할과정 row
    data = [[c.value for c in row] for row in ws.iter_rows(min_row=2, values_only=False)]
    by_course = {row[3]: row for row in data}
    assert by_course["분할과정"][5] == "2024-05-01"
    assert by_course["분할과정"][6] == 3
    assert by_course["다른과정"][6] == 1
    wb.close()
