import io
from datetime import date
from decimal import Decimal

import pytest
from openpyxl import Workbook
from sqlalchemy import func, select

from app.models import SeparateSettlement
from app.routes.separate_settlements import (
    EXCEL_HEADERS,
    _resolve_settlement_rate,
    _row_to_fields,
)


def _xlsx_bytes(headers: list[str], rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


HEADERS = [label for label, _ in EXCEL_HEADERS]

EXAMPLE_ROW = [
    "2024-02-22",
    "김승훈",
    "연간임대",
    "(주)한국이러닝인재개발원",
    "4차 산업혁명의 본질과 미래",
    "4차 산업혁명의 본질과 미래",
    10000000,
    "25%",
    2500000,
    "2022.08.29~2023.08.28",
    "",  # 차감액
    2500000,
    "4차 산업혁명의 본질과 미래(연간임대)",
    2500000,
    250000,
    2750000,
    "rmptax@thermp.co.kr",
]


def test_resolve_settlement_rate_numeric_percent():
    rate, raw = _resolve_settlement_rate("25%", Decimal("10000000"), Decimal("2500000"))
    assert rate == Decimal("0.25")
    assert raw == "25%"


def test_resolve_settlement_rate_from_amounts_when_non_numeric():
    rate, raw = _resolve_settlement_rate(
        "별도협의", Decimal("10000000"), Decimal("2500000")
    )
    assert rate == Decimal("0.25")
    assert raw == "별도협의"


def test_resolve_settlement_rate_none_when_base_zero():
    rate, raw = _resolve_settlement_rate("별도협의", Decimal("0"), Decimal("2500000"))
    assert rate is None
    assert raw == "별도협의"


def test_row_to_fields_example():
    fields = _row_to_fields(
        {
            "invoice_deadline_date": "2024-02-22",
            "sales_rep": "김승훈",
            "category": "연간임대",
            "client_name": "(주)한국이러닝인재개발원",
            "business_detail": "4차 산업혁명의 본질과 미래",
            "course_name": "4차 산업혁명의 본질과 미래",
            "base_revenue": 10000000,
            "settlement_rate": "25%",
            "calculated_amount": 2500000,
            "contract_period": "2022.08.29~2023.08.28",
            "deduction_amount": "",
            "final_amount": 2500000,
            "invoice_item": "4차 산업혁명의 본질과 미래(연간임대)",
            "supply_amount": 2500000,
            "tax_amount": 250000,
            "total_amount": 2750000,
            "invoice_issuer": "rmptax@thermp.co.kr",
        }
    )
    assert fields["invoice_deadline_date"].isoformat() == "2024-02-22"
    assert fields["invoice_deadline_year"] == 2024
    assert fields["settlement_rate"] == Decimal("0.25")
    assert fields["settlement_rate_raw"] == "25%"
    assert fields["deduction_amount"] is None
    assert fields["client_name"] == "(주)한국이러닝인재개발원"


@pytest.mark.asyncio
async def test_import_example_and_full_replace(
    test_client, authenticated_user, db_session
):
    headers = authenticated_user["headers"]

    first = _xlsx_bytes(HEADERS, [EXAMPLE_ROW])
    res1 = await test_client.post(
        "/settlements/separate/import",
        headers=headers,
        files={
            "file": (
                "separate.xlsx",
                first,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert res1.status_code == 200, res1.text
    body1 = res1.json()
    assert body1["deleted"] == 0
    assert body1["created"] == 1
    assert body1["failed"] == 0

    list1 = await test_client.get(
        "/settlements/separate",
        headers=headers,
        params={"page": 1, "size": 20},
    )
    assert list1.status_code == 200
    item = list1.json()["items"][0]
    assert item["client_name"] == "(주)한국이러닝인재개발원"
    assert item["invoice_deadline_year"] == 2024
    assert Decimal(str(item["settlement_rate"])) == Decimal("0.25")
    assert item["settlement_rate_raw"] == "25%"

    second_row = list(EXAMPLE_ROW)
    second_row[3] = "교체고객"
    second_row[5] = "교체과정"
    second_row[7] = "별도협의"
    second = _xlsx_bytes(HEADERS, [second_row])
    res2 = await test_client.post(
        "/settlements/separate/import",
        headers=headers,
        files={
            "file": (
                "separate.xlsx",
                second,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert res2.status_code == 200, res2.text
    body2 = res2.json()
    assert body2["deleted"] == 1
    assert body2["created"] == 1

    count = await db_session.scalar(select(func.count()).select_from(SeparateSettlement))
    assert count == 1
    row = (
        await db_session.execute(select(SeparateSettlement))
    ).scalars().one()
    assert row.client_name == "교체고객"
    assert row.course_name == "교체과정"
    assert row.settlement_rate_raw == "별도협의"
    assert row.settlement_rate == Decimal("0.25")


@pytest.mark.asyncio
async def test_list_separate_settlements_filters(
    test_client, authenticated_user, db_session
):
    headers = authenticated_user["headers"]
    db_session.add_all(
        [
            SeparateSettlement(
                invoice_deadline_date=date(2024, 2, 22),
                invoice_deadline_year=2024,
                client_name="한국이러닝",
                course_name="과정A",
            ),
            SeparateSettlement(
                invoice_deadline_date=date(2024, 5, 1),
                invoice_deadline_year=2024,
                client_name="다른고객",
                course_name="과정B",
            ),
            SeparateSettlement(
                invoice_deadline_date=date(2025, 1, 1),
                invoice_deadline_year=2025,
                client_name="한국이러닝",
                course_name="과정C",
            ),
        ]
    )
    await db_session.commit()

    res = await test_client.get(
        "/settlements/separate",
        headers=headers,
        params={"year": 2024, "client_name": "한국", "page": 1, "size": 20},
    )
    assert res.status_code == 200
    data = res.json()
    assert data["total"] == 1
    assert data["items"][0]["course_name"] == "과정A"

    res2 = await test_client.get(
        "/settlements/separate",
        headers=headers,
        params={"course_name": "과정B", "page": 1, "size": 20},
    )
    assert res2.status_code == 200
    assert res2.json()["total"] == 1
    assert res2.json()["items"][0]["client_name"] == "다른고객"


@pytest.mark.asyncio
async def test_export_and_template(test_client, authenticated_user, db_session):
    headers = authenticated_user["headers"]
    content = _xlsx_bytes(HEADERS, [EXAMPLE_ROW])
    await test_client.post(
        "/settlements/separate/import",
        headers=headers,
        files={
            "file": (
                "separate.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    template = await test_client.get(
        "/settlements/separate/import/template",
        headers=headers,
    )
    assert template.status_code == 200
    assert "spreadsheetml" in (template.headers.get("content-type") or "")

    export = await test_client.get(
        "/settlements/separate/export",
        headers=headers,
    )
    assert export.status_code == 200, export.text
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(export.content))
    ws = wb.active
    assert [c.value for c in ws[1]] == HEADERS
    assert ws.max_row == 2
    assert ws.cell(2, 4).value == "(주)한국이러닝인재개발원"
    assert ws.cell(2, 8).value == "25%"  # settlement_rate_raw
    wb.close()
