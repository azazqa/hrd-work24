from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from fastapi_pagination.ext.sqlalchemy import apaginate
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, ConfigDict, Field
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_async_session
from app.models import OwnedCourse, User
from app.pagination import MAX_PAGE_SIZE, Page, Params
from app.users import current_active_user

router = APIRouter()

EXCEL_HEADERS: list[tuple[str, str]] = [
    ("개발년도", "dev_year"),
    ("개발차수", "dev_round"),
    ("심사차수", "review_round"),
    ("구분", "division"),
    ("ncs과정개발구분", "ncs_dev_category"),
    ("과정명", "course_name"),
    ("차시", "session_count"),
    ("평가활동훈련분량", "eval_training_volume"),
    ("결과", "result"),
    ("등급(최초)", "grade_initial"),
    ("등급(23)", "grade_23"),
    ("NCS(신청)", "ncs_applied"),
    ("NCS(인정)", "ncs_approved"),
]

HEADER_TO_FIELD = {label: field for label, field in EXCEL_HEADERS}
FIELD_TO_HEADER = {field: label for label, field in EXCEL_HEADERS}


class OwnedCourseListItem(BaseModel):
    id: int
    dev_year: int | None
    division: str | None
    course_name: str
    is_active: bool


class OwnedCourseRead(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: int
    dev_year: int | None
    dev_round: str | None
    review_round: str | None
    division: str | None
    ncs_dev_category: str | None
    course_name: str
    session_count: int | None
    eval_training_volume: str | None
    result: str | None
    grade_initial: str | None
    grade_23: str | None
    ncs_applied: str | None
    ncs_approved: str | None
    is_active: bool
    created_at: datetime
    updated_at: datetime


class OwnedCourseCreate(BaseModel):
    dev_year: int | None = None
    dev_round: str | None = Field(default=None, max_length=50)
    review_round: str | None = Field(default=None, max_length=50)
    division: str | None = Field(default=None, max_length=100)
    ncs_dev_category: str | None = Field(default=None, max_length=100)
    course_name: str = Field(min_length=1, max_length=500)
    session_count: int | None = None
    eval_training_volume: str | None = Field(default=None, max_length=100)
    result: str | None = Field(default=None, max_length=100)
    grade_initial: str | None = Field(default=None, max_length=50)
    grade_23: str | None = Field(default=None, max_length=50)
    ncs_applied: str | None = Field(default=None, max_length=100)
    ncs_approved: str | None = Field(default=None, max_length=100)
    is_active: bool = True


class OwnedCourseUpdate(BaseModel):
    dev_year: int | None = None
    dev_round: str | None = Field(default=None, max_length=50)
    review_round: str | None = Field(default=None, max_length=50)
    division: str | None = Field(default=None, max_length=100)
    ncs_dev_category: str | None = Field(default=None, max_length=100)
    course_name: str | None = Field(default=None, min_length=1, max_length=500)
    session_count: int | None = None
    eval_training_volume: str | None = Field(default=None, max_length=100)
    result: str | None = Field(default=None, max_length=100)
    grade_initial: str | None = Field(default=None, max_length=50)
    grade_23: str | None = Field(default=None, max_length=50)
    ncs_applied: str | None = Field(default=None, max_length=100)
    ncs_approved: str | None = Field(default=None, max_length=100)
    is_active: bool | None = None


class OwnedCourseImportError(BaseModel):
    row: int
    message: str


class OwnedCourseImportResult(BaseModel):
    created: int
    updated: int
    failed: int
    errors: list[OwnedCourseImportError]


def _to_list_item(row: OwnedCourse) -> OwnedCourseListItem:
    return OwnedCourseListItem(
        id=row.id,
        dev_year=row.dev_year,
        division=row.division,
        course_name=row.course_name,
        is_active=row.is_active,
    )


def _transform_list(rows: list[OwnedCourse]) -> list[OwnedCourseListItem]:
    return [_to_list_item(r) for r in rows]


def _parse_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    s = str(value).strip()
    if not s:
        return None
    return int(float(s))


def _parse_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _row_to_fields(row_values: dict[str, Any]) -> dict[str, Any]:
    data: dict[str, Any] = {}
    for field in FIELD_TO_HEADER:
        raw = row_values.get(field)
        if field in ("dev_year", "session_count"):
            parsed = _parse_int(raw)
            if parsed is not None:
                data[field] = parsed
        else:
            parsed = _parse_str(raw)
            if parsed is not None:
                data[field] = parsed
    return data


async def _find_existing_for_import(
    session: AsyncSession,
    *,
    course_name: str,
    dev_year: int | None,
) -> OwnedCourse | str:
    """과정명(+개발년도)으로 기존 행을 찾는다. 'ambiguous'면 동명이 2건 이상."""
    stmt = select(OwnedCourse).where(
        OwnedCourse.is_delete == False,  # noqa: E712
        OwnedCourse.course_name == course_name,
    )
    if dev_year is not None:
        stmt = stmt.where(OwnedCourse.dev_year == dev_year)
    rows = list((await session.scalars(stmt)).all())
    if len(rows) > 1:
        return "ambiguous"
    return rows[0] if rows else None


@router.get("", response_model=Page[OwnedCourseListItem])
async def list_owned_courses(
    page: int = 1,
    size: int = 20,
    q: str | None = Query(default=None, description="과정명 검색"),
    is_active: bool | None = Query(default=None),
    dev_year: int | None = Query(default=None),
    division: str | None = Query(default=None),
    session: AsyncSession = Depends(get_async_session),
    _: User = Depends(current_active_user),
):
    if size < 1:
        raise HTTPException(status_code=400, detail="size must be >= 1")
    if size > MAX_PAGE_SIZE:
        raise HTTPException(status_code=400, detail=f"size must be <= {MAX_PAGE_SIZE}")

    params = Params(page=page, size=size)
    stmt = select(OwnedCourse).where(OwnedCourse.is_delete == False)  # noqa: E712
    if q and q.strip():
        stmt = stmt.where(OwnedCourse.course_name.ilike(f"%{q.strip()}%"))
    if is_active is not None:
        stmt = stmt.where(OwnedCourse.is_active == is_active)
    if dev_year is not None:
        stmt = stmt.where(OwnedCourse.dev_year == dev_year)
    if division and division.strip():
        stmt = stmt.where(OwnedCourse.division == division.strip())
    stmt = stmt.order_by(OwnedCourse.id.desc())
    return await apaginate(session, stmt, params, transformer=_transform_list)


@router.get("/import/template")
async def download_import_template(
    _: User = Depends(current_active_user),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "보유과정"
    ws.append([label for label, _ in EXCEL_HEADERS])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="owned_courses_template.xlsx"'
        },
    )


@router.post("/import", response_model=OwnedCourseImportResult)
async def import_owned_courses(
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_async_session),
    _: User = Depends(current_active_user),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="xlsx 파일만 업로드할 수 있습니다.")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"엑셀 파일을 읽을 수 없습니다: {exc}") from exc

    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        raise HTTPException(status_code=400, detail="헤더 행이 없습니다.")

    col_map: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        label = str(cell).strip()
        field = HEADER_TO_FIELD.get(label)
        if field:
            col_map[idx] = field

    if "course_name" not in col_map.values():
        raise HTTPException(status_code=400, detail="과정명 열이 필요합니다.")

    created = 0
    updated = 0
    failed = 0
    errors: list[OwnedCourseImportError] = []

    for row_num, row in enumerate(rows, start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue

        row_values: dict[str, Any] = {}
        for idx, field in col_map.items():
            if idx < len(row):
                row_values[field] = row[idx]

        course_name = _parse_str(row_values.get("course_name"))
        if not course_name:
            failed += 1
            errors.append(OwnedCourseImportError(row=row_num, message="과정명이 비어 있습니다."))
            continue

        try:
            fields = _row_to_fields(row_values)
            fields["course_name"] = course_name
            fields["is_active"] = True

            dev_year = fields.get("dev_year")
            existing = await _find_existing_for_import(
                session, course_name=course_name, dev_year=dev_year
            )
            if existing == "ambiguous":
                failed += 1
                errors.append(
                    OwnedCourseImportError(
                        row=row_num,
                        message="동일 과정명(개발년도)이 여러 건입니다. 화면에서 개별 수정하세요.",
                    )
                )
                continue
            if existing is not None:
                for k, v in fields.items():
                    setattr(existing, k, v)
                updated += 1
            else:
                session.add(OwnedCourse(**fields))
                created += 1
        except Exception as exc:
            failed += 1
            errors.append(
                OwnedCourseImportError(row=row_num, message=str(exc) or "처리 실패")
            )

    await session.commit()
    wb.close()
    return OwnedCourseImportResult(
        created=created, updated=updated, failed=failed, errors=errors
    )


@router.get("/{course_id}", response_model=OwnedCourseRead)
async def get_owned_course(
    course_id: int,
    session: AsyncSession = Depends(get_async_session),
    _: User = Depends(current_active_user),
):
    row = await session.get(OwnedCourse, course_id)
    if row is None or row.is_delete:
        raise HTTPException(status_code=404, detail="보유 과정을 찾을 수 없습니다.")
    return row


@router.post("", response_model=OwnedCourseRead, status_code=201)
async def create_owned_course(
    body: OwnedCourseCreate,
    session: AsyncSession = Depends(get_async_session),
    _: User = Depends(current_active_user),
):
    row = OwnedCourse(**body.model_dump())
    session.add(row)
    await session.commit()
    await session.refresh(row)
    return row


@router.put("/{course_id}", response_model=OwnedCourseRead)
async def update_owned_course(
    course_id: int,
    body: OwnedCourseUpdate,
    session: AsyncSession = Depends(get_async_session),
    _: User = Depends(current_active_user),
):
    row = await session.get(OwnedCourse, course_id)
    if row is None or row.is_delete:
        raise HTTPException(status_code=404, detail="보유 과정을 찾을 수 없습니다.")

    data = body.model_dump(exclude_unset=True)
    if not data:
        raise HTTPException(status_code=400, detail="수정할 필드가 없습니다.")

    for k, v in data.items():
        setattr(row, k, v)
    await session.commit()
    await session.refresh(row)
    return row


@router.delete("/{course_id}", status_code=204)
async def delete_owned_course(
    course_id: int,
    session: AsyncSession = Depends(get_async_session),
    _: User = Depends(current_active_user),
):
    row = await session.get(OwnedCourse, course_id)
    if row is None or row.is_delete:
        raise HTTPException(status_code=404, detail="보유 과정을 찾을 수 없습니다.")
    row.is_delete = True
    await session.commit()
