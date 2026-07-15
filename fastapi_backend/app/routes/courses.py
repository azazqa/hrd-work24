from __future__ import annotations

import asyncio
import logging
import os
import re
import tempfile
from collections.abc import AsyncIterator
from datetime import datetime

from elasticsearch import NotFoundError
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from fastapi_pagination.ext.sqlalchemy import apaginate
from openpyxl import Workbook
from pydantic import BaseModel, ConfigDict, Field
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.database import get_async_session
from app.es import get_es
from app.models import (
    CourseExportJob,
    OwnedCourse,
    SchedulerJob,
    SchedulerJobQueue,
    User,
)
from app.pagination import MAX_PAGE_SIZE, Page, Params
from app.users import current_active_user, current_superuser
from app.work24 import (
    INDEX_TEST_PARAMS,
    ensure_course_index,
    fetch_courses,
    index_courses,
    iter_month_ranges,
    parse_year_month,
)

logger = logging.getLogger(__name__)

router = APIRouter()

_DATE_RE = re.compile(r"^\d{8}$")
_WILDCARD_ESCAPE_RE = re.compile(r"([\\*?])")

MAX_EXPORT_ROWS = 100_000
SCROLL_BATCH_SIZE = 1_000
SCROLL_KEEPALIVE = "2m"
EXPORT_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
EXPORT_OVER_LIMIT_MESSAGE = (
    "조회 결과가 100,000건을 초과합니다. 검색 조건을 좁혀 주세요."
)

EXPORT_HEADERS: list[tuple[str, str]] = [
    ("훈련기관명", "inst_name"),
    ("훈련과정명", "course_name"),
    ("훈련과정차수", "trpr_degr"),
    ("훈련시작일", "tra_start_date"),
    ("훈련종료일", "tra_end_date"),
    ("주소", "address"),
    ("전화번호", "tel_no"),
    ("정원", "yard_man"),
    ("수강신청 인원", "reg_course_man"),
    ("실제 훈련비", "real_man"),
    ("Work24 링크", "title_link"),
    ("훈련기관ID", "trainst_cst_id"),
    ("훈련과정ID", "trpr_id"),
]


class CourseSearchHit(BaseModel):
    model_config = ConfigDict(extra="allow")

    trpr_id: str | None = None
    trng_crse_nm: str | None = None
    inst_nm: str | None = None
    score: float


class CourseIndexResult(BaseModel):
    fetched: int
    indexed: int
    total_count: int


class LegacyIndexRequest(BaseModel):
    start_month: str = Field(..., pattern=r"^\d{4}-\d{2}$")
    end_month: str = Field(..., pattern=r"^\d{4}-\d{2}$")


class LegacyIndexResponse(BaseModel):
    queue_ids: list[int]
    start_month: str
    end_month: str
    month_count: int
    message: str


class CourseExportJobCreate(BaseModel):
    memo: str | None = Field(default=None, max_length=1000)
    srch_tra_st_dt: str | None = None
    srch_tra_end_dt: str | None = None
    srch_tra_organ_nm: str | None = None
    srch_tra_process_nm: str | None = None
    has_reg_course_man: bool = False
    owned_year: int | None = Field(default=None, ge=2023, le=2100)
    min_score: float = Field(default=0, ge=0)


class CourseExportJobRead(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: int
    status: str
    memo: str | None
    conditions_summary: str | None
    row_count: int | None
    file_name: str | None
    file_size: int | None
    error_message: str | None
    created_at: datetime
    updated_at: datetime


QUEUE_ACTION_RUN_NOW = "RUN_NOW"
QUEUE_STATUS_PENDING = "PENDING"


class CourseListItem(BaseModel):
    trainst_cst_id: str | None = None
    trpr_id: str | None = None
    trpr_degr: str | None = None
    course_name: str | None = None
    inst_name: str | None = None
    tra_start_date: str | None = None
    tra_end_date: str | None = None
    address: str | None = None
    tel_no: str | None = None
    title_link: str | None = None
    reg_course_man: str | None = Field(default=None, description="수강신청 인원 (Work24 regCourseMan)")
    yard_man: str | None = Field(default=None, description="정원 (Work24 yardMan)")
    real_man: str | None = Field(default=None, description="실제 훈련비 (Work24 realMan)")


class CourseListResponse(BaseModel):
    items: list[CourseListItem]
    total_count: int
    page_num: int
    page_size: int


def _normalize_work24_date(value: str, field_name: str) -> str:
    normalized = value.strip().replace("-", "")
    if not _DATE_RE.match(normalized):
        raise HTTPException(
            status_code=400,
            detail=f"{field_name} must be YYYYMMDD or YYYY-MM-DD",
        )
    return normalized


def _parse_course_from_es(source: dict) -> CourseListItem:
    return CourseListItem(
        trainst_cst_id=source.get("trainstCstId"),
        trpr_id=source.get("trprId"),
        trpr_degr=source.get("trprDegr"),
        course_name=source.get("trngCrseNm") or source.get("title"),
        inst_name=source.get("instNm") or source.get("subTitle"),
        tra_start_date=source.get("traStartDate"),
        tra_end_date=source.get("traEndDate"),
        address=source.get("address"),
        tel_no=source.get("telNo"),
        title_link=source.get("titleLink"),
        reg_course_man=source.get("regCourseMan"),
        yard_man=source.get("yardMan"),
        real_man=source.get("realMan"),
    )


def _to_es_date(value: str) -> str:
    """YYYYMMDD → YYYY-MM-DD (ES date range용)."""
    if len(value) == 8 and value.isdigit():
        return f"{value[:4]}-{value[4:6]}-{value[6:8]}"
    return value


def _tra_start_date_range(gte: str, lte: str) -> dict:
    """traStartDate(date) 필드 range — ISO YYYY-MM-DD."""
    return {
        "range": {
            "traStartDate": {
                "gte": _to_es_date(gte),
                "lte": _to_es_date(lte),
            }
        }
    }


def _tra_start_date_year_filter(year: int) -> dict:
    return _tra_start_date_range(f"{year}0101", f"{year}1231")


_COURSE_ID_SORT_FIELDS = ("trainstCstId", "trprId", "trprDegr")


def _course_id_sort(*, include_score: bool = False) -> list[dict]:
    """TRAINST_CST_ID → TRPR_ID → TRPR_DEGR 순 정렬 (.keyword: text 매핑 인덱스 호환)."""
    sort: list[dict] = []
    if include_score:
        sort.append({"_score": "desc"})
    for field in _COURSE_ID_SORT_FIELDS:
        sort.append({f"{field}.keyword": {"order": "asc", "missing": "_last"}})
    return sort


def _reg_course_man_gt_zero_filter() -> dict:
    """Work24 regCourseMan(문자열)이 0보다 큰 문서만 필터."""
    return {
        "script": {
            "script": {
                "source": (
                    "if (doc.containsKey('regCourseMan.keyword') && "
                    "doc['regCourseMan.keyword'].size() > 0) {"
                    "  try { return Double.parseDouble("
                    "doc['regCourseMan.keyword'].value) > 0; }"
                    "  catch (Exception e) { return false; }"
                    "}"
                    "if (doc.containsKey('regCourseMan') && "
                    "doc['regCourseMan'].size() > 0) {"
                    "  try { return Double.parseDouble("
                    "doc['regCourseMan'].value) > 0; }"
                    "  catch (Exception e) { return false; }"
                    "}"
                    "return false;"
                ),
                "lang": "painless",
            }
        }
    }


def _build_list_query(
    st_dt: str,
    end_dt: str,
    organ_nm: str | None,
    process_nm: str | None,
    has_reg_course_man: bool = False,
) -> dict:
    must: list[dict] = [
        _tra_start_date_range(st_dt, end_dt),
    ]
    if organ_nm and organ_nm.strip():
        must.append(
            {"wildcard": {"instNm": {"value": f"*{organ_nm.strip()}*"}}}
        )
    if process_nm and process_nm.strip():
        keyword = process_nm.strip()
        must.append(
            {
                "bool": {
                    "should": [
                        {"match": {"trngCrseNm.nori": {"query": keyword}}},
                        {"wildcard": {"trngCrseNm.raw": {"value": f"*{keyword}*"}}},
                    ],
                    "minimum_should_match": 1,
                }
            }
        )
    if has_reg_course_man:
        must.append(_reg_course_man_gt_zero_filter())
    return {"bool": {"must": must}}


def _build_list_body(
    st_dt: str,
    end_dt: str,
    organ_nm: str | None,
    process_nm: str | None,
    page_num: int,
    page_size: int,
    has_reg_course_man: bool = False,
) -> dict:
    return {
        "query": _build_list_query(
            st_dt, end_dt, organ_nm, process_nm, has_reg_course_man
        ),
        "from": (page_num - 1) * page_size,
        "size": page_size,
        "sort": _course_id_sort(),
        "track_total_hits": True,
    }


def _build_list_scroll_body(
    st_dt: str,
    end_dt: str,
    organ_nm: str | None,
    process_nm: str | None,
    has_reg_course_man: bool = False,
) -> dict:
    return {
        "query": _build_list_query(
            st_dt, end_dt, organ_nm, process_nm, has_reg_course_man
        ),
        "size": SCROLL_BATCH_SIZE,
        "sort": _course_id_sort(),
        "track_total_hits": True,
    }


async def _load_active_owned_names(session: AsyncSession) -> list[str]:
    stmt = select(OwnedCourse.course_name).where(
        OwnedCourse.is_delete == False,  # noqa: E712
        OwnedCourse.is_active == True,  # noqa: E712
    )
    rows = (await session.execute(stmt)).scalars().all()
    return sorted({(n or "").strip() for n in rows if n and n.strip()})


def _escape_es_wildcard(value: str) -> str:
    return _WILDCARD_ESCAPE_RE.sub(r"\\\1", value)


def _build_owned_name_clause(name: str) -> dict:
    """보유과정명과 Work24 과정명을 엄격히 매칭 (부분 토큰·오타 fuzzy 제외)."""
    trimmed = name.strip()
    escaped = _escape_es_wildcard(trimmed)
    return {
        "bool": {
            "minimum_should_match": 1,
            "should": [
                {"term": {"trngCrseNm.raw": {"value": trimmed, "boost": 10}}},
                {
                    "wildcard": {
                        "trngCrseNm.raw": {
                            "value": f"*{escaped}*",
                            "case_insensitive": True,
                            "boost": 8,
                        }
                    }
                },
                {
                    "match_phrase": {
                        "trngCrseNm.nori": {
                            "query": trimmed,
                            "slop": 2,
                            "boost": 5,
                        }
                    }
                },
            ],
        }
    }


def _build_owned_match_query(
    names: list[str],
    year: int,
    min_score: float,
    has_reg_course_man: bool = False,
) -> tuple[dict, float | None]:
    should = [_build_owned_name_clause(name) for name in names]
    filters: list[dict] = [
        _tra_start_date_year_filter(year),
    ]
    if has_reg_course_man:
        filters.append(_reg_course_man_gt_zero_filter())
    query: dict = {
        "bool": {
            "filter": filters,
            "should": should,
            "minimum_should_match": 1,
        }
    }
    min_score_value = min_score if min_score > 0 else None
    return query, min_score_value


def _build_owned_match_body(
    names: list[str],
    year: int,
    page_num: int,
    page_size: int,
    min_score: float,
    has_reg_course_man: bool = False,
) -> dict:
    # TODO: 활성 보유과정이 매우 많으면 bool.max_clause_count 초과 가능 → 배치 분할 검토
    query, min_score_value = _build_owned_match_query(
        names, year, min_score, has_reg_course_man
    )
    body: dict = {
        "query": query,
        "from": (page_num - 1) * page_size,
        "size": page_size,
        "sort": _course_id_sort(include_score=True),
        "track_total_hits": True,
    }
    if min_score_value is not None:
        body["min_score"] = min_score_value
    return body


def _build_owned_scroll_body(
    names: list[str],
    year: int,
    min_score: float,
    has_reg_course_man: bool = False,
) -> dict:
    query, min_score_value = _build_owned_match_query(
        names, year, min_score, has_reg_course_man
    )
    body: dict = {
        "query": query,
        "size": SCROLL_BATCH_SIZE,
        "sort": _course_id_sort(include_score=True),
        "track_total_hits": True,
    }
    if min_score_value is not None:
        body["min_score"] = min_score_value
    return body


def _course_list_from_es_response(
    response: dict, page_num: int, page_size: int
) -> CourseListResponse:
    hits = response.get("hits", {})
    total_count = _extract_total_count(response)
    items = [
        _parse_course_from_es(hit.get("_source", {}))
        for hit in hits.get("hits", [])
    ]
    return CourseListResponse(
        items=items,
        total_count=total_count,
        page_num=page_num,
        page_size=page_size,
    )


def _extract_total_count(response: dict) -> int:
    total_raw = response.get("hits", {}).get("total")
    if isinstance(total_raw, dict):
        return int(total_raw.get("value", 0))
    return int(total_raw or 0)


def _course_item_to_row(item: CourseListItem) -> list[str]:
    return [str(getattr(item, attr) or "") for _, attr in EXPORT_HEADERS]


def _assert_export_within_limit(total_count: int) -> None:
    if total_count > MAX_EXPORT_ROWS:
        raise HTTPException(status_code=400, detail=EXPORT_OVER_LIMIT_MESSAGE)


async def _clear_scroll_safe(es, scroll_id: str | None) -> None:
    if not scroll_id:
        return
    try:
        await es.clear_scroll(scroll_id=scroll_id)
    except Exception:
        logger.exception("Elasticsearch clear_scroll failed")


async def _write_courses_xlsx(es, body: dict, path: str) -> int:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("과정목록")
    ws.append([label for label, _ in EXPORT_HEADERS])

    scroll_id: str | None = None
    written = 0
    try:
        response = await es.search(
            index=settings.ES_COURSE_INDEX,
            body=body,
            scroll=SCROLL_KEEPALIVE,
        )
        scroll_id = response.get("_scroll_id")
        _assert_export_within_limit(_extract_total_count(response))

        while True:
            hits = response.get("hits", {}).get("hits", [])
            if not hits:
                break
            for hit in hits:
                if written >= MAX_EXPORT_ROWS:
                    break
                item = _parse_course_from_es(hit.get("_source", {}))
                ws.append(_course_item_to_row(item))
                written += 1
            if written >= MAX_EXPORT_ROWS or len(hits) < SCROLL_BATCH_SIZE:
                break
            response = await es.scroll(scroll_id=scroll_id, scroll=SCROLL_KEEPALIVE)
    finally:
        await _clear_scroll_safe(es, scroll_id)

    await asyncio.to_thread(wb.save, path)
    return written


async def _write_empty_courses_xlsx(path: str) -> int:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("과정목록")
    ws.append([label for label, _ in EXPORT_HEADERS])
    await asyncio.to_thread(wb.save, path)
    return 0


async def _stream_file_chunks(path: str, *, unlink_after: bool = True) -> AsyncIterator[bytes]:
    try:
        with open(path, "rb") as file_obj:
            while True:
                chunk = await asyncio.to_thread(file_obj.read, 65_536)
                if not chunk:
                    break
                yield chunk
    finally:
        if unlink_after:
            try:
                await asyncio.to_thread(os.unlink, path)
            except OSError:
                logger.exception("Failed to remove temporary export file %s", path)


async def _export_courses_response(
    es, body: dict | None, params: dict | None = None
) -> StreamingResponse:
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        if body is None:
            await _write_empty_courses_xlsx(path)
        else:
            try:
                await _write_courses_xlsx(es, body, path)
            except NotFoundError:
                await _write_empty_courses_xlsx(path)
    except Exception:
        try:
            await asyncio.to_thread(os.unlink, path)
        except OSError:
            pass
        raise

    filename = _export_download_filename(params)
    return StreamingResponse(
        _stream_file_chunks(path),
        media_type=EXPORT_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _normalize_export_params(
    *,
    srch_tra_st_dt: str | None,
    srch_tra_end_dt: str | None,
    srch_tra_organ_nm: str | None,
    srch_tra_process_nm: str | None,
    has_reg_course_man: bool,
    owned_year: int | None,
    min_score: float,
) -> dict:
    """요청 파라미터를 검증하고 course_export_job.params 저장용 dict로 정규화한다."""
    if owned_year is not None:
        return {
            "owned_year": int(owned_year),
            "min_score": float(min_score or 0),
            "has_reg_course_man": bool(has_reg_course_man),
        }
    if not srch_tra_st_dt or not srch_tra_end_dt:
        raise HTTPException(
            status_code=400,
            detail="srch_tra_st_dt and srch_tra_end_dt are required",
        )
    st_dt = _normalize_work24_date(srch_tra_st_dt, "srch_tra_st_dt")
    end_dt = _normalize_work24_date(srch_tra_end_dt, "srch_tra_end_dt")
    params: dict = {
        "srch_tra_st_dt": st_dt,
        "srch_tra_end_dt": end_dt,
        "has_reg_course_man": bool(has_reg_course_man),
    }
    organ_nm = srch_tra_organ_nm.strip() if srch_tra_organ_nm else None
    process_nm = srch_tra_process_nm.strip() if srch_tra_process_nm else None
    if organ_nm:
        params["srch_tra_organ_nm"] = organ_nm
    if process_nm:
        params["srch_tra_process_nm"] = process_nm
    return params


def _build_conditions_summary(params: dict) -> str:
    """저장된 params를 사람이 읽을 수 있는 검색조건 요약으로 변환한다."""
    parts: list[str] = []
    if params.get("owned_year") is not None:
        parts.append(f"보유과정 {params['owned_year']}년")
        min_score = params.get("min_score")
        if min_score:
            parts.append(f"관련도≥{min_score}")
    else:
        st = _to_es_date(str(params.get("srch_tra_st_dt", "")))
        end = _to_es_date(str(params.get("srch_tra_end_dt", "")))
        parts.append(f"훈련시작일 {st} ~ {end}")
        if params.get("srch_tra_organ_nm"):
            parts.append(f"기관명 '{params['srch_tra_organ_nm']}'")
        if params.get("srch_tra_process_nm"):
            parts.append(f"과정명 '{params['srch_tra_process_nm']}'")
    if params.get("has_reg_course_man"):
        parts.append("수강신청 인원 있음")
    return ", ".join(parts)


def _export_year_label(params: dict | None) -> str | None:
    """다운로드 파일명에 넣을 과정 년도 라벨 (예: '2025', '2024-2025')."""
    if not params:
        return None
    if params.get("owned_year") is not None:
        return str(int(params["owned_year"]))
    st = str(params.get("srch_tra_st_dt") or "").strip()
    end = str(params.get("srch_tra_end_dt") or "").strip()
    st_es = _to_es_date(st) if st else ""
    end_es = _to_es_date(end) if end else ""
    if len(st_es) >= 4 and st_es[:4].isdigit():
        start_year = st_es[:4]
        if len(end_es) >= 4 and end_es[:4].isdigit() and end_es[:4] != start_year:
            return f"{start_year}-{end_es[:4]}"
        return start_year
    return None


def _export_download_filename(
    params: dict | None = None,
    *,
    when: datetime | None = None,
) -> str:
    """다운로드용 파일명. 예: courses_2025_20260715_113000.xlsx"""
    ts = (when or datetime.now()).strftime("%Y%m%d_%H%M%S")
    year = _export_year_label(params)
    if year:
        return f"courses_{year}_{ts}.xlsx"
    return f"courses_{ts}.xlsx"


def _export_body_from_params(params: dict, names: list[str] | None) -> dict | None:
    """저장된 params로 스크롤 ES body를 재구성한다. owned 모드는 names가 필요하다."""
    if params.get("owned_year") is not None:
        if not names:
            return None
        return _build_owned_scroll_body(
            names,
            int(params["owned_year"]),
            float(params.get("min_score") or 0),
            bool(params.get("has_reg_course_man")),
        )
    return _build_list_scroll_body(
        str(params["srch_tra_st_dt"]),
        str(params["srch_tra_end_dt"]),
        params.get("srch_tra_organ_nm"),
        params.get("srch_tra_process_nm"),
        bool(params.get("has_reg_course_man")),
    )


def _build_search_body(keyword: str, size: int) -> dict:
    return {
        "query": {
            "bool": {
                "should": [
                    {"term": {"trngCrseNm.raw": {"value": keyword, "boost": 5}}},
                    {"match": {"trngCrseNm.nori": {"query": keyword, "boost": 3}}},
                    {"match": {"trngCrseNm.ngram": {"query": keyword, "boost": 1}}},
                    {
                        "fuzzy": {
                            "trngCrseNm.raw": {
                                "value": keyword,
                                "fuzziness": "AUTO",
                                "boost": 2,
                            }
                        }
                    },
                ],
                "minimum_should_match": 1,
            }
        },
        "size": size,
    }


def _parse_hits(response: dict) -> list[CourseSearchHit]:
    hits: list[CourseSearchHit] = []
    for hit in response.get("hits", {}).get("hits", []):
        source = hit.get("_source", {})
        hits.append(
            CourseSearchHit(
                trpr_id=source.get("trprId"),
                trng_crse_nm=source.get("trngCrseNm"),
                inst_nm=source.get("instNm"),
                score=float(hit.get("_score", 0.0)),
            )
        )
    return hits


@router.get("", response_model=CourseListResponse)
async def list_courses(
    srch_tra_st_dt: str = Query(..., description="훈련시작일 From (YYYYMMDD 또는 YYYY-MM-DD)"),
    srch_tra_end_dt: str = Query(..., description="훈련시작일 To (YYYYMMDD 또는 YYYY-MM-DD)"),
    srch_tra_organ_nm: str | None = Query(None, description="훈련기관명"),
    srch_tra_process_nm: str | None = Query(None, description="훈련과정명"),
    has_reg_course_man: bool = Query(
        False, description="수강신청 인원 있음 (regCourseMan > 0)"
    ),
    page_num: int = Query(1, ge=1, le=1000),
    page_size: int = Query(20, ge=1, le=100),
    _user=Depends(current_active_user),
) -> CourseListResponse:
    st_dt = _normalize_work24_date(srch_tra_st_dt, "srch_tra_st_dt")
    end_dt = _normalize_work24_date(srch_tra_end_dt, "srch_tra_end_dt")
    organ_nm = srch_tra_organ_nm.strip() if srch_tra_organ_nm else None
    process_nm = srch_tra_process_nm.strip() if srch_tra_process_nm else None

    es = get_es()
    body = _build_list_body(
        st_dt, end_dt, organ_nm, process_nm, page_num, page_size, has_reg_course_man
    )
    try:
        response = await es.search(index=settings.ES_COURSE_INDEX, body=body)
    except NotFoundError:
        return CourseListResponse(
            items=[],
            total_count=0,
            page_num=page_num,
            page_size=page_size,
        )
    except Exception:
        logger.exception("Elasticsearch list request failed")
        raise HTTPException(status_code=502, detail="Course search failed") from None

    return _course_list_from_es_response(response, page_num, page_size)


@router.get("/owned-search", response_model=CourseListResponse)
async def search_owned_courses(
    year: int = Query(..., ge=2023, le=2100, description="훈련시작일 기준 조회 년도"),
    min_score: float = Query(0, ge=0, description="관련도 임계치 (0이면 미적용)"),
    has_reg_course_man: bool = Query(
        False, description="수강신청 인원 있음 (regCourseMan > 0)"
    ),
    page_num: int = Query(1, ge=1, le=1000),
    page_size: int = Query(20, ge=1, le=100),
    session: AsyncSession = Depends(get_async_session),
    _user=Depends(current_active_user),
) -> CourseListResponse:
    names = await _load_active_owned_names(session)
    if not names:
        return CourseListResponse(
            items=[],
            total_count=0,
            page_num=page_num,
            page_size=page_size,
        )

    es = get_es()
    body = _build_owned_match_body(
        names, year, page_num, page_size, min_score, has_reg_course_man
    )
    logger.info(
        "owned-search year=%s names=%d page=%d size=%d min_score=%s",
        year,
        len(names),
        page_num,
        page_size,
        min_score,
    )
    try:
        response = await es.search(index=settings.ES_COURSE_INDEX, body=body)
    except NotFoundError:
        return CourseListResponse(
            items=[],
            total_count=0,
            page_num=page_num,
            page_size=page_size,
        )
    except Exception:
        logger.exception(
            "Elasticsearch owned-search failed for year=%s names=%d",
            year,
            len(names),
        )
        raise HTTPException(status_code=502, detail="Owned course search failed") from None

    return _course_list_from_es_response(response, page_num, page_size)


@router.get("/export")
async def export_courses(
    srch_tra_st_dt: str | None = Query(
        None, description="훈련시작일 From (YYYYMMDD 또는 YYYY-MM-DD)"
    ),
    srch_tra_end_dt: str | None = Query(
        None, description="훈련시작일 To (YYYYMMDD 또는 YYYY-MM-DD)"
    ),
    srch_tra_organ_nm: str | None = Query(None, description="훈련기관명"),
    srch_tra_process_nm: str | None = Query(None, description="훈련과정명"),
    has_reg_course_man: bool = Query(
        False, description="수강신청 인원 있음 (regCourseMan > 0)"
    ),
    owned_year: int | None = Query(
        None, ge=2023, le=2100, description="보유 과정 조회 년도"
    ),
    min_score: float = Query(0, ge=0, description="관련도 임계치 (보유 과정 조회)"),
    session: AsyncSession = Depends(get_async_session),
    _user=Depends(current_active_user),
) -> StreamingResponse:
    es = get_es()
    params = _normalize_export_params(
        srch_tra_st_dt=srch_tra_st_dt,
        srch_tra_end_dt=srch_tra_end_dt,
        srch_tra_organ_nm=srch_tra_organ_nm,
        srch_tra_process_nm=srch_tra_process_nm,
        has_reg_course_man=has_reg_course_man,
        owned_year=owned_year,
        min_score=min_score,
    )
    is_owned_search = params.get("owned_year") is not None

    if is_owned_search:
        names = await _load_active_owned_names(session)
        body = (
            _build_owned_scroll_body(
                names,
                int(params["owned_year"]),
                float(params.get("min_score") or 0),
                bool(params.get("has_reg_course_man")),
            )
            if names
            else None
        )
    else:
        body = _build_list_scroll_body(
            str(params["srch_tra_st_dt"]),
            str(params["srch_tra_end_dt"]),
            params.get("srch_tra_organ_nm"),
            params.get("srch_tra_process_nm"),
            bool(params.get("has_reg_course_man")),
        )

    try:
        return await _export_courses_response(es, body, params)
    except HTTPException:
        raise
    except Exception:
        logger.exception("Course export failed")
        raise HTTPException(status_code=502, detail="Course export failed") from None


@router.post("/export-jobs", response_model=CourseExportJobRead, status_code=201)
async def create_course_export_job(
    body: CourseExportJobCreate,
    session: AsyncSession = Depends(get_async_session),
    user: User = Depends(current_active_user),
) -> CourseExportJobRead:
    params = _normalize_export_params(
        srch_tra_st_dt=body.srch_tra_st_dt,
        srch_tra_end_dt=body.srch_tra_end_dt,
        srch_tra_organ_nm=body.srch_tra_organ_nm,
        srch_tra_process_nm=body.srch_tra_process_nm,
        has_reg_course_man=body.has_reg_course_man,
        owned_year=body.owned_year,
        min_score=body.min_score,
    )

    job_def = await session.get(SchedulerJob, "course_export")
    if job_def is None or job_def.is_delete:
        raise HTTPException(status_code=404, detail="course_export job not found")

    memo = body.memo.strip() if body.memo else None
    export = CourseExportJob(
        status=QUEUE_STATUS_PENDING,
        memo=memo,
        conditions_summary=_build_conditions_summary(params),
        params=params,
        requested_by_user_id=user.id,
    )
    session.add(export)
    await session.commit()
    await session.refresh(export)

    q = SchedulerJobQueue(
        job_key="course_export",
        action=QUEUE_ACTION_RUN_NOW,
        status=QUEUE_STATUS_PENDING,
        requested_by_user_id=user.id,
        payload={"export_id": export.id},
    )
    session.add(q)
    await session.commit()
    await session.refresh(q)

    export.queue_id = q.id
    await session.commit()
    await session.refresh(export)
    return export


@router.get("/export-jobs", response_model=Page[CourseExportJobRead])
async def list_course_export_jobs(
    page: int = 1,
    size: int = 20,
    session: AsyncSession = Depends(get_async_session),
    user: User = Depends(current_active_user),
):
    if size < 1:
        raise HTTPException(status_code=400, detail="size must be >= 1")
    if size > MAX_PAGE_SIZE:
        raise HTTPException(status_code=400, detail=f"size must be <= {MAX_PAGE_SIZE}")
    params = Params(page=page, size=size)
    stmt = select(CourseExportJob).where(CourseExportJob.is_delete == False)  # noqa: E712
    if not user.is_superuser:
        stmt = stmt.where(CourseExportJob.requested_by_user_id == user.id)
    stmt = stmt.order_by(CourseExportJob.id.desc())
    return await apaginate(session, stmt, params)


@router.get("/export-jobs/{export_id}/download")
async def download_course_export_job(
    export_id: int,
    session: AsyncSession = Depends(get_async_session),
    user: User = Depends(current_active_user),
) -> StreamingResponse:
    job = await session.get(CourseExportJob, export_id)
    if job is None or job.is_delete:
        raise HTTPException(status_code=404, detail="Export job not found")
    if not user.is_superuser and job.requested_by_user_id != user.id:
        raise HTTPException(status_code=403, detail="Forbidden")
    if job.status != "SUCCEEDED" or not job.file_path:
        raise HTTPException(status_code=404, detail="Export file not ready")
    if not os.path.exists(job.file_path):
        raise HTTPException(status_code=404, detail="Export file missing")

    filename = job.file_name or f"courses_{export_id}.xlsx"
    return StreamingResponse(
        _stream_file_chunks(job.file_path, unlink_after=False),
        media_type=EXPORT_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/search", response_model=list[CourseSearchHit])
async def search_courses(
    keyword: str = Query(..., min_length=1),
    size: int = Query(5, ge=1, le=100),
    _user=Depends(current_active_user),
) -> list[CourseSearchHit]:
    es = get_es()
    body = _build_search_body(keyword, size)
    try:
        response = await es.search(index=settings.ES_COURSE_INDEX, body=body)
        return _parse_hits(response)
    except NotFoundError:
        return []
    except Exception:
        logger.exception("Elasticsearch search failed for keyword=%r", keyword)
        return []


@router.post("/index", response_model=CourseIndexResult)
async def index_courses_from_work24(
    _user=Depends(current_superuser),
) -> CourseIndexResult:
    es = get_es()
    items, total = await fetch_courses(
        source="courses_index",
        **INDEX_TEST_PARAMS,
    )
    await ensure_course_index(es)
    indexed = await index_courses(es, items)
    return CourseIndexResult(
        fetched=len(items),
        indexed=indexed,
        total_count=total,
    )


@router.post("/legacy-index", response_model=LegacyIndexResponse)
async def enqueue_legacy_course_index(
    body: LegacyIndexRequest,
    session: AsyncSession = Depends(get_async_session),
    user: User = Depends(current_superuser),
) -> LegacyIndexResponse:
    try:
        parse_year_month(body.start_month, "start_month")
        parse_year_month(body.end_month, "end_month")
        month_ranges = iter_month_ranges(body.start_month, body.end_month)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e

    job = await session.get(SchedulerJob, "legacy_course_index")
    if job is None or job.is_delete:
        raise HTTPException(status_code=404, detail="legacy_course_index job not found")

    queues: list[SchedulerJobQueue] = []
    for st, _en in month_ranges:
        month_label = f"{st[:4]}-{st[4:6]}"
        q = SchedulerJobQueue(
            job_key="legacy_course_index",
            action=QUEUE_ACTION_RUN_NOW,
            status=QUEUE_STATUS_PENDING,
            requested_by_user_id=user.id,
            payload={"month": month_label},
        )
        session.add(q)
        queues.append(q)

    await session.commit()
    for q in queues:
        await session.refresh(q)
        q.payload = {**q.payload, "queue_id": q.id}
    await session.commit()

    month_count = len(queues)
    return LegacyIndexResponse(
        queue_ids=[q.id for q in queues],
        start_month=body.start_month,
        end_month=body.end_month,
        month_count=month_count,
        message=f"과거 과정 색인 작업 {month_count}건(월 단위)이 큐에 등록되었습니다.",
    )
