from __future__ import annotations

import io
import logging
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from fastapi_pagination.ext.sqlalchemy import apaginate
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, ConfigDict, Field
from sqlalchemy import delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_async_session
from app.models import SeparateSettlement, User
from app.pagination import MAX_PAGE_SIZE, Page, Params
from app.routes.courses import MAX_EXPORT_ROWS
from app.routes.settlements import (
    _normalize_compare_date,
    _parse_decimal,
    _parse_rate,
    _parse_str,
)
from app.users import current_active_user

logger = logging.getLogger(__name__)

router = APIRouter()

EXCEL_HEADERS: list[tuple[str, str]] = [
    ("계산서(수취)마감일", "invoice_deadline_date"),
    ("영업대표", "sales_rep"),
    ("구분", "category"),
    ("고객사", "client_name"),
    ("사업 내역", "business_detail"),
    ("과정명", "course_name"),
    ("기준매출액", "base_revenue"),
    ("정산율", "settlement_rate"),
    ("산출정산액", "calculated_amount"),
    ("계약기간", "contract_period"),
    ("차감액", "deduction_amount"),
    ("최종정산액(차감 후)", "final_amount"),
    ("발행 항목", "invoice_item"),
    ("공급가액", "supply_amount"),
    ("세액", "tax_amount"),
    ("총액", "total_amount"),
    ("계산서발행처", "invoice_issuer"),
]

HEADER_TO_FIELD = {label: field for label, field in EXCEL_HEADERS}
HEADER_TO_FIELD_NORM = {
    label.replace(" ", ""): field for label, field in EXCEL_HEADERS
}


class SeparateSettlementListItem(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: int
    invoice_deadline_date: date | None
    invoice_deadline_year: int | None
    sales_rep: str | None
    category: str | None
    client_name: str
    business_detail: str | None
    course_name: str
    base_revenue: Decimal | None
    settlement_rate: Decimal | None
    settlement_rate_raw: str | None
    calculated_amount: Decimal | None
    contract_period: str | None
    deduction_amount: Decimal | None
    final_amount: Decimal | None
    invoice_item: str | None
    supply_amount: Decimal | None
    tax_amount: Decimal | None
    total_amount: Decimal | None
    invoice_issuer: str | None


class SeparateSettlementImportError(BaseModel):
    row: int
    message: str


class SeparateSettlementImportResult(BaseModel):
    deleted: int
    created: int
    failed: int
    errors: list[SeparateSettlementImportError] = Field(default_factory=list)


def _transform_list(
    items: list[SeparateSettlement],
) -> list[SeparateSettlementListItem]:
    return [SeparateSettlementListItem.model_validate(row) for row in items]


def _resolve_settlement_rate(
    raw_value: Any,
    base_revenue: Decimal | None,
    calculated_amount: Decimal | None,
) -> tuple[Decimal | None, str | None]:
    """정산율: 숫자면 파싱, 비숫자면 산출정산액/기준매출액."""
    raw_str = _parse_str(raw_value)
    if raw_str is None and raw_value is not None and raw_value != "":
        raw_str = str(raw_value).strip() or None

    rate: Decimal | None = None
    if raw_value is not None and raw_value != "":
        try:
            rate = _parse_rate(raw_value)
        except (InvalidOperation, ValueError):
            rate = None

    if rate is None:
        if (
            base_revenue is not None
            and base_revenue != 0
            and calculated_amount is not None
        ):
            rate = calculated_amount / base_revenue

    return rate, raw_str


def _row_to_fields(row_values: dict[str, Any]) -> dict[str, Any]:
    client_name = _parse_str(row_values.get("client_name"))
    if not client_name:
        raise ValueError("고객사가 비어 있습니다.")

    course_name = _parse_str(row_values.get("course_name"))
    if not course_name:
        raise ValueError("과정명이 비어 있습니다.")

    try:
        invoice_deadline_date = _normalize_compare_date(
            row_values.get("invoice_deadline_date")
        )
        base_revenue = _parse_decimal(row_values.get("base_revenue"))
        calculated_amount = _parse_decimal(row_values.get("calculated_amount"))
        settlement_rate, settlement_rate_raw = _resolve_settlement_rate(
            row_values.get("settlement_rate"),
            base_revenue,
            calculated_amount,
        )
        fields: dict[str, Any] = {
            "invoice_deadline_date": invoice_deadline_date,
            "invoice_deadline_year": (
                invoice_deadline_date.year if invoice_deadline_date else None
            ),
            "sales_rep": _parse_str(row_values.get("sales_rep")),
            "category": _parse_str(row_values.get("category")),
            "client_name": client_name,
            "business_detail": _parse_str(row_values.get("business_detail")),
            "course_name": course_name,
            "base_revenue": base_revenue,
            "settlement_rate": settlement_rate,
            "settlement_rate_raw": settlement_rate_raw,
            "calculated_amount": calculated_amount,
            "contract_period": _parse_str(row_values.get("contract_period")),
            "deduction_amount": _parse_decimal(row_values.get("deduction_amount")),
            "final_amount": _parse_decimal(row_values.get("final_amount")),
            "invoice_item": _parse_str(row_values.get("invoice_item")),
            "supply_amount": _parse_decimal(row_values.get("supply_amount")),
            "tax_amount": _parse_decimal(row_values.get("tax_amount")),
            "total_amount": _parse_decimal(row_values.get("total_amount")),
            "invoice_issuer": _parse_str(row_values.get("invoice_issuer")),
        }
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"숫자 파싱 실패: {exc}") from exc
    return fields


def _excel_cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, date) and not isinstance(value, datetime):
        return value.isoformat()
    return value


@router.get("", response_model=Page[SeparateSettlementListItem])
async def list_separate_settlements(
    page: int = 1,
    size: int = 20,
    year: int | None = Query(default=None, description="계산서(수취)마감일 연도"),
    client_name: str | None = Query(default=None, description="고객사 검색"),
    course_name: str | None = Query(default=None, description="과정명 검색"),
    session: AsyncSession = Depends(get_async_session),
    _: User = Depends(current_active_user),
):
    if size < 1:
        raise HTTPException(status_code=400, detail="size must be >= 1")
    if size > MAX_PAGE_SIZE:
        raise HTTPException(status_code=400, detail=f"size must be <= {MAX_PAGE_SIZE}")

    params = Params(page=page, size=size)
    stmt = select(SeparateSettlement).where(
        SeparateSettlement.is_delete == False  # noqa: E712
    )
    if year is not None:
        stmt = stmt.where(SeparateSettlement.invoice_deadline_year == year)
    if client_name and client_name.strip():
        stmt = stmt.where(
            SeparateSettlement.client_name.ilike(f"%{client_name.strip()}%")
        )
    if course_name and course_name.strip():
        stmt = stmt.where(
            SeparateSettlement.course_name.ilike(f"%{course_name.strip()}%")
        )
    stmt = stmt.order_by(
        SeparateSettlement.invoice_deadline_date.desc().nulls_last(),
        SeparateSettlement.id.desc(),
    )
    return await apaginate(session, stmt, params, transformer=_transform_list)


@router.get("/import/template")
async def download_import_template(
    _: User = Depends(current_active_user),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "별도정산"
    ws.append([label for label, _ in EXCEL_HEADERS])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                'attachment; filename="separate_settlements_template.xlsx"'
            )
        },
    )


@router.get("/export")
async def export_separate_settlements(
    session: AsyncSession = Depends(get_async_session),
    _: User = Depends(current_active_user),
):
    total = await session.scalar(
        select(func.count())
        .select_from(SeparateSettlement)
        .where(SeparateSettlement.is_delete == False)  # noqa: E712
    )
    total = int(total or 0)
    if total > MAX_EXPORT_ROWS:
        raise HTTPException(
            status_code=400,
            detail=f"내보낼 데이터가 {MAX_EXPORT_ROWS:,}건을 초과합니다.",
        )

    rows = (
        await session.execute(
            select(SeparateSettlement)
            .where(SeparateSettlement.is_delete == False)  # noqa: E712
            .order_by(
                SeparateSettlement.invoice_deadline_date.desc().nulls_last(),
                SeparateSettlement.id.desc(),
            )
        )
    ).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "별도정산"
    ws.append([label for label, _ in EXCEL_HEADERS])
    for row in rows:
        cells: list[Any] = []
        for label, field in EXCEL_HEADERS:
            if field == "settlement_rate":
                cells.append(_excel_cell_value(row.settlement_rate_raw))
            else:
                cells.append(_excel_cell_value(getattr(row, field)))
        ws.append(cells)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                'attachment; filename="separate_settlements.xlsx"'
            )
        },
    )


@router.post("/import", response_model=SeparateSettlementImportResult)
async def import_separate_settlements(
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_async_session),
    _: User = Depends(current_active_user),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="xlsx 파일만 업로드할 수 있습니다.")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=400, detail=f"엑셀 파일을 읽을 수 없습니다: {exc}"
        ) from exc

    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        wb.close()
        raise HTTPException(status_code=400, detail="헤더 행이 없습니다.")

    col_map: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        label = str(cell).strip()
        field = HEADER_TO_FIELD.get(label) or HEADER_TO_FIELD_NORM.get(
            label.replace(" ", "")
        )
        if field:
            col_map[idx] = field

    required = {"client_name", "course_name"}
    if not required.issubset(set(col_map.values())):
        wb.close()
        raise HTTPException(
            status_code=400, detail="고객사, 과정명 열이 필요합니다."
        )

    delete_result = await session.execute(delete(SeparateSettlement))
    deleted = delete_result.rowcount or 0

    created = 0
    failed = 0
    errors: list[SeparateSettlementImportError] = []
    to_insert: list[SeparateSettlement] = []

    for row_num, row in enumerate(rows, start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue

        row_values: dict[str, Any] = {}
        for idx, field in col_map.items():
            if idx < len(row):
                row_values[field] = row[idx]

        try:
            fields = _row_to_fields(row_values)
            to_insert.append(SeparateSettlement(**fields))
            created += 1
        except Exception as exc:
            failed += 1
            errors.append(
                SeparateSettlementImportError(
                    row=row_num, message=str(exc) or "처리 실패"
                )
            )

    if to_insert:
        session.add_all(to_insert)
    await session.commit()
    wb.close()
    return SeparateSettlementImportResult(
        deleted=deleted, created=created, failed=failed, errors=errors
    )
