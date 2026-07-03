from io import BytesIO
from datetime import datetime, timezone, date

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Form
from fastapi_pagination import Page, Params
from fastapi_pagination.ext.sqlalchemy import apaginate
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import func, or_
from sqlalchemy.orm import selectinload
from sqlalchemy.exc import SQLAlchemyError
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from starlette.responses import StreamingResponse
from urllib.parse import quote
from uuid import UUID
from pydantic import BaseModel, Field

from app.database import get_async_session
from app.models import (
    Shipment,
    ShipmentItem,
    Order,
    OrderStatus,
    OrderHistory,
    OrderHistoryActionType,
    Settlement,
    SettlementHistory,
    SettlementState,
    SettlementHistoryActionType,
    Channel,
    Receiver,
    Product,
    Stock,
    StockHistory,
    StockHistoryActionType,
)
from app.schemas import ShipmentListRead
from app.users import current_active_user
from app.utils import datetime_as_utc_aware
from app.permissions import require_permission


def _stock_snapshot(stock: Stock) -> dict:
    """수정 전 재고 스냅샷(StockHistory.before_update 저장용)."""
    return {
        "product_id": str(stock.product_id) if stock.product_id is not None else None,
        "logistics_location_id": str(getattr(stock, "logistics_location_id", None))
        if getattr(stock, "logistics_location_id", None) is not None
        else None,
        "quantity": stock.quantity,
        "batch_code": stock.batch_code,
        "stock_date": stock.stock_date.isoformat() if getattr(stock, "stock_date", None) else None,
        "expiration_date": stock.expiration_date.isoformat()
        if getattr(stock, "expiration_date", None)
        else None,
        "condition": stock.condition.value if hasattr(stock, "condition") and hasattr(stock.condition, "value") else stock.condition,
        "memo": stock.memo,
        "product_barcode": stock.product_barcode,
    }


def _build_outbound_history(*, stock: Stock, action_quantity: int, user_id, before_update: dict) -> StockHistory:
    # NOTE: StockHistory.quantity는 (이미 차감 반영된) 현재 stock.quantity를 넣는다.
    return StockHistory(
        stock_id=stock.id,
        product_id=stock.product_id,
        quantity=stock.quantity,
        batch_code=stock.batch_code,
        stock_date=datetime_as_utc_aware(stock.stock_date),
        expiration_date=datetime_as_utc_aware(stock.expiration_date),
        action_type=StockHistoryActionType.OUTBOUND,
        action_quantity=action_quantity,
        update_user_id=user_id,
        before_update=before_update,
        reason="송장 업로드",
    )


router = APIRouter(tags=["shipment"])


def _ser(v: object):
    if v is None:
        return None
    if hasattr(v, "hex"):
        return str(v)
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if hasattr(v, "value"):
        try:
            return getattr(v, "value")
        except Exception:
            return str(v)
    return v


def _order_snapshot(order: Order) -> dict:
    return {
        "id": _ser(order.id),
        "receiver_id": _ser(order.receiver_id),
        "channel_id": _ser(order.channel_id),
        "mall_product_name": _ser(getattr(order, "mall_product_name", None)),
        "price": float(order.price) if getattr(order, "price", None) is not None else None,
        "quantity": _ser(getattr(order, "quantity", None)),
        "memo": _ser(getattr(order, "memo", None)),
        "order_date": _ser(getattr(order, "order_date", None)),
        "invoice_number": _ser(getattr(order, "invoice_number", None)),
        "status": _ser(getattr(order, "status", None)),
        "order_placed_date": _ser(getattr(order, "order_placed_date", None)),
        "shipping_date": _ser(getattr(order, "shipping_date", None)),
    }


async def _try_add_order_history(db: AsyncSession, h: OrderHistory) -> None:
    from sqlalchemy.exc import SQLAlchemyError

    try:
        async with db.begin_nested():
            db.add(h)
            await db.flush()
    except SQLAlchemyError:
        return


async def _try_add_settlement_history(db: AsyncSession, h: SettlementHistory) -> None:
    """
    정산 이력은 감사 로그지만, 정산/배송 핵심 플로우를 깨지 않도록
    savepoint로 insert를 격리한다.
    """
    try:
        async with db.begin_nested():
            db.add(h)
            await db.flush()
    except SQLAlchemyError:
        return

class ShipmentSelectedExcelRequest(BaseModel):
    shipment_ids: list[UUID] = Field(min_length=1, max_length=500)


def _to_shipment_list_read(s: Shipment) -> ShipmentListRead:
    order = s.order
    if order is None:
        # shipments.order_id is NOT NULL, so this should never happen.
        raise ValueError("Shipment.order is None")
    receiver = None
    if order is not None and order.receiver is not None:
        r = order.receiver
        receiver = {
            "id": r.id,
            "name": r.name,
            "phone": r.phone,
            "zip_code": r.zip_code,
            "address": r.address,
            "address_detail": r.address_detail,
        }

    items = []
    total_quantity = 0
    for si in (s.shipment_items or []):
        p = si.product
        if p is None:
            continue
        total_quantity += si.quantity
        items.append(
            {
                "product": {
                    "id": p.id,
                    "product_code": p.product_code,
                    "name": p.name,
                },
                "quantity": si.quantity,
            }
        )

    return ShipmentListRead(
        id=s.id,
        order_id=s.order_id,
        invoice_number=s.invoice_number,
        receiver=receiver,
        items=items,
        total_quantity=total_quantity,
        order_placed_date=s.order_placed_date,
        shipping_date=s.shipping_date,
        order_date=order.order_date,
        order_status=order.status,
        channel=(
            {
                "id": order.channel.id,
                "name": order.channel.name,
                "courier_name": getattr(getattr(order.channel, "courier", None), "name", None),
                "courier_url": getattr(getattr(order.channel, "courier", None), "url", None),
                "url": getattr(order.channel, "url", None),
            }
            if order.channel is not None
            else None
        ),
        memo=order.memo,
    )


def _transform_shipments(rows):
    return [_to_shipment_list_read(s) for s in rows]


def _append_excel_header(ws, headers: list[str]) -> None:
    header_fill = PatternFill(patternType="solid", fgColor="FFD9D9D9")  # light gray
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cells: list[WriteOnlyCell] = []
    for h in headers:
        c = WriteOnlyCell(ws, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        cells.append(c)
    ws.append(cells)


PLACE_ORDER_EXCEL_HEADERS = [
    "받는분성명",
    "받는분전화번호",
    "받는분주소(전체, 분할)",
    "받는분상세주소(분할)",
    "고객주문번호",
    "발주용 상품 수량",
    "박스타입",
    "박스수량",
    "배송메세지1",
]


def _shipment_to_place_order_row(sh: Shipment) -> list[object]:
    """
    발주서 엑셀(/order-placed/excel)과 동일한 포맷으로 한 줄을 만든다.
    NOTE: customer_order_no(E)는 송장 업로드 매칭 키로 shipments.id를 사용한다.
    """
    o = sh.order
    if o is None:
        return []
    r = o.receiver
    if r is None:
        return []

    raw = o.raw or {}
    customer_order_no = str(sh.id)
    box_type = "극소"
    box_qty = 1
    msg1 = raw.get("배송메세지1") or raw.get("배송메세지") or ""

    parts: list[str] = []
    for si in sh.shipment_items or []:
        p = si.product
        label = p.name if p is not None else str(si.product_id)
        parts.append(f"{label} x{si.quantity}")
    products_cell = ", ".join(parts)

    return [
        r.name,
        r.phone,
        f"{r.address} {r.address_detail or ''}".strip(),
        r.address_detail or "",
        customer_order_no,
        products_cell,
        box_type,
        box_qty,
        msg1,
    ]


@router.get("", response_model=Page[ShipmentListRead])
async def list_shipments(
    params: Params = Depends(),
    order_id: UUID | None = Query(default=None),
    invoice_number: str | None = Query(default=None),
    order_status: OrderStatus | None = Query(default=None),
    channel_id: UUID | None = Query(default=None),
    channel_ids: str | None = Query(default=None, description="Comma-separated channel ids"),
    channel_name: str | None = Query(default=None),
    receiver_name: str | None = Query(default=None),
    receiver_phone: str | None = Query(default=None),
    receiver_zip_code: str | None = Query(default=None),
    receiver_address: str | None = Query(default=None),
    product_query: str | None = Query(default=None),
    order_date_start: date | None = Query(default=None),
    order_date_end: date | None = Query(default=None),
    order_placed_date_start: date | None = Query(default=None),
    order_placed_date_end: date | None = Query(default=None),
    shipping_date_start: date | None = Query(default=None),
    shipping_date_end: date | None = Query(default=None),
    db: AsyncSession = Depends(get_async_session),
    user=Depends(current_active_user),
):
    _ = await require_permission("shipments", "read")(user, db)
    q = (
        select(Shipment)
        .join(Shipment.order)
        .outerjoin(Receiver, Receiver.id == Order.receiver_id)
        .outerjoin(Channel, Channel.id == Order.channel_id)
        .options(
            selectinload(Shipment.order).selectinload(Order.receiver),
            selectinload(Shipment.order).selectinload(Order.channel).selectinload(Channel.courier),
            selectinload(Shipment.shipment_items)
            .selectinload(ShipmentItem.product),
        )
        .order_by(Shipment.created_at.desc())
    )
    if order_id is not None:
        # 합배송 하위 주문 id로 조회된 경우에도 대표 주문의 shipment가 노출되어야 한다.
        rep_id_subq = (
            select(Order.consolidated_to_order_id)
            .where(Order.id == order_id)
            .scalar_subquery()
        )
        q = q.where(or_(Shipment.order_id == order_id, Shipment.order_id == rep_id_subq))
    if invoice_number is not None and invoice_number.strip():
        q = q.where(Shipment.invoice_number.ilike(f"%{invoice_number.strip()}%"))
    if order_status is not None:
        q = q.where(Order.status == order_status)
    if channel_ids is not None and channel_ids.strip():
        tokens = [t.strip() for t in channel_ids.split(",") if t.strip()]
        parsed: list[UUID] = []
        for t in tokens:
            try:
                parsed.append(UUID(t))
            except Exception:
                raise HTTPException(status_code=422, detail="Invalid channel_ids")
        if parsed:
            q = q.where(Order.channel_id.in_(parsed))
    elif channel_id is not None:
        q = q.where(Order.channel_id == channel_id)
    if channel_name is not None and channel_name.strip():
        q = q.where(Channel.name.ilike(f"%{channel_name.strip()}%"))
    if receiver_name is not None and receiver_name.strip():
        q = q.where(Receiver.name.ilike(f"%{receiver_name.strip()}%"))
    if receiver_phone is not None and receiver_phone.strip():
        q = q.where(Receiver.phone.ilike(f"%{receiver_phone.strip()}%"))
    if receiver_zip_code is not None and receiver_zip_code.strip():
        q = q.where(Receiver.zip_code.ilike(f"%{receiver_zip_code.strip()}%"))
    if receiver_address is not None and receiver_address.strip():
        ra = receiver_address.strip()
        q = q.where(or_(Receiver.address.ilike(f"%{ra}%"), Receiver.address_detail.ilike(f"%{ra}%")))
    if order_date_start is not None:
        q = q.where(func.date(Order.order_date) >= order_date_start)
    if order_date_end is not None:
        q = q.where(func.date(Order.order_date) <= order_date_end)
    if order_placed_date_start is not None:
        q = q.where(func.date(Shipment.order_placed_date) >= order_placed_date_start)
    if order_placed_date_end is not None:
        q = q.where(func.date(Shipment.order_placed_date) <= order_placed_date_end)
    if shipping_date_start is not None:
        q = q.where(func.date(Shipment.shipping_date) >= shipping_date_start)
    if shipping_date_end is not None:
        q = q.where(func.date(Shipment.shipping_date) <= shipping_date_end)

    if product_query is not None and product_query.strip():
        pq = product_query.strip()
        q = (
            q.join(ShipmentItem, ShipmentItem.shipment_id == Shipment.id)
            .join(Product, Product.id == ShipmentItem.product_id)
            .where(or_(Product.product_code.ilike(f"%{pq}%"), Product.name.ilike(f"%{pq}%")))
            .distinct(Shipment.id)
        )

    return await apaginate(db, q, params, transformer=_transform_shipments)


@router.post("/excel/selected")
async def download_selected_shipments_excel(
    body: ShipmentSelectedExcelRequest,
    db: AsyncSession = Depends(get_async_session),
    user=Depends(current_active_user),
):
    _ = await require_permission("shipments", "read")(user, db)
    """
    배송 선택 다운로드(엑셀):
    - 사용자가 선택한 shipment_ids만 대상으로 엑셀 생성
    - UI '선택 다운로드' 용도
    """
    ids = list(dict.fromkeys(body.shipment_ids or []))
    if not ids:
        raise HTTPException(status_code=422, detail="shipment_ids is required")

    q = (
        select(Shipment)
        .join(Shipment.order)
        .outerjoin(Receiver, Receiver.id == Order.receiver_id)
        .outerjoin(Channel, Channel.id == Order.channel_id)
        .options(
            selectinload(Shipment.order).selectinload(Order.receiver),
            selectinload(Shipment.order).selectinload(Order.channel).selectinload(Channel.courier),
            selectinload(Shipment.shipment_items).selectinload(ShipmentItem.product),
        )
        .where(Shipment.id.in_(ids))
        .order_by(Shipment.created_at.desc())
    )
    result = await db.execute(q)
    shipments: list[Shipment] = result.scalars().unique().all()
    if not shipments:
        raise HTTPException(status_code=404, detail="No shipments found")

    wb = Workbook()
    ws = wb.active
    ws.title = "발주서"
    _append_excel_header(ws, PLACE_ORDER_EXCEL_HEADERS)
    for sh in shipments:
        row = _shipment_to_place_order_row(sh)
        if row:
            ws.append(row)

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"shipments_selected_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"
        },
    )


@router.get("/excel")
async def download_shipments_excel(
    order_id: UUID | None = Query(default=None),
    invoice_number: str | None = Query(default=None),
    order_status: OrderStatus | None = Query(default=None),
    channel_id: UUID | None = Query(default=None),
    channel_name: str | None = Query(default=None),
    receiver_name: str | None = Query(default=None),
    receiver_phone: str | None = Query(default=None),
    receiver_zip_code: str | None = Query(default=None),
    receiver_address: str | None = Query(default=None),
    product_query: str | None = Query(default=None),
    order_date_start: date | None = Query(default=None),
    order_date_end: date | None = Query(default=None),
    order_placed_date_start: date | None = Query(default=None),
    order_placed_date_end: date | None = Query(default=None),
    shipping_date_start: date | None = Query(default=None),
    shipping_date_end: date | None = Query(default=None),
    db: AsyncSession = Depends(get_async_session),
    user=Depends(current_active_user),
):
    _ = await require_permission("shipments", "read")(user, db)
    """
    배송 전체 다운로드(엑셀):
    - 검색 조건에 맞는 shipments를 조회해 엑셀 생성
    - 포맷은 발주서 엑셀(/order-placed/excel)과 동일
    """
    q = (
        select(Shipment)
        .join(Shipment.order)
        .outerjoin(Receiver, Receiver.id == Order.receiver_id)
        .outerjoin(Channel, Channel.id == Order.channel_id)
        .options(
            selectinload(Shipment.order).selectinload(Order.receiver),
            selectinload(Shipment.order).selectinload(Order.channel).selectinload(Channel.courier),
            selectinload(Shipment.shipment_items).selectinload(ShipmentItem.product),
        )
        .order_by(Shipment.created_at.desc())
    )
    if order_id is not None:
        q = q.where(Shipment.order_id == order_id)
    if invoice_number is not None and invoice_number.strip():
        q = q.where(Shipment.invoice_number.ilike(f"%{invoice_number.strip()}%"))
    if order_status is not None:
        q = q.where(Order.status == order_status)
    if channel_id is not None:
        q = q.where(Order.channel_id == channel_id)
    if channel_name is not None and channel_name.strip():
        q = q.where(Channel.name.ilike(f"%{channel_name.strip()}%"))
    if receiver_name is not None and receiver_name.strip():
        q = q.where(Receiver.name.ilike(f"%{receiver_name.strip()}%"))
    if receiver_phone is not None and receiver_phone.strip():
        q = q.where(Receiver.phone.ilike(f"%{receiver_phone.strip()}%"))
    if receiver_zip_code is not None and receiver_zip_code.strip():
        q = q.where(Receiver.zip_code.ilike(f"%{receiver_zip_code.strip()}%"))
    if receiver_address is not None and receiver_address.strip():
        ra = receiver_address.strip()
        q = q.where(or_(Receiver.address.ilike(f"%{ra}%"), Receiver.address_detail.ilike(f"%{ra}%")))
    if order_date_start is not None:
        q = q.where(func.date(Order.order_date) >= order_date_start)
    if order_date_end is not None:
        q = q.where(func.date(Order.order_date) <= order_date_end)
    if order_placed_date_start is not None:
        q = q.where(func.date(Shipment.order_placed_date) >= order_placed_date_start)
    if order_placed_date_end is not None:
        q = q.where(func.date(Shipment.order_placed_date) <= order_placed_date_end)
    if shipping_date_start is not None:
        q = q.where(func.date(Shipment.shipping_date) >= shipping_date_start)
    if shipping_date_end is not None:
        q = q.where(func.date(Shipment.shipping_date) <= shipping_date_end)
    if product_query is not None and product_query.strip():
        pq = product_query.strip()
        q = (
            q.join(ShipmentItem, ShipmentItem.shipment_id == Shipment.id)
            .join(Product, Product.id == ShipmentItem.product_id)
            .where(or_(Product.product_code.ilike(f"%{pq}%"), Product.name.ilike(f"%{pq}%")))
            .distinct(Shipment.id)
        )

    result = await db.execute(q)
    shipments: list[Shipment] = result.scalars().unique().all()
    if not shipments:
        raise HTTPException(status_code=404, detail="No shipments found")

    wb = Workbook()
    ws = wb.active
    ws.title = "발주서"
    _append_excel_header(ws, PLACE_ORDER_EXCEL_HEADERS)
    for sh in shipments:
        row = _shipment_to_place_order_row(sh)
        if row:
            ws.append(row)

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename_utf8 = f"배송목록_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    filename_ascii = f"shipments_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    disposition = (
        f'attachment; filename="{filename_ascii}"; '
        f"filename*=UTF-8''{quote(filename_utf8)}"
    )
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": disposition},
    )


@router.post("/order-placed/excel")
async def download_order_placed_excel(
    db: AsyncSession = Depends(get_async_session),
    user=Depends(current_active_user),
):
    _ = await require_permission("shipments", "read")(user, db)
    _ = await require_permission("shipments", "update")(user, db)
    """
    발주서 엑셀 다운로드:
    - 주문 상태가 ORDER_PLACED인 배송 건만 조회해 엑셀 생성
    - 같은 요청 안에서 해당 배송에 연결된 주문을 SHIPPING_WAITING으로 전환하고 커밋한 뒤 파일을 반환한다.
    """
    q = (
        select(Shipment)
        .join(Shipment.order)
        .where(Order.is_delete == False, Order.status == OrderStatus.ORDER_PLACED)
        .options(
            selectinload(Shipment.order).selectinload(Order.receiver),
            selectinload(Shipment.order).selectinload(Order.channel).selectinload(Channel.courier),
            selectinload(Shipment.shipment_items).selectinload(ShipmentItem.product),
        )
        .order_by(Shipment.created_at.desc())
    )
    result = await db.execute(q)
    shipments: list[Shipment] = result.scalars().unique().all()

    if not shipments:
        raise HTTPException(status_code=404, detail="No ORDER_PLACED shipments to export")

    wb = Workbook()
    ws = wb.active
    ws.title = "발주서"

    _append_excel_header(ws, PLACE_ORDER_EXCEL_HEADERS)

    for sh in shipments:
        row = _shipment_to_place_order_row(sh)
        if row:
            ws.append(row)

    orders_by_id: dict[UUID, Order] = {}
    for sh in shipments:
        o = sh.order
        if o is None:
            continue
        orders_by_id[o.id] = o

    # 합배송 하위 주문도 함께 배송대기로 전환해야 한다.
    rep_order_ids = list(orders_by_id.keys())
    if rep_order_ids:
        sub_res = await db.execute(
            select(Order).where(
                Order.is_delete == False,
                Order.consolidated_to_order_id.in_(rep_order_ids),
            )
        )
        for sub in sub_res.scalars().all():
            orders_by_id[sub.id] = sub

    for o in orders_by_id.values():
        if o.status != OrderStatus.ORDER_PLACED:
            continue
        before = _order_snapshot(o)
        from_status = o.status
        o.status = OrderStatus.SHIPPING_WAITING
        o.update_user_id = user.id
        await _try_add_order_history(
            db,
            OrderHistory(
                order_id=o.id,
                action_type=OrderHistoryActionType.SHIPPING_WAITING,
                update_user_id=user.id,
                from_status=from_status,
                to_status=o.status,
                before_update=before,
                after_update=_order_snapshot(o),
                reason="order_placed_excel_download",
            ),
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    await db.commit()
    # Header values must be latin-1 encodable. Use RFC5987 filename* for UTF-8.
    filename_utf8 = f"발주서_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filename_ascii = f"order_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    disposition = (
        f'attachment; filename="{filename_ascii}"; '
        f"filename*=UTF-8''{quote(filename_utf8)}"
    )
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": disposition},
    )


@router.post("/order-placed/mark-shipping-waiting")
async def mark_order_placed_as_shipping_waiting(
    db: AsyncSession = Depends(get_async_session),
    user=Depends(current_active_user),
):
    _ = await require_permission("shipments", "update")(user, db)
    """
    발주(order_placed) 상태 주문 전체를 배송대기(shipping_waiting)로 전환한다.
    발주서 다운로드는 /order-placed/excel에서 엑셀과 함께 처리되며, 본 엔드포인트는
    다운로드 없이 상태만 맞출 때(예: 운영 보정)를 위한 보조용이다.
    """
    q = select(Order).where(Order.is_delete == False, Order.status == OrderStatus.ORDER_PLACED)
    result = await db.execute(q)
    orders = result.scalars().all()
    for o in orders:
        before = _order_snapshot(o)
        from_status = o.status
        o.status = OrderStatus.SHIPPING_WAITING
        o.update_user_id = user.id
        await _try_add_order_history(
            db,
            OrderHistory(
                order_id=o.id,
                action_type=OrderHistoryActionType.SHIPPING_WAITING,
                update_user_id=user.id,
                from_status=from_status,
                to_status=o.status,
                before_update=before,
                after_update=_order_snapshot(o),
                reason="mark_shipping_waiting",
            ),
        )

    await db.commit()
    return {"updated": len(orders)}


@router.post("/invoices/upload")
async def upload_invoices_excel(
    file: UploadFile = File(...),
    logistics_location_id: UUID | None = Form(default=None),
    db: AsyncSession = Depends(get_async_session),
    user=Depends(current_active_user),
):
    _ = await require_permission("shipments", "update")(user, db)
    """
    송장 업로드:
    - E(고객주문번호) = shipments.id
    - K(운송장번호) -> shipments.invoice_number 업데이트
    - 해당 주문 상태를 shipping 으로 전환
    """
    is_superuser = bool(getattr(user, "is_superuser", False))
    user_location_id = getattr(user, "logistics_location_id", None)
    chosen_location_id: UUID | None = logistics_location_id

    if is_superuser:
        if chosen_location_id is None:
            raise HTTPException(status_code=422, detail="logistics_location_id is required")
    else:
        if user_location_id is None:
            raise HTTPException(status_code=422, detail="담당 출고지가 설정되지 않았습니다.")
        if chosen_location_id is None:
            chosen_location_id = user_location_id
        if chosen_location_id != user_location_id:
            raise HTTPException(status_code=403, detail="담당 출고지와 다른 출고지는 선택할 수 없습니다.")

    if not file.filename:
        raise HTTPException(status_code=422, detail="No file provided")

    content = await file.read()
    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Invalid excel file: {e}")

    ws = wb.active
    # Header row assumed at 1
    updates: dict[UUID, str] = {}
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # E=4, K=10 (0-based)
        shipment_id_raw = row[4] if len(row) > 4 else None
        invoice_raw = row[10] if len(row) > 10 else None
        if shipment_id_raw is None or invoice_raw is None:
            continue
        sid = str(shipment_id_raw).strip()
        inv = str(invoice_raw).strip()
        if not sid or not inv:
            continue
        try:
            shipment_id = UUID(sid)
        except Exception:
            continue
        updates[shipment_id] = inv

    if not updates:
        return {"updated_shipments": 0, "missing_shipments": [], "message": "No rows to update"}

    result = await db.execute(
        select(Shipment)
        .where(Shipment.id.in_(list(updates.keys())))
        .options(
            selectinload(Shipment.order).selectinload(Order.consolidated_sub_orders),
        )
    )
    shipments: list[Shipment] = result.scalars().unique().all()
    shipment_by_id = {s.id: s for s in shipments}

    missing = [str(sid) for sid in updates.keys() if sid not in shipment_by_id]

    now = datetime.now(timezone.utc)
    updated_count = 0
    touched_orders: set[UUID] = set()
    # 합배송 하위 주문에 전파할 송장: sub_order_id -> invoice_number
    sub_order_invoice: dict[UUID, str] = {}
    deduct_shipment_ids: set[UUID] = set()
    for sid, inv in updates.items():
        sh = shipment_by_id.get(sid)
        if not sh:
            continue
        # idempotency: invoice_number이 처음 세팅되는 shipment만 재고 차감
        should_deduct = sh.invoice_number is None
        sh.invoice_number = inv
        sh.logistics_location_id = chosen_location_id
        if sh.shipping_date is None:
            sh.shipping_date = now
        updated_count += 1
        if should_deduct:
            deduct_shipment_ids.add(sh.id)
        if sh.order is not None:
            touched_orders.add(sh.order.id)
            # 합배송 하위 주문들도 함께 상태/송장/정산 전파 대상으로 포함
            for sub in sh.order.consolidated_sub_orders or []:
                touched_orders.add(sub.id)
                sub_order_invoice.setdefault(sub.id, inv)

    if touched_orders:
        # shipment 변경을 동일 트랜잭션에서 조회에 반영
        await db.flush()

        result2 = await db.execute(
            select(Order)
            .where(Order.id.in_(list(touched_orders)))
            .options(selectinload(Order.shipments))
        )
        orders = result2.scalars().unique().all()
        for o in orders:
            before = _order_snapshot(o)
            from_status = o.status
            o.status = OrderStatus.SHIPPING
            if o.shipping_date is None:
                o.shipping_date = now
            o.update_user_id = user.id
            # 주문.invoice_number는 단일 값만 가능 → 다중 배송 시 "첫 번째" 배송건 송장만 반영
            shs = sorted(
                o.shipments or [],
                key=lambda s: (s.order_placed_date, s.created_at, s.id),
            )
            if shs:
                o.invoice_number = shs[0].invoice_number
            elif o.id in sub_order_invoice:
                # 합배송 하위 주문: 대표 주문의 shipment 송장으로 전파
                o.invoice_number = sub_order_invoice[o.id]
            await _try_add_order_history(
                db,
                OrderHistory(
                    order_id=o.id,
                    action_type=OrderHistoryActionType.SHIPPING,
                    update_user_id=user.id,
                    from_status=from_status,
                    to_status=o.status,
                    before_update=before,
                    after_update=_order_snapshot(o),
                    reason="송장 업로드",
                ),
            )

        # 정산 대상 생성(대기): 배송(SHIPPING)된 주문 중 settlement가 없는 건만 생성 (멱등)
        existing_res = await db.execute(
            select(Settlement).where(Settlement.order_id.in_([o.id for o in orders]))
        )
        existing = existing_res.scalars().all()
        existing_order_ids = {s.order_id for s in existing}
        for o in orders:
            if o.id in existing_order_ids:
                continue
            order_price = getattr(o, "price", None)
            if order_price is None:
                continue
            commission = int(getattr(o, "commission", 0) or 0)
            st = Settlement(
                order_id=o.id,
                order_price=order_price,
                price=order_price,
                commission=commission,
                state=SettlementState.PENDING,
                update_user_id=user.id,
            )
            db.add(st)
            await db.flush()
            await _try_add_settlement_history(
                db,
                SettlementHistory(
                    settlement_id=st.id,
                    order_id=o.id,
                    action_type=SettlementHistoryActionType.CREATED,
                    update_user_id=user.id,
                    reason="송장 업로드(배송) - 정산 대기 생성",
                    from_state=None,
                    to_state=st.state,
                    before_price=None,
                    after_price=st.price,
                ),
            )

    # 재고 차감 (OUTBOUND): 유통기한(=expiration_date)이 짧은 것부터
    if deduct_shipment_ids:
        inventory_warnings: list[str] = []
        items_result = await db.execute(
            select(ShipmentItem).where(ShipmentItem.shipment_id.in_(list(deduct_shipment_ids)))
        )
        items: list[ShipmentItem] = items_result.scalars().all()

        required_by_product: dict[UUID, int] = {}
        for it in items:
            if it.product_id is None:
                continue
            required_by_product[it.product_id] = required_by_product.get(it.product_id, 0) + int(it.quantity)

        for product_id, required_qty in required_by_product.items():
            remaining = required_qty

            stocks_result = await db.execute(
                select(Stock)
                .where(
                    Stock.product_id == product_id,
                    Stock.is_delete == False,
                    Stock.logistics_location_id == chosen_location_id,
                )
                .order_by(
                    Stock.expiration_date.asc(),
                    Stock.stock_date.asc(),
                    Stock.created_at.asc(),
                    Stock.id.asc(),
                )
            )
            stocks: list[Stock] = stocks_result.scalars().all()

            if not stocks:
                inventory_warnings.append(
                    f"재고 없음: product_id={product_id}, 필요수량={required_qty}"
                )
                continue

            for i, st in enumerate(stocks):
                if remaining <= 0:
                    break

                # 재고 부족이어도 실패하지 않도록:
                # - 마지막 배치에 대해서는 남은 수량을 그대로 차감하여 음수가 가능하게 둔다.
                is_last = i == len(stocks) - 1
                take_qty = int(remaining if is_last else min(int(st.quantity), remaining))
                if take_qty <= 0:
                    continue

                before_snapshot = _stock_snapshot(st)
                st.quantity -= take_qty

                if st.quantity < 0:
                    inventory_warnings.append(
                        f"재고 음수 발생: product_id={product_id}, batch={st.batch_code}, 음수수량={st.quantity}"
                    )

                history = _build_outbound_history(
                    stock=st,
                    action_quantity=-take_qty,
                    user_id=user.id,
                    before_update=before_snapshot,
                )
                db.add(history)

                remaining -= take_qty

    await db.commit()

    return {
        "updated_shipments": updated_count,
        "missing_shipments": missing,
        "inventory_warnings": locals().get("inventory_warnings", []),
    }

