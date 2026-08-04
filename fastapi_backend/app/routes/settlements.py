from __future__ import annotations

import io
import logging
import re
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from typing import Any, Literal

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from fastapi_pagination.ext.sqlalchemy import apaginate
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, ConfigDict, Field
from sqlalchemy import (
    Integer,
    and_,
    case,
    cast,
    delete,
    exists,
    func,
    literal,
    or_,
    select,
    text,
    update,
)
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm.attributes import flag_modified

from app.config import settings
from app.database import get_async_session
from app.models import (
    ClientNameMapping,
    OwnedCourseOpening,
    OwnedSettlementCompareResultRow,
    SchedulerJob,
    SchedulerJobQueue,
    Settlement,
    SettlementConsolidated,
    User,
)
from app.pagination import MAX_PAGE_SIZE, Page, Params
from app.routes.courses import (
    MAX_EXPORT_ROWS,
    SCROLL_BATCH_SIZE,
    SCROLL_KEEPALIVE,
    _clear_scroll_safe,
    _parse_course_from_es,
)
from app.users import current_active_user

logger = logging.getLogger(__name__)

router = APIRouter()

QUEUE_ACTION_RUN_NOW = "RUN_NOW"
QUEUE_STATUS_PENDING = "PENDING"
OWNED_OPENING_EXTRACT_JOB_KEY = "owned_course_opening_extract"

SETTLEMENTS_CONSOLIDATED_MV = "settlements_consolidated"

CREATE_SETTLEMENTS_CONSOLIDATED_SQL = """
CREATE MATERIALIZED VIEW settlements_consolidated AS
SELECT
  purchase_ym, purchase_year, sales_ym, client_name, course_name,
  education_period, education_period_date,
  SUM(headcount) AS headcount,
  base_tuition, textbook_fee, exclude_amount, share_rate,
  SUM(net_sales) AS net_sales,
  settlement_rate,
  SUM(settlement_amount) AS settlement_amount,
  note, sales_rep
FROM settlements
WHERE is_delete = false
GROUP BY
  purchase_ym, purchase_year, sales_ym, client_name, course_name,
  education_period, education_period_date,
  base_tuition, textbook_fee, exclude_amount, share_rate,
  settlement_rate, note, sales_rep
WITH DATA
"""

CREATE_SETTLEMENTS_CONSOLIDATED_INDEX_SQL = """
CREATE INDEX ix_settlements_consolidated_match
  ON settlements_consolidated (client_name, course_name, education_period_date)
"""

EXCEL_HEADERS: list[tuple[str, str]] = [
    ("매입년월", "purchase_ym"),
    ("매출년월", "sales_ym"),
    ("고객사", "client_name"),
    ("과정명", "course_name"),
    ("교육기간", "education_period"),
    ("인원", "headcount"),
    ("기준수강료", "base_tuition"),
    ("교재비", "textbook_fee"),
    ("정산제외금", "exclude_amount"),
    ("배분율", "share_rate"),
    ("순매출액", "net_sales"),
    ("정산율", "settlement_rate"),
    ("정산액", "settlement_amount"),
    ("비고", "note"),
    ("영업대표", "sales_rep"),
]

# 내보내기 전용: 교육기간 오른쪽에 변환 날짜 컬럼
EXPORT_EXCEL_HEADERS: list[tuple[str, str]] = [
    ("매입년월", "purchase_ym"),
    ("매출년월", "sales_ym"),
    ("고객사", "client_name"),
    ("과정명", "course_name"),
    ("교육기간", "education_period"),
    ("교육기간(변환)", "education_period_date"),
    ("인원", "headcount"),
    ("기준수강료", "base_tuition"),
    ("교재비", "textbook_fee"),
    ("정산제외금", "exclude_amount"),
    ("배분율", "share_rate"),
    ("순매출액", "net_sales"),
    ("정산율", "settlement_rate"),
    ("정산액", "settlement_amount"),
    ("비고", "note"),
    ("영업대표", "sales_rep"),
]

HEADER_TO_FIELD = {label: field for label, field in EXCEL_HEADERS}
HEADER_TO_FIELD_NORM = {
    label.replace(" ", ""): field for label, field in EXCEL_HEADERS
}

_YM_DIGITS_RE = re.compile(r"\D+")
# 전처리 후 구간: YYYYMMDD-YYYYMMDD / YYYYMMDD~YYYYMMDD
_COMPACT_RANGE_RE = re.compile(r"^(\d{8})[~-](\d{8})$")
# 전처리 후 구간: 2023-11-17~2023-12-14 / 2023-11-17-2023-12-14
_HYPHEN_DATE_RANGE_RE = re.compile(
    r"^(\d{4}-\d{1,2}-\d{1,2})[~-](\d{4}-\d{1,2}-\d{1,2})$"
)
# 전처리 후 단일 날짜: 숫자 3부분 (구분자 - 또는 /)
_DATE_PARTS_RE = re.compile(
    r"^(\d{1,4})[-/](\d{1,2})[-/](\d{1,4})$"
)
_DATE_FORMATS_YEAR_FIRST = (
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%Y%m%d",
    "%y-%m-%d",
    "%y/%m/%d",
)

def _parse_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


async def _refresh_settlements_consolidated(session: AsyncSession) -> None:
    """정산 테이블 변경 후 consolidated MV 갱신."""
    await session.execute(
        text(f"REFRESH MATERIALIZED VIEW {SETTLEMENTS_CONSOLIDATED_MV}")
    )
    await session.commit()


def ensure_settlements_consolidated_mv(connection) -> None:
    """테스트/초기화용: MV를 (재)생성한다. sync connection."""
    connection.execute(
        text("DROP MATERIALIZED VIEW IF EXISTS settlements_consolidated CASCADE")
    )
    connection.execute(text(CREATE_SETTLEMENTS_CONSOLIDATED_SQL))
    connection.execute(text(CREATE_SETTLEMENTS_CONSOLIDATED_INDEX_SQL))


def drop_settlements_consolidated_mv(connection) -> None:
    """테스트 teardown용. sync connection."""
    connection.execute(
        text("DROP MATERIALIZED VIEW IF EXISTS settlements_consolidated CASCADE")
    )


def _normalize_ym(value: Any) -> str | None:
    """매입/매출년월 → YYYYMM 문자열."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        text = str(int(value))
    else:
        text = _YM_DIGITS_RE.sub("", str(value).strip())
    if len(text) >= 6 and text[:6].isdigit():
        return text[:6]
    if len(text) == 4 and text.isdigit():
        return None
    return None


def _year_from_ym(ym: str) -> int:
    return int(ym[:4])


def _parse_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        raise ValueError("정수 값이 아닙니다.")
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    return int(Decimal(text))


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        raise ValueError("숫자 값이 아닙니다.")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    return Decimal(text)


def _parse_rate(value: Any) -> Decimal | None:
    """배분율/정산율: '30%' → 0.3, '0.36' → 0.36, 0.5 → 0.5."""
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        raise ValueError("비율 값이 아닙니다.")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    if text.endswith("%"):
        num = Decimal(text[:-1].strip())
        return num / Decimal("100")
    return Decimal(text)


def _safe_date(year: int, month: int, day: int) -> date | None:
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _normalize_year(year: int) -> int:
    """2자리 연도 → 2000년대 기준 4자리."""
    if year < 100:
        return 2000 + year
    return year


def _parse_dmy_or_mdy(a: int, b: int, year: int) -> date | None:
    """연도가 끝인 경우. 모호하면 일-월-년(DMY). 한쪽만 타당하면 그쪽."""
    year = _normalize_year(year)
    dmy = _safe_date(year, b, a)  # a=day, b=month
    mdy = _safe_date(year, a, b)  # a=month, b=day
    if a > 12 and b <= 12:
        return dmy
    if b > 12 and a <= 12:
        return mdy
    # 둘 다 가능하면 DMY 우선
    return dmy or mdy


def _preprocess_education_period(text: str) -> str:
    """교육기간 전처리: 공백 제거, '.' → '-'."""
    normalized = re.sub(r"\s+", "", text.strip())
    normalized = normalized.replace(".", "-")
    # 유사 구간/대시 문자를 통일
    normalized = re.sub(r"[～∼〜⁓]", "~", normalized)
    normalized = re.sub(r"[–—]", "-", normalized)
    return normalized


def _extract_range_start(text: str) -> str:
    """전처리된 교육기간 구간이면 시작일 텍스트만 반환."""
    for pattern in (_COMPACT_RANGE_RE, _HYPHEN_DATE_RANGE_RE):
        m = pattern.match(text)
        if m:
            return m.group(1)

    if "~" in text:
        return text.split("~", maxsplit=1)[0]
    return text


def _parse_single_date_text(text: str) -> date | None:
    if not text:
        return None

    for fmt in _DATE_FORMATS_YEAR_FIRST:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    # YYYYMMDD (또는 앞 8자리)
    digits = _YM_DIGITS_RE.sub("", text)
    if len(digits) >= 8 and digits[:8].isdigit():
        parsed = _safe_date(int(digits[:4]), int(digits[4:6]), int(digits[6:8]))
        if parsed is not None:
            return parsed

    m = _DATE_PARTS_RE.match(text)
    if not m:
        return None

    p1, p2, p3 = m.group(1), m.group(2), m.group(3)
    n1, n2, n3 = int(p1), int(p2), int(p3)

    # 연도 선행: YYYY-M-D
    if len(p1) == 4:
        return _safe_date(n1, n2, n3)

    # 연도 후행: D-M-YYYY / D-M-YY (모호 시 DMY)
    if len(p3) in (2, 4):
        return _parse_dmy_or_mdy(n1, n2, n3)

    return None


def _parse_education_period_date(value: Any) -> date | None:
    """교육기간 → YYYY-MM-DD 날짜. 구간이면 앞쪽 날짜만. 실패 시 None."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    normalized = _preprocess_education_period(text)
    first = _extract_range_start(normalized)
    return _parse_single_date_text(first)


class SettlementListItem(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: int
    purchase_ym: str
    purchase_year: int
    sales_ym: str | None
    client_name: str
    course_name: str
    education_period: str | None
    education_period_date: date | None
    headcount: int | None
    base_tuition: Decimal | None
    net_sales: Decimal | None
    settlement_amount: Decimal | None
    sales_rep: str | None


class SettlementRead(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: int
    purchase_ym: str
    purchase_year: int
    sales_ym: str | None
    client_name: str
    course_name: str
    education_period: str | None
    education_period_date: date | None
    headcount: int | None
    base_tuition: Decimal | None
    textbook_fee: Decimal | None
    exclude_amount: Decimal | None
    share_rate: Decimal | None
    net_sales: Decimal | None
    settlement_rate: Decimal | None
    settlement_amount: Decimal | None
    note: str | None
    sales_rep: str | None


class SettlementImportError(BaseModel):
    row: int
    message: str


class SettlementImportResult(BaseModel):
    deleted: int
    created: int
    failed: int
    errors: list[SettlementImportError] = Field(default_factory=list)


def _transform_list(items: list[Settlement]) -> list[SettlementListItem]:
    return [SettlementListItem.model_validate(row) for row in items]


def _row_to_fields(row_values: dict[str, Any]) -> dict[str, Any]:
    purchase_ym = _normalize_ym(row_values.get("purchase_ym"))
    if not purchase_ym:
        raise ValueError("매입년월이 올바르지 않습니다.")

    client_name = _parse_str(row_values.get("client_name"))
    if not client_name:
        raise ValueError("고객사가 비어 있습니다.")

    course_name = _parse_str(row_values.get("course_name"))
    if not course_name:
        raise ValueError("과정명이 비어 있습니다.")

    try:
        fields: dict[str, Any] = {
            "purchase_ym": purchase_ym,
            "purchase_year": _year_from_ym(purchase_ym),
            "sales_ym": _normalize_ym(row_values.get("sales_ym")),
            "client_name": client_name,
            "course_name": course_name,
            "education_period": _parse_str(row_values.get("education_period")),
            "education_period_date": _parse_education_period_date(
                row_values.get("education_period")
            ),
            "headcount": _parse_int(row_values.get("headcount")),
            "base_tuition": _parse_decimal(row_values.get("base_tuition")),
            "textbook_fee": _parse_decimal(row_values.get("textbook_fee")),
            "exclude_amount": _parse_decimal(row_values.get("exclude_amount")),
            "share_rate": _parse_rate(row_values.get("share_rate")),
            "net_sales": _parse_decimal(row_values.get("net_sales")),
            "settlement_rate": _parse_rate(row_values.get("settlement_rate")),
            "settlement_amount": _parse_decimal(row_values.get("settlement_amount")),
            "note": _parse_str(row_values.get("note")),
            "sales_rep": _parse_str(row_values.get("sales_rep")),
        }
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"숫자 파싱 실패: {exc}") from exc
    return fields


@router.get("", response_model=Page[SettlementListItem])
async def list_settlements(
    page: int = 1,
    size: int = 20,
    year: int | None = Query(default=None, description="매입년도"),
    client_name: str | None = Query(default=None, description="고객사 검색"),
    course_name: str | None = Query(default=None, description="과정명 검색"),
    session: AsyncSession = Depends(get_async_session),
    _: User = Depends(current_active_user),
):
    if size < 1:
        raise HTTPException(status_code=400, detail="size must be >= 1")
    if size > MAX_PAGE_SIZE:
        raise HTTPException(status_code=400, detail=f"size must be <= {MAX_PAGE_SIZE}")

    params = Params(page=page, size=size)
    stmt = select(Settlement).where(Settlement.is_delete == False)  # noqa: E712
    if year is not None:
        stmt = stmt.where(Settlement.purchase_year == year)
    if client_name and client_name.strip():
        stmt = stmt.where(Settlement.client_name.ilike(f"%{client_name.strip()}%"))
    if course_name and course_name.strip():
        stmt = stmt.where(Settlement.course_name.ilike(f"%{course_name.strip()}%"))
    stmt = stmt.order_by(Settlement.purchase_ym.desc(), Settlement.id.desc())
    return await apaginate(session, stmt, params, transformer=_transform_list)


@router.get("/import/template")
async def download_import_template(
    _: User = Depends(current_active_user),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "정산"
    ws.append([label for label, _ in EXCEL_HEADERS])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="settlements_template.xlsx"'
        },
    )


def _excel_cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, date) and not isinstance(value, datetime):
        return value.isoformat()
    return value


@router.get("/export")
async def export_settlements_consolidated(
    session: AsyncSession = Depends(get_async_session),
    _: User = Depends(current_active_user),
):
    """정리된 정산 MV(settlements_consolidated) 전체를 xlsx로 내보낸다."""
    total = await session.scalar(select(func.count()).select_from(SettlementConsolidated))
    total = int(total or 0)
    if total > MAX_EXPORT_ROWS:
        raise HTTPException(
            status_code=400,
            detail=f"내보낼 데이터가 {MAX_EXPORT_ROWS:,}건을 초과합니다.",
        )

    rows = (
        await session.execute(
            select(SettlementConsolidated).order_by(
                SettlementConsolidated.purchase_ym.desc(),
                SettlementConsolidated.client_name.asc(),
                SettlementConsolidated.course_name.asc(),
            )
        )
    ).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "정산"
    ws.append([label for label, _ in EXPORT_EXCEL_HEADERS])
    for row in rows:
        ws.append(
            [
                _excel_cell_value(getattr(row, field))
                for _, field in EXPORT_EXCEL_HEADERS
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="settlements_consolidated.xlsx"'
        },
    )


@router.post("/import", response_model=SettlementImportResult)
async def import_settlements(
    year: int = Form(..., ge=2000, le=2100, description="교체할 매입년도"),
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_async_session),
    _: User = Depends(current_active_user),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="xlsx 파일만 업로드할 수 있습니다.")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=400, detail=f"엑셀 파일을 읽을 수 없습니다: {exc}"
        ) from exc

    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        wb.close()
        raise HTTPException(status_code=400, detail="헤더 행이 없습니다.")

    col_map: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        label = str(cell).strip()
        field = HEADER_TO_FIELD.get(label) or HEADER_TO_FIELD_NORM.get(
            label.replace(" ", "")
        )
        if field:
            col_map[idx] = field

    required = {"purchase_ym", "client_name", "course_name"}
    if not required.issubset(set(col_map.values())):
        wb.close()
        raise HTTPException(
            status_code=400, detail="매입년월, 고객사, 과정명 열이 필요합니다."
        )

    soft_delete_result = await session.execute(
        update(Settlement)
        .where(
            Settlement.is_delete == False,  # noqa: E712
            Settlement.purchase_year == year,
        )
        .values(is_delete=True)
    )
    deleted = soft_delete_result.rowcount or 0

    created = 0
    failed = 0
    errors: list[SettlementImportError] = []
    to_insert: list[Settlement] = []

    for row_num, row in enumerate(rows, start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue

        row_values: dict[str, Any] = {}
        for idx, field in col_map.items():
            if idx < len(row):
                row_values[field] = row[idx]

        try:
            fields = _row_to_fields(row_values)
            if fields["purchase_year"] != year:
                failed += 1
                errors.append(
                    SettlementImportError(
                        row=row_num,
                        message=(
                            f"매입년월 년도({fields['purchase_year']})가 "
                            f"선택 년도({year})와 다릅니다."
                        ),
                    )
                )
                continue
            to_insert.append(Settlement(**fields))
            created += 1
        except Exception as exc:
            failed += 1
            errors.append(
                SettlementImportError(row=row_num, message=str(exc) or "처리 실패")
            )

    if to_insert:
        session.add_all(to_insert)
    await session.commit()
    wb.close()
    await _refresh_settlements_consolidated(session)
    return SettlementImportResult(
        deleted=deleted, created=created, failed=failed, errors=errors
    )


CompareStatus = Literal["matched", "partial", "unsettled", "unmapped"]


class OwnedSettlementCompareItem(BaseModel):
    institution_name: str | None = None
    client_name: str | None = None
    course_name: str | None = None
    tra_start_date: str | None = None
    tra_end_date: str | None = None
    reg_course_man: str | None = None
    status: CompareStatus


class OwnedSettlementCompareResult(BaseModel):
    """비교 요약 (목록은 /compare-owned/items)."""

    year: int
    total: int
    matched: int
    partial: int = 0
    unsettled: int
    unmapped: int
    has_result: bool = False
    cache_hit: bool = False  # has_result와 동일 (기존 프론트 호환)
    extracted_at: datetime | None = None
    compared_at: datetime | None = None


class OwnedOpeningExtractQueueRead(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: int
    year: int
    status: str
    row_count: int | None = None
    extracted_at: datetime | None = None
    error_message: str | None = None
    created_at: datetime
    updated_at: datetime


def _queue_to_extract_read(row: SchedulerJobQueue) -> OwnedOpeningExtractQueueRead:
    payload = dict(row.payload or {})
    year_raw = payload.get("year")
    year = int(year_raw) if year_raw is not None else 0
    row_count_raw = payload.get("row_count")
    extracted_raw = payload.get("extracted_at")
    extracted_at: datetime | None = None
    if isinstance(extracted_raw, str) and extracted_raw:
        try:
            extracted_at = datetime.fromisoformat(extracted_raw)
        except ValueError:
            extracted_at = None
    return OwnedOpeningExtractQueueRead(
        id=row.id,
        year=year,
        status=row.status,
        row_count=int(row_count_raw) if row_count_raw is not None else None,
        extracted_at=extracted_at,
        error_message=row.error_message,
        created_at=row.created_at,
        updated_at=row.updated_at,
    )


def _normalize_compare_date(value: Any) -> date | None:
    """ES traStartDate / 문자열 → date."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip().replace(".", "-").replace("/", "-")
    if not text:
        return None
    digits = re.sub(r"\D+", "", text)
    if len(digits) >= 8 and digits[:8].isdigit():
        try:
            return date(int(digits[:4]), int(digits[4:6]), int(digits[6:8]))
        except ValueError:
            return None
    try:
        return datetime.strptime(text[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def _format_date_cell(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value)


def _sql_strip_all_whitespace(column):
    """SQL 비교용: 공백·탭·개행 등 모든 공백 문자 제거."""
    return func.regexp_replace(func.coalesce(column, ""), r"\s+", "", "g")


def _compare_status_expr():
    """보유과정 row → matched | partial | unsettled | unmapped SQL 식."""
    owned_course_key = _sql_strip_all_whitespace(OwnedCourseOpening.course_name)
    settlement_key = and_(
        _sql_strip_all_whitespace(SettlementConsolidated.client_name)
        == _sql_strip_all_whitespace(ClientNameMapping.client_name),
        _sql_strip_all_whitespace(SettlementConsolidated.course_name)
        == owned_course_key,
        SettlementConsolidated.education_period_date
        == OwnedCourseOpening.tra_start_date,
    )
    has_settlement = exists(select(literal(1)).where(settlement_key))
    settlement_headcount = (
        select(func.coalesce(func.sum(SettlementConsolidated.headcount), 0))
        .where(settlement_key)
        .correlate_except(SettlementConsolidated)
        .scalar_subquery()
    )
    owned_man = func.trim(OwnedCourseOpening.reg_course_man)
    owned_headcount = case(
        (owned_man.op("~")(r"^[0-9]+$"), cast(owned_man, Integer)),
        else_=None,
    )
    return case(
        (ClientNameMapping.id.is_(None), literal("unmapped")),
        (
            or_(
                OwnedCourseOpening.tra_start_date.is_(None),
                owned_course_key == "",
                ~has_settlement,
            ),
            literal("unsettled"),
        ),
        (
            and_(
                has_settlement,
                owned_headcount.is_not(None),
                owned_headcount == settlement_headcount,
            ),
            literal("matched"),
        ),
        (has_settlement, literal("partial")),
        else_=literal("unsettled"),
    )


def _compare_base_stmt(year: int):
    status_expr = _compare_status_expr()
    inst_trim = func.trim(OwnedCourseOpening.institution_name)
    return (
        select(
            OwnedCourseOpening.institution_name,
            ClientNameMapping.client_name,
            OwnedCourseOpening.course_name,
            OwnedCourseOpening.tra_start_date,
            OwnedCourseOpening.tra_end_date,
            OwnedCourseOpening.reg_course_man,
            status_expr.label("status"),
            OwnedCourseOpening.id,
        )
        .outerjoin(
            ClientNameMapping,
            and_(
                ClientNameMapping.institution_name == inst_trim,
                ClientNameMapping.is_delete == False,  # noqa: E712
            ),
        )
        .where(
            OwnedCourseOpening.is_delete == False,  # noqa: E712
            OwnedCourseOpening.year == year,
        )
    )


def _normalize_name_for_mapping(value: str | None) -> str:
    """비교용 정규화: 모든 공백(공백·탭·개행 등) 제거."""
    if value is None:
        return ""
    return "".join(str(value).split())


async def _auto_register_identity_mappings(session: AsyncSession, year: int) -> int:
    """공백 제거 후 훈련기관명==고객사명이면 맵핑 자동 등록.

    soft-delete된 행은 되살리지 않음. 저장 값은 원본(양쪽 trim)을 유지한다.
    """
    owned_names = [
        n
        for n in (
            await session.execute(
                select(func.distinct(func.trim(OwnedCourseOpening.institution_name))).where(
                    OwnedCourseOpening.is_delete == False,  # noqa: E712
                    OwnedCourseOpening.year == year,
                    OwnedCourseOpening.institution_name.is_not(None),
                    func.trim(OwnedCourseOpening.institution_name) != "",
                )
            )
        ).scalars().all()
        if n
    ]
    client_names = [
        n
        for n in (
            await session.execute(
                select(func.distinct(func.trim(Settlement.client_name))).where(
                    Settlement.is_delete == False,  # noqa: E712
                    Settlement.client_name.is_not(None),
                    func.trim(Settlement.client_name) != "",
                )
            )
        ).scalars().all()
        if n
    ]
    if not owned_names or not client_names:
        return 0

    owned_by_key: dict[str, list[str]] = {}
    for name in owned_names:
        key = _normalize_name_for_mapping(name)
        if not key:
            continue
        owned_by_key.setdefault(key, []).append(name)

    clients_by_key: dict[str, list[str]] = {}
    for name in client_names:
        key = _normalize_name_for_mapping(name)
        if not key:
            continue
        clients_by_key.setdefault(key, []).append(name)

    matched_keys = set(owned_by_key) & set(clients_by_key)
    if not matched_keys:
        return 0

    candidate_institutions = sorted(
        {inst for key in matched_keys for inst in owned_by_key[key]}
    )
    existing = set(
        (
            await session.execute(
                select(ClientNameMapping.institution_name).where(
                    ClientNameMapping.institution_name.in_(candidate_institutions)
                )
            )
        ).scalars().all()
    )

    created = 0
    for key in sorted(matched_keys):
        client_candidates = sorted(clients_by_key[key])
        client_set = set(client_candidates)
        for inst in sorted(owned_by_key[key]):
            if inst in existing:
                continue
            # 표기가 동일한 고객사명을 우선, 없으면 공백 제거 결과가 같은 첫 고객사명
            client = inst if inst in client_set else client_candidates[0]
            session.add(
                ClientNameMapping(institution_name=inst, client_name=client)
            )
            existing.add(inst)
            created += 1

    if created:
        await session.commit()
    return created


async def _scroll_owned_courses(es, body: dict) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    scroll_id: str | None = None
    try:
        response = await es.search(
            index=settings.ES_COURSE_INDEX,
            body=body,
            scroll=SCROLL_KEEPALIVE,
        )
        scroll_id = response.get("_scroll_id")
        total_raw = response.get("hits", {}).get("total")
        if isinstance(total_raw, dict):
            total_count = int(total_raw.get("value", 0))
        else:
            total_count = int(total_raw or 0)
        if total_count > MAX_EXPORT_ROWS:
            raise HTTPException(
                status_code=400,
                detail=f"비교 대상이 {MAX_EXPORT_ROWS:,}건을 초과합니다. 조건을 좁혀 주세요.",
            )

        while True:
            hits = response.get("hits", {}).get("hits", [])
            if not hits:
                break
            for hit in hits:
                if len(rows) >= MAX_EXPORT_ROWS:
                    break
                item = _parse_course_from_es(hit.get("_source", {}))
                rows.append(
                    {
                        "institution_name": item.inst_name,
                        "course_name": item.course_name,
                        "tra_start_date": item.tra_start_date,
                        "tra_end_date": item.tra_end_date,
                        "reg_course_man": item.reg_course_man,
                    }
                )
            if len(rows) >= MAX_EXPORT_ROWS or len(hits) < SCROLL_BATCH_SIZE:
                break
            response = await es.scroll(scroll_id=scroll_id, scroll=SCROLL_KEEPALIVE)
    finally:
        await _clear_scroll_safe(es, scroll_id)
    return rows


@router.post(
    "/compare-owned/refresh",
    response_model=OwnedOpeningExtractQueueRead,
    status_code=202,
)
async def enqueue_owned_course_opening_refresh(
    year: int = Query(..., ge=2000, le=2100, description="추출할 훈련시작일 연도"),
    min_score: float = Query(default=0, ge=0, description="보유과정 ES 매칭 min_score"),
    session: AsyncSession = Depends(get_async_session),
    user: User = Depends(current_active_user),
):
    """ES 추출은 scheduler_job_queue에서 비동기로 처리. queue 행을 즉시 반환한다."""
    in_flight_rows = (
        await session.execute(
            select(SchedulerJobQueue)
            .where(
                SchedulerJobQueue.is_delete == False,  # noqa: E712
                SchedulerJobQueue.job_key == OWNED_OPENING_EXTRACT_JOB_KEY,
                SchedulerJobQueue.status.in_(("PENDING", "PROCESSING")),
            )
            .order_by(SchedulerJobQueue.id.desc())
            .limit(50)
        )
    ).scalars().all()
    for row in in_flight_rows:
        payload = dict(row.payload or {})
        if int(payload.get("year") or 0) == year:
            return _queue_to_extract_read(row)

    job_def = await session.get(SchedulerJob, OWNED_OPENING_EXTRACT_JOB_KEY)
    if job_def is None:
        session.add(
            SchedulerJob(
                job_key=OWNED_OPENING_EXTRACT_JOB_KEY,
                title="개설 보유과정 추출",
                enabled=False,
                cron_hour=3,
                cron_minute=0,
                timezone="Asia/Seoul",
                description="ES에서 개설 보유과정을 추출해 캐시 테이블에 연도별 적재",
            )
        )
        await session.commit()
    elif job_def.is_delete:
        job_def.is_delete = False
        job_def.enabled = False
        await session.commit()

    q = SchedulerJobQueue(
        job_key=OWNED_OPENING_EXTRACT_JOB_KEY,
        action=QUEUE_ACTION_RUN_NOW,
        status=QUEUE_STATUS_PENDING,
        requested_by_user_id=user.id,
        payload={"year": year, "min_score": min_score},
    )
    session.add(q)
    await session.commit()
    await session.refresh(q)

    q.payload = {**(q.payload or {}), "queue_id": q.id, "year": year, "min_score": min_score}
    flag_modified(q, "payload")
    await session.commit()
    await session.refresh(q)
    return _queue_to_extract_read(q)


@router.get(
    "/compare-owned/refresh-jobs/{queue_id}",
    response_model=OwnedOpeningExtractQueueRead,
)
async def get_owned_course_opening_refresh_queue(
    queue_id: int,
    session: AsyncSession = Depends(get_async_session),
    _: User = Depends(current_active_user),
):
    row = await session.get(SchedulerJobQueue, queue_id)
    if (
        row is None
        or row.is_delete
        or row.job_key != OWNED_OPENING_EXTRACT_JOB_KEY
    ):
        raise HTTPException(status_code=404, detail="추출 작업을 찾을 수 없습니다.")
    return _queue_to_extract_read(row)


async def _load_compare_summary(
    session: AsyncSession, year: int
) -> OwnedSettlementCompareResult:
    extracted_at = await session.scalar(
        select(func.max(OwnedCourseOpening.extracted_at)).where(
            OwnedCourseOpening.is_delete == False,  # noqa: E712
            OwnedCourseOpening.year == year,
        )
    )
    counts = (
        await session.execute(
            select(
                func.count().label("total"),
                func.coalesce(
                    func.sum(
                        case(
                            (OwnedSettlementCompareResultRow.status == "matched", 1),
                            else_=0,
                        )
                    ),
                    0,
                ).label("matched"),
                func.coalesce(
                    func.sum(
                        case(
                            (OwnedSettlementCompareResultRow.status == "partial", 1),
                            else_=0,
                        )
                    ),
                    0,
                ).label("partial"),
                func.coalesce(
                    func.sum(
                        case(
                            (OwnedSettlementCompareResultRow.status == "unsettled", 1),
                            else_=0,
                        )
                    ),
                    0,
                ).label("unsettled"),
                func.coalesce(
                    func.sum(
                        case(
                            (OwnedSettlementCompareResultRow.status == "unmapped", 1),
                            else_=0,
                        )
                    ),
                    0,
                ).label("unmapped"),
                func.max(OwnedSettlementCompareResultRow.compared_at).label(
                    "compared_at"
                ),
            ).where(
                OwnedSettlementCompareResultRow.is_delete == False,  # noqa: E712
                OwnedSettlementCompareResultRow.year == year,
            )
        )
    ).one()
    total = int(counts.total or 0)
    has_result = total > 0 or counts.compared_at is not None
    return OwnedSettlementCompareResult(
        year=year,
        total=total,
        matched=int(counts.matched or 0),
        partial=int(counts.partial or 0),
        unsettled=int(counts.unsettled or 0),
        unmapped=int(counts.unmapped or 0),
        has_result=has_result,
        cache_hit=has_result,
        extracted_at=extracted_at,
        compared_at=counts.compared_at,
    )


async def _persist_compare_results(
    session: AsyncSession, year: int
) -> OwnedSettlementCompareResult:
    """자동 맵핑 + openings 분류 후 해당 연도 결과 hard delete & insert."""
    await _auto_register_identity_mappings(session, year)

    extracted_at = await session.scalar(
        select(func.max(OwnedCourseOpening.extracted_at)).where(
            OwnedCourseOpening.is_delete == False,  # noqa: E712
            OwnedCourseOpening.year == year,
        )
    )
    if extracted_at is None:
        raise HTTPException(
            status_code=400,
            detail="해당 연도 추출 캐시가 없습니다. 먼저 추출/갱신을 실행하세요.",
        )

    classified = (
        await session.execute(
            _compare_base_stmt(year).order_by(
                OwnedCourseOpening.institution_name.asc().nulls_last(),
                OwnedCourseOpening.course_name.asc().nulls_last(),
                OwnedCourseOpening.tra_start_date.asc().nulls_last(),
                OwnedCourseOpening.id.asc(),
            )
        )
    ).all()
    if len(classified) > MAX_EXPORT_ROWS:
        raise HTTPException(
            status_code=400,
            detail=f"비교 대상이 {MAX_EXPORT_ROWS:,}건을 초과합니다.",
        )

    compared_at = datetime.now(timezone.utc)
    await session.execute(
        delete(OwnedSettlementCompareResultRow).where(
            OwnedSettlementCompareResultRow.year == year
        )
    )
    to_insert = [
        OwnedSettlementCompareResultRow(
            year=year,
            status=row.status,
            institution_name=row.institution_name,
            client_name=row.client_name,
            course_name=row.course_name,
            tra_start_date=row.tra_start_date
            if isinstance(row.tra_start_date, date)
            else _normalize_compare_date(row.tra_start_date),
            tra_end_date=row.tra_end_date
            if isinstance(row.tra_end_date, date)
            else _normalize_compare_date(row.tra_end_date),
            reg_course_man=row.reg_course_man,
            compared_at=compared_at,
        )
        for row in classified
    ]
    if to_insert:
        session.add_all(to_insert)
    await session.commit()

    return await _load_compare_summary(session, year)


def _transform_stored_compare_items(
    rows: list[OwnedSettlementCompareResultRow],
) -> list[OwnedSettlementCompareItem]:
    return [
        OwnedSettlementCompareItem(
            institution_name=row.institution_name,
            client_name=row.client_name,
            course_name=row.course_name,
            tra_start_date=_format_date_cell(row.tra_start_date),
            tra_end_date=_format_date_cell(row.tra_end_date),
            reg_course_man=row.reg_course_man,
            status=row.status,  # type: ignore[arg-type]
        )
        for row in rows
    ]


@router.get("/compare-owned", response_model=OwnedSettlementCompareResult)
async def get_owned_settlement_compare(
    year: int = Query(..., ge=2000, le=2100, description="비교 연도"),
    session: AsyncSession = Depends(get_async_session),
    _: User = Depends(current_active_user),
):
    """저장된 비교 결과 요약 조회 (재분류하지 않음)."""
    return await _load_compare_summary(session, year)


@router.post("/compare-owned", response_model=OwnedSettlementCompareResult)
async def run_owned_settlement_compare(
    year: int = Query(..., ge=2000, le=2100, description="비교 연도"),
    session: AsyncSession = Depends(get_async_session),
    _: User = Depends(current_active_user),
):
    """비교 실행: 해당 연도 결과를 delete & insert로 갱신."""
    return await _persist_compare_results(session, year)


@router.get("/compare-owned/items", response_model=Page[OwnedSettlementCompareItem])
async def list_owned_settlement_compare_items(
    year: int = Query(..., ge=2000, le=2100, description="비교 연도"),
    status: CompareStatus = Query(..., description="탭 상태"),
    page: int = 1,
    size: int = 50,
    session: AsyncSession = Depends(get_async_session),
    _: User = Depends(current_active_user),
):
    if size < 1:
        raise HTTPException(status_code=400, detail="size must be >= 1")
    if size > MAX_PAGE_SIZE:
        raise HTTPException(status_code=400, detail=f"size must be <= {MAX_PAGE_SIZE}")

    params = Params(page=page, size=size)
    stmt = (
        select(OwnedSettlementCompareResultRow)
        .where(
            OwnedSettlementCompareResultRow.is_delete == False,  # noqa: E712
            OwnedSettlementCompareResultRow.year == year,
            OwnedSettlementCompareResultRow.status == status,
        )
        .order_by(
            OwnedSettlementCompareResultRow.institution_name.asc().nulls_last(),
            OwnedSettlementCompareResultRow.course_name.asc().nulls_last(),
            OwnedSettlementCompareResultRow.tra_start_date.asc().nulls_last(),
            OwnedSettlementCompareResultRow.id.asc(),
        )
    )
    return await apaginate(
        session, stmt, params, transformer=_transform_stored_compare_items
    )


_COMPARE_EXPORT_HEADERS = [
    "훈련기관명",
    "고객사명",
    "과정명",
    "훈련시작일",
    "훈련종료일",
    "수강신청 인원",
]

_COMPARE_STATUS_SHEETS: list[tuple[CompareStatus, str]] = [
    ("unsettled", "미정산"),
    ("partial", "일부정산"),
    ("unmapped", "맵핑없음"),
    ("matched", "정산됨"),
]


def _append_compare_rows(ws, rows: list[OwnedSettlementCompareResultRow]) -> None:
    ws.append(list(_COMPARE_EXPORT_HEADERS))
    for row in rows:
        ws.append(
            [
                row.institution_name,
                row.client_name,
                row.course_name,
                _format_date_cell(row.tra_start_date),
                _format_date_cell(row.tra_end_date),
                row.reg_course_man,
            ]
        )


@router.get("/compare-owned/export")
async def export_owned_settlement_compare(
    year: int = Query(..., ge=2000, le=2100, description="비교 연도"),
    session: AsyncSession = Depends(get_async_session),
    _: User = Depends(current_active_user),
):
    """저장된 비교 결과 테이블을 시트별 xlsx로 내보낸다."""
    summary = await _load_compare_summary(session, year)
    if not summary.has_result:
        raise HTTPException(
            status_code=404,
            detail="해당 연도 비교 결과가 없습니다. 먼저 비교를 실행하세요.",
        )
    if summary.total > MAX_EXPORT_ROWS:
        raise HTTPException(
            status_code=400,
            detail=f"내보낼 데이터가 {MAX_EXPORT_ROWS:,}건을 초과합니다.",
        )

    all_rows = (
        await session.execute(
            select(OwnedSettlementCompareResultRow)
            .where(
                OwnedSettlementCompareResultRow.is_delete == False,  # noqa: E712
                OwnedSettlementCompareResultRow.year == year,
            )
            .order_by(
                OwnedSettlementCompareResultRow.institution_name.asc().nulls_last(),
                OwnedSettlementCompareResultRow.course_name.asc().nulls_last(),
                OwnedSettlementCompareResultRow.tra_start_date.asc().nulls_last(),
                OwnedSettlementCompareResultRow.id.asc(),
            )
        )
    ).scalars().all()

    by_status: dict[str, list[OwnedSettlementCompareResultRow]] = {
        "matched": [],
        "partial": [],
        "unsettled": [],
        "unmapped": [],
    }
    for row in all_rows:
        by_status.setdefault(row.status, []).append(row)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "요약"
    ws_summary.append(["항목", "값"])
    ws_summary.append(["연도", year])
    ws_summary.append(["전체", summary.total])
    ws_summary.append(["정산됨", summary.matched])
    ws_summary.append(["일부 정산", summary.partial])
    ws_summary.append(["미정산", summary.unsettled])
    ws_summary.append(["맵핑 없음", summary.unmapped])
    ws_summary.append(
        [
            "비교일시",
            summary.compared_at.isoformat() if summary.compared_at else None,
        ]
    )

    for status_key, sheet_name in _COMPARE_STATUS_SHEETS:
        ws = wb.create_sheet(sheet_name)
        _append_compare_rows(ws, by_status.get(status_key, []))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"owned_settlement_compare_{year}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        },
    )
