from uuid import UUID
from io import BytesIO
import json
import zipfile
from urllib.parse import quote
from decimal import Decimal
from datetime import datetime, timezone, date
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from fastapi_pagination import Page, Params
from fastapi_pagination.ext.sqlalchemy import apaginate
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import and_, exists, func, or_, collate
from sqlalchemy.orm import selectinload
from sqlalchemy.exc import SQLAlchemyError
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field

from app.database import User, get_async_session
from app.models import (
    Order,
    OrderItem,
    OrderMemo,
    Shipment,
    ShipmentItem,
    Receiver,
    Channel,
    ChannelOrderExcelMappingVersion,
    Product,
    ProductState,
    ProductAliasDict,
    ProductAliasItem,
    OrderHistory,
    OrderHistoryActionType,
    Settlement,
    SettlementHistory,
    SettlementState,
    SettlementHistoryActionType,
)
from app.models import OrderStatus
from app.schemas import (
    OrderRead,
    OrderCreate,
    OrderUpdate,
    OrderListRead,
    OrderItemRead,
    OrderItemListRead,
    ProductSummary,
    ReceiverSummary,
    ChannelSummary,
    OrderHistoryRead,
)
from app.schemas import (
    ExcelOrderPreviewResponse,
    ExcelOrderPreviewRowOut,
    ExcelOrderUploadItem,
    ExcelOrderUploadRequest,
    ExcelOrderUploadResponse,
)
from app.schemas import (
    ProductRead,
    PlaceOrderRequest,
    PlaceOrderResponse,
    PlaceOrderOrderShipments,
    PlaceOrderShipment,
    OrderMemoRead,
    OrderMemoCreate,
)
from app.users import current_active_user
from app.permissions import require_permission

router = APIRouter(tags=["order"])


def _append_excel_header(ws, headers: list[str]) -> None:
    header_fill = PatternFill(patternType="solid", fgColor="FFD9D9D9")  # light gray
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cells: list[WriteOnlyCell] = []
    for h in headers:
        c = WriteOnlyCell(ws, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        cells.append(c)
    ws.append(cells)

EXPORT_MAX_ROWS = 10_000
PLACE_ORDER_ALL_LIMIT = 1_000
ORDERS_EXCEL_MAX_ROWS = 10_000

# orders list sorting
ORDER_SORT_BY_VALUES = ("created_at", "order_date", "updated_at", "receiver_name")
ORDER_SORT_DIR_VALUES = ("asc", "desc")

# 채널별 raw 엑셀: raw 컬럼 뒤에 붙는 고정 메타 열
_RAW_EXPORT_META_HEADERS = (
    "주문일",
    "발주일",
    "배송일",
    "주문상태",
    "배송번호",
)

_SEOUL_TZ = ZoneInfo("Asia/Seoul")

def _decrypt_xlsx_bytes_if_needed(*, content: bytes, password: str) -> bytes:
    """
    Decrypt an Office-encrypted xlsx into plain xlsx bytes.
    Used only for order excel preview upload when password is provided.
    """
    try:
        import msoffcrypto  # type: ignore
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"엑셀 복호화 라이브러리가 설치되지 않았습니다: {e}",
        )

    try:
        of = msoffcrypto.OfficeFile(BytesIO(content))
        of.load_key(password=password)
        out = BytesIO()
        of.decrypt(out)
        return out.getvalue()
    except Exception:
        # Avoid leaking sensitive details.
        raise HTTPException(
            status_code=422,
            detail="엑셀 암호가 올바르지 않거나 파일을 복호화할 수 없습니다.",
        )


def _fmt_dt_excel(dt: datetime | None) -> str | None:
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(_SEOUL_TZ).strftime("%Y-%m-%d %H:%M")


def _order_status_label_ko(status: OrderStatus | None) -> str | None:
    if status is None:
        return None
    labels: dict[OrderStatus, str] = {
        OrderStatus.ORDER: "주문",
        OrderStatus.ORDER_PLACED: "발주",
        OrderStatus.SHIPPING_WAITING: "배송 대기",
        OrderStatus.SHIPPING: "배송",
        OrderStatus.CANCELLED: "취소",
    }
    return labels.get(status, status.value)


def _active_shipments_for_export(order: Order) -> list[Shipment]:
    rows = [
        s
        for s in (getattr(order, "shipments", None) or [])
        if not getattr(s, "is_delete", False)
    ]
    return sorted(rows, key=lambda s: (s.created_at, s.id))


def _raw_export_meta_cells(order: Order) -> list[object]:
    """배송 1:N이면 배송일·배송번호는 줄바꿈으로 한 셀에 나열."""
    ss = _active_shipments_for_export(order)
    order_date = _fmt_dt_excel(getattr(order, "order_date", None))
    placed = _fmt_dt_excel(getattr(order, "order_placed_date", None))
    status_ko = _order_status_label_ko(getattr(order, "status", None))
    if ss:
        ship_dates = "\n".join(_fmt_dt_excel(s.shipping_date) or "" for s in ss)
        invoices = "\n".join((s.invoice_number or "").strip() for s in ss)
    else:
        ship_dates = _fmt_dt_excel(getattr(order, "shipping_date", None)) or ""
        inv = (getattr(order, "invoice_number", None) or "").strip()
        invoices = inv
    return [
        order_date,
        placed,
        ship_dates.strip() and ship_dates or None,
        status_ko,
        invoices.strip() and invoices or None,
    ]


async def _load_mapping_by_version_id(
    db: AsyncSession, version_id
) -> dict | None:
    if version_id is None:
        return None
    res = await db.execute(
        select(ChannelOrderExcelMappingVersion).filter(
            ChannelOrderExcelMappingVersion.id == version_id
        )
    )
    v = res.scalars().first()
    if not v:
        return None
    m = getattr(v, "mapping", None)
    return m if isinstance(m, dict) else None

def _ser(v: object):
    if v is None:
        return None
    # UUID
    if hasattr(v, "hex"):
        return str(v)
    # datetime / date
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date):
        return v.isoformat()
    # Enum
    if hasattr(v, "value"):
        try:
            return getattr(v, "value")
        except Exception:
            return str(v)
    return v


def _stable_sort_keys(keys: list[str]) -> list[str]:
    """Extra columns should be stable across runs for diffability."""
    return sorted(keys, key=lambda s: (s is None, str(s)))


def _json_cell_for_excel(v: object) -> object:
    """Return a value openpyxl can write (str/float/int/bool/None)."""
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        return v
    try:
        return json.dumps(v, ensure_ascii=False, default=str)
    except Exception:
        return str(v)


def _build_headers_for_version(
    mapping: dict | None,
    raw_keys_union: set[str],
) -> list[str]:
    source_headers: list[str] = []
    if isinstance(mapping, dict):
        sh = mapping.get("source_headers")
        if isinstance(sh, list):
            source_headers = [str(x).strip() for x in sh if str(x).strip()]
    base = []
    seen = set()
    for h in source_headers:
        if h not in seen:
            base.append(h)
            seen.add(h)
    extras = [k for k in raw_keys_union if k and k not in seen]
    return base + _stable_sort_keys(extras)


def _order_snapshot(order: Order) -> dict:
    return {
        "id": _ser(order.id),
        "receiver_id": _ser(order.receiver_id),
        "channel_id": _ser(order.channel_id),
        "mall_product_name": _ser(getattr(order, "mall_product_name", None)),
        "price": float(order.price) if getattr(order, "price", None) is not None else None,
        "quantity": _ser(getattr(order, "quantity", None)),
        "memo": _ser(getattr(order, "memo", None)),
        "order_date": _ser(getattr(order, "order_date", None)),
        "invoice_number": _ser(getattr(order, "invoice_number", None)),
        "status": _ser(getattr(order, "status", None)),
        "order_placed_date": _ser(getattr(order, "order_placed_date", None)),
        "shipping_date": _ser(getattr(order, "shipping_date", None)),
        "items": [
            {"product_id": _ser(oi.product_id), "quantity": _ser(oi.quantity)}
            for oi in (getattr(order, "order_items", None) or [])
        ],
    }


async def _try_add_order_history(db: AsyncSession, h: OrderHistory) -> None:
    """
    이력 테이블 미생성/권한 이슈 등으로 주문 핵심 플로우가 깨지지 않도록,
    nested transaction(savepoint)로 이력 insert를 격리한다.
    """
    try:
        async with db.begin_nested():
            db.add(h)
            await db.flush()
    except SQLAlchemyError:
        # ignore (do not fail main transaction)
        return


async def _try_add_settlement_history(db: AsyncSession, h: SettlementHistory) -> None:
    try:
        async with db.begin_nested():
            db.add(h)
            await db.flush()
    except SQLAlchemyError:
        return


def _order_list_transform(rows):
    out = []
    for r in rows:
        base = OrderRead.model_validate(r)
        items: list[OrderItemListRead] = []
        for oi in getattr(r, "order_items", []) or []:
            items.append(
                OrderItemListRead(
                    product_id=oi.product_id,
                    quantity=oi.quantity,
                    product=ProductSummary.model_validate(oi.product) if oi.product else None,
                )
            )
        receiver = ReceiverSummary.model_validate(r.receiver) if r.receiver else None
        ch = getattr(r, "channel", None)
        if ch is not None:
            courier = getattr(ch, "courier", None)
            courier_name = courier.name if courier is not None else None
            courier_url = courier.url if courier is not None else None
            channel = ChannelSummary(
                id=ch.id,
                name=ch.name,
                courier_name=courier_name,
                courier_url=courier_url,
                url=getattr(ch, "url", None),
            )
        else:
            channel = None
        memo_count = int(getattr(r, "memo_count", 0) or 0)
        out.append(
            OrderListRead(
                **base.model_dump(exclude={"items"}),
                items=items,
                receiver=receiver,
                channel=channel,
                memo_count=memo_count,
            )
        )
    return out

def _norm(s: str) -> str:
    return "".join((s or "").strip().lower().split())

EXPECTED_HEADERS = {
    "채널": "channel",
    "수취인명": "receiver_name",
    "상품명": "product_name",
    "수량": "quantity",
    "총 주문금액": "total_price",
    "수취인연락처": "receiver_phone",
    "우편번호": "zip_code",
    "통합배송지": "integrated_address",
    "배송메세지": "shipping_message",
}


def _parse_channel_order_excel_mapping(
    mapping: dict | list | None,
) -> tuple[int, dict[str, str], list[str] | None, list[str]] | None:
    """
    Returns (header_row_1based, canonical_korean -> excel_header_string, source_headers, warnings).
    source_headers: 저장된 엑셀 헤더 전체(없으면 None).
    None if mapping is absent — caller uses legacy row-1 header and direct column names.
    알 수 없는 label·중복 label은 건너뛰고 warnings에만 적는다(예외 없음).
    """
    warnings: list[str] = []
    if not mapping or not isinstance(mapping, dict):
        return None
    source_headers: list[str] | None = None
    sh = mapping.get("source_headers")
    if isinstance(sh, list):
        source_headers = [_cell_to_str(x) for x in sh]
    header_row = mapping.get("header_row")
    try:
        hr = int(header_row) if header_row is not None else 1
    except (TypeError, ValueError):
        hr = 1
        warnings.append("주문 엑셀 매핑의 header_row 값이 올바르지 않아 1로 처리했습니다.")
    if hr < 1:
        hr = 1
        warnings.append("주문 엑셀 매핑의 header_row는 1 이상이어야 합니다. 1행으로 처리했습니다.")
    cols = mapping.get("columns")
    if not cols or not isinstance(cols, dict):
        return hr, {}, source_headers, warnings
    canonical_to_excel: dict[str, str] = {}
    for excel_name, meta in cols.items():
        if not isinstance(meta, dict):
            continue
        label = str(meta.get("label", "")).strip()
        if not label:
            continue
        if label not in EXPECTED_HEADERS:
            warnings.append(
                f"주문 엑셀 매핑에서 인식하지 않는 표준 필드가 제외되었습니다: 「{label}」(엑셀 열: {excel_name})"
            )
            continue
        if label in canonical_to_excel:
            warnings.append(
                f"주문 엑셀 매핑에서 표준 필드 「{label}」이(가) 중복되어 첫 매핑만 사용합니다."
            )
            continue
        ex_key = _cell_to_str(excel_name)
        if not ex_key:
            continue
        canonical_to_excel[label] = ex_key
    return hr, canonical_to_excel, source_headers, warnings


def _excel_header_and_data_iter(ws, header_row_1based: int):
    """Skip header_row_1based - 1 rows, read header row, return (data iterator, header_to_idx, headers_full)."""
    rows_iter = ws.iter_rows(values_only=True)
    skip = max(0, header_row_1based - 1)
    for _ in range(skip):
        try:
            next(rows_iter)
        except StopIteration:
            raise HTTPException(
                status_code=422,
                detail=f"Excel has fewer rows than header_row ({header_row_1based})",
            )
    try:
        header_cells = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=422, detail="Empty excel file")
    headers = [_cell_to_str(h) for h in header_cells]
    header_to_idx = {h: i for i, h in enumerate(headers) if h}
    return rows_iter, header_to_idx, headers


def _json_cell(v: object) -> object:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, bool | int | float | str):
        return v
    return str(v)


def _build_raw_row_from_excel(headers_full: list[str], excel_row: tuple | None) -> dict[str, object]:
    """엑셀 원본 헤더 문자열을 키로 한 행 전체(JSON 직렬화 가능 값). 빈 헤더 열은 __col_{n} 키 사용."""
    raw: dict[str, object] = {}
    row = excel_row or ()
    n = max(len(headers_full), len(row))
    used: dict[str, int] = {}
    for i in range(n):
        header = headers_full[i] if i < len(headers_full) else ""
        base = header if header else f"__col_{i + 1}"
        c = used.get(base, 0) + 1
        used[base] = c
        key = base if c == 1 else f"{base}__{c}"
        val = row[i] if i < len(row) else None
        raw[key] = _json_cell(val)
    return raw


def _get_raw_field(
    raw: dict,
    canonical: str,
    canonical_to_excel: dict[str, str] | None,
) -> object | None:
    """업로드 raw는 원본 엑셀 헤더 키이므로, 채널 매핑이 있으면 그 열 이름으로 조회한다."""
    if canonical_to_excel:
        ex = canonical_to_excel.get(canonical)
        if ex:
            if ex in raw:
                return raw.get(ex)
            return None
    return raw.get(canonical)


def _cell_to_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()

def _cell_to_int(v) -> int:
    if v is None or str(v).strip() == "":
        raise ValueError("empty")
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    s = str(v).strip().replace(",", "")
    return int(s)

def _cell_to_decimal(v) -> Decimal:
    if v is None or str(v).strip() == "":
        raise ValueError("empty")
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).strip().replace(",", "")
    return Decimal(s)


class OrdersExcelRequest(BaseModel):
    order_ids: list[UUID] = Field(default_factory=list, max_length=1000)


@router.get("/", response_model=Page[OrderListRead])
async def list_orders(
    page: int = 1,
    size: int = 10,
    receiver_id: UUID | None = None,
    status: OrderStatus | None = None,
    sort_by: str | None = Query(default=None, description="created_at|order_date|updated_at|receiver_name"),
    sort_dir: str | None = Query(default="desc", description="asc|desc"),
    channel_id: UUID | None = Query(default=None),
    channel_ids: str | None = Query(default=None, description="Comma-separated channel ids"),
    channel_name: str | None = Query(default=None),
    receiver_name: str | None = Query(default=None),
    receiver_phone: str | None = Query(default=None),
    receiver_zip_code: str | None = Query(default=None),
    receiver_address: str | None = Query(default=None),
    invoice_number: str | None = Query(default=None),
    product_query: str | None = Query(default=None),
    order_date_start: date | None = Query(default=None),
    order_date_end: date | None = Query(default=None),
    has_memos: bool | None = Query(default=None),
    db: AsyncSession = Depends(get_async_session),
    user: User = Depends(current_active_user),
):
    _ = await require_permission("orders", "read")(user, db)
    params = Params(page=page, size=size)
    sort_by_norm = (sort_by or "").strip()
    sort_dir_norm = (sort_dir or "desc").strip().lower()
    if sort_dir_norm not in ORDER_SORT_DIR_VALUES:
        raise HTTPException(status_code=422, detail=f"Invalid sort_dir: {sort_dir}")
    if sort_by_norm and sort_by_norm not in ORDER_SORT_BY_VALUES:
        raise HTTPException(status_code=422, detail=f"Invalid sort_by: {sort_by}")

    # Default sort: created_at desc (stable with id)
    if not sort_by_norm:
        sort_by_norm = "created_at"

    is_asc = sort_dir_norm == "asc"
    query = (
        select(Order)
        .filter(Order.is_delete == False)
        .options(
            selectinload(Order.order_items).selectinload(OrderItem.product),
            selectinload(Order.receiver),
            selectinload(Order.channel).selectinload(Channel.courier),
        )
    )
    query = _apply_order_filters(
        query,
        receiver_id=receiver_id,
        status=status,
        channel_id=channel_id,
        channel_ids=channel_ids,
        channel_name=channel_name,
        receiver_name=receiver_name,
        receiver_phone=receiver_phone,
        receiver_zip_code=receiver_zip_code,
        receiver_address=receiver_address,
        invoice_number=invoice_number,
        product_query=product_query,
        order_date_start=order_date_start,
        order_date_end=order_date_end,
        has_memos=has_memos,
        force_receiver_join=(sort_by_norm == "receiver_name"),
    )
    if sort_by_norm == "receiver_name":
        name_expr = collate(func.trim(Receiver.name), "ko-KR-x-icu")
        query = query.order_by(name_expr.asc() if is_asc else name_expr.desc(), Order.id.asc())
    elif sort_by_norm == "created_at":
        query = query.order_by(
            Order.created_at.asc() if is_asc else Order.created_at.desc(),
            Order.id.desc(),
        )
    elif sort_by_norm == "updated_at":
        query = query.order_by(
            Order.updated_at.asc() if is_asc else Order.updated_at.desc(),
            Order.id.desc(),
        )
    else:
        # order_date (default)
        query = query.order_by(
            Order.order_date.asc() if is_asc else Order.order_date.desc(),
            Order.id.desc(),
        )
    return await apaginate(db, query, params, transformer=_order_list_transform)


@router.get("/order/all", response_model=list[OrderListRead])
async def list_all_order_status_orders(
    db: AsyncSession = Depends(get_async_session),
    user: User = Depends(current_active_user),
):
    """
    발주(주문 → 발주) 기능 전용: ORDER 상태 주문을 한 번에 조회한다.
    - 화면 페이지네이션 반복 호출을 금지하고, 서버에서 최대 1,000건으로 제한한다.
    """
    _ = await require_permission("orders", "read")(user, db)

    q = (
        select(Order)
        .filter(
            Order.is_delete == False,
            Order.status == OrderStatus.ORDER,
        )
        .options(
            selectinload(Order.order_items).selectinload(OrderItem.product),
            selectinload(Order.receiver),
            selectinload(Order.channel).selectinload(Channel.courier),
        )
        .order_by(Order.order_date.desc(), Order.id.desc())
        .limit(PLACE_ORDER_ALL_LIMIT + 1)
    )
    result = await db.execute(q)
    orders = result.scalars().unique().all()
    if len(orders) > PLACE_ORDER_ALL_LIMIT:
        raise HTTPException(
            status_code=422,
            detail=f"발주 대상이 너무 많습니다. ({PLACE_ORDER_ALL_LIMIT}건 초과) 주문 조건을 줄여 주세요.",
        )
    return _order_list_transform(orders)


@router.post("/export/excel")
async def download_orders_excel(
    payload: OrdersExcelRequest | None = None,
    order_date_start: date | None = Query(default=None),
    order_date_end: date | None = Query(default=None),
    status: OrderStatus | None = Query(default=None),
    channel_id: UUID | None = Query(default=None),
    channel_name: str | None = Query(default=None),
    receiver_name: str | None = Query(default=None),
    receiver_phone: str | None = Query(default=None),
    receiver_address: str | None = Query(default=None),
    invoice_number: str | None = Query(default=None),
    product_query: str | None = Query(default=None),
    has_memos: bool | None = Query(default=None),
    db: AsyncSession = Depends(get_async_session),
    user: User = Depends(current_active_user),
):
    """
    주문 엑셀 다운로드(xlsx).
    - payload.order_ids가 있으면 해당 주문만 다운로드
    - 없으면 query(filter) 조건으로 조회해 다운로드 (최대 10,000건)
    """
    _ = await require_permission("orders", "read")(user, db)

    q = (
        select(Order)
        .filter(Order.is_delete == False)
        .options(
            selectinload(Order.order_items).selectinload(OrderItem.product),
            selectinload(Order.receiver),
            selectinload(Order.channel).selectinload(Channel.courier),
        )
    )

    ids = list((payload.order_ids if payload else []) or [])
    if ids:
        q = q.filter(Order.id.in_(ids))
    else:
        q = _apply_order_filters(
            q,
            status=status,
            channel_id=channel_id,
            channel_name=channel_name,
            receiver_name=receiver_name,
            receiver_phone=receiver_phone,
            receiver_address=receiver_address,
            invoice_number=invoice_number,
            product_query=product_query,
            order_date_start=order_date_start,
            order_date_end=order_date_end,
            has_memos=has_memos,
        )
        q = q.order_by(Order.order_date.desc(), Order.id.desc()).limit(ORDERS_EXCEL_MAX_ROWS + 1)

    res = await db.execute(q)
    orders = res.scalars().unique().all()
    if not orders:
        raise HTTPException(status_code=404, detail="다운로드할 주문이 없습니다.")
    if not ids and len(orders) > ORDERS_EXCEL_MAX_ROWS:
        raise HTTPException(
            status_code=422,
            detail=f"다운로드 대상이 너무 많습니다. (최대 {ORDERS_EXCEL_MAX_ROWS}건) 검색 조건을 좁혀 주세요.",
        )

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("orders")
    headers = [
        "주문 ID",
        "주문일",
        "주문 상태",
        "배송번호",
        "채널",
        "택배사",
        "상품",
        "수취인",
        "연락처",
        "우편번호",
        "주소",
        "금액",
        "수량",
        "메모",
    ]
    _append_excel_header(ws, headers)

    for o in orders:
        r = getattr(o, "receiver", None)
        ch = getattr(o, "channel", None)
        courier_name = getattr(ch, "courier_name", None) if ch is not None else None
        if courier_name is None and ch is not None:
            courier = getattr(ch, "courier", None)
            courier_name = getattr(courier, "name", None) if courier is not None else None
        addr = ""
        if r is not None:
            addr = " ".join([x for x in [getattr(r, "address", None), getattr(r, "address_detail", None)] if x])

        products = []
        for oi in getattr(o, "order_items", None) or []:
            p = getattr(oi, "product", None)
            label = f"[{getattr(p, 'product_code', '')}] {getattr(p, 'name', '')}".strip() if p is not None else str(getattr(oi, "product_id", ""))
            products.append(f"{label} x {getattr(oi, 'quantity', 0)}")
        products_cell = " | ".join(products)

        ws.append(
            [
                str(o.id),
                _fmt_dt_excel(getattr(o, "order_date", None)),
                _order_status_label_ko(getattr(o, "status", None)) or "",
                getattr(o, "invoice_number", None) or "",
                getattr(ch, "name", None) or "",
                courier_name or "",
                products_cell,
                getattr(r, "name", None) if r is not None else "",
                getattr(r, "phone", None) if r is not None else "",
                getattr(r, "zip_code", None) if r is not None else "",
                addr,
                float(getattr(o, "price", 0) or 0),
                int(getattr(o, "quantity", 0) or 0),
                getattr(o, "memo", None) or "",
            ]
        )

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename_utf8 = f"orders_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    filename_ascii = f"orders_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    disposition = (
        f'attachment; filename="{filename_ascii}"; '
        f"filename*=UTF-8''{quote(filename_utf8)}"
    )
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": disposition},
    )


def _apply_order_filters(
    query,
    *,
    receiver_id: UUID | None = None,
    status: OrderStatus | None = None,
    channel_id: UUID | None = None,
    channel_ids: str | None = None,
    channel_name: str | None = None,
    receiver_name: str | None = None,
    receiver_phone: str | None = None,
    receiver_zip_code: str | None = None,
    receiver_address: str | None = None,
    invoice_number: str | None = None,
    product_query: str | None = None,
    order_date_start: date | None = None,
    order_date_end: date | None = None,
    has_memos: bool | None = None,
    force_receiver_join: bool = False,
):
    if receiver_id is not None:
        query = query.filter(Order.receiver_id == receiver_id)
    if status is not None:
        query = query.filter(Order.status == status)
    # channel filter (multi > single)
    if channel_ids is not None and str(channel_ids).strip():
        tokens = [t.strip() for t in str(channel_ids).split(",") if t.strip()]
        parsed: list[UUID] = []
        for t in tokens:
            try:
                parsed.append(UUID(t))
            except Exception:
                raise HTTPException(status_code=422, detail="Invalid channel_ids")
        if parsed:
            query = query.filter(Order.channel_id.in_(parsed))
    elif channel_id is not None:
        query = query.filter(Order.channel_id == channel_id)

    if channel_name is not None and channel_name.strip():
        query = query.join(Channel, Channel.id == Order.channel_id).filter(
            Channel.name.ilike(f"%{channel_name.strip()}%")
        )

    receiver_join_needed = force_receiver_join or any(
        bool(x and str(x).strip())
        for x in [receiver_name, receiver_phone, receiver_zip_code, receiver_address]
    )
    if receiver_join_needed:
        query = query.join(Receiver, Receiver.id == Order.receiver_id)

    if receiver_name is not None and receiver_name.strip():
        query = query.filter(Receiver.name.ilike(f"%{receiver_name.strip()}%"))

    if receiver_phone is not None and receiver_phone.strip():
        query = query.filter(Receiver.phone.ilike(f"%{receiver_phone.strip()}%"))

    if receiver_zip_code is not None and receiver_zip_code.strip():
        query = query.filter(Receiver.zip_code.ilike(f"%{receiver_zip_code.strip()}%"))

    if receiver_address is not None and receiver_address.strip():
        s = receiver_address.strip()
        query = query.filter(
            or_(
                Receiver.address.ilike(f"%{s}%"),
                Receiver.address_detail.ilike(f"%{s}%"),
            )
        )

    if invoice_number is not None and invoice_number.strip():
        query = query.filter(Order.invoice_number.ilike(f"%{invoice_number.strip()}%"))

    if product_query is not None and product_query.strip():
        pq = product_query.strip()
        query = (
            query.join(OrderItem, OrderItem.order_id == Order.id)
            .join(Product, Product.id == OrderItem.product_id)
            .filter(
                or_(
                    Product.product_code.ilike(f"%{pq}%"),
                    Product.name.ilike(f"%{pq}%"),
                )
            )
            .distinct()
        )

    if order_date_start is not None:
        query = query.filter(func.date(Order.order_date) >= order_date_start)
    if order_date_end is not None:
        query = query.filter(func.date(Order.order_date) <= order_date_end)

    if has_memos is True:
        query = query.where(
            exists().where(
                and_(OrderMemo.order_id == Order.id, OrderMemo.is_delete == False)
            )
        )
    return query


@router.get("/order-placed/count")
async def count_order_placed_orders(
    db: AsyncSession = Depends(get_async_session),
    user: User = Depends(current_active_user),
):
    _ = await require_permission("orders", "read")(user, db)
    result = await db.execute(
        select(func.count())
        .select_from(Order)
        .where(Order.is_delete == False, Order.status == OrderStatus.ORDER_PLACED)
    )
    return {"count": int(result.scalar() or 0)}


@router.get("/histories", response_model=Page[OrderHistoryRead])
async def list_order_histories(
    page: int = 1,
    size: int = 10,
    order_id: UUID | None = None,
    action_type: OrderHistoryActionType | None = None,
    update_user_id: UUID | None = None,
    from_status: OrderStatus | None = None,
    to_status: OrderStatus | None = None,
    date_start: date | None = Query(default=None, description="created_at 기준 (YYYY-MM-DD)"),
    date_end: date | None = Query(default=None, description="created_at 기준 (YYYY-MM-DD)"),
    db: AsyncSession = Depends(get_async_session),
    user: User = Depends(current_active_user),
):
    _ = await require_permission("orders", "read")(user, db)
    params = Params(page=page, size=size)
    q = select(OrderHistory).where(OrderHistory.is_delete == False)
    if order_id is not None:
        q = q.where(OrderHistory.order_id == order_id)
    if action_type is not None:
        q = q.where(OrderHistory.action_type == action_type)
    if update_user_id is not None:
        q = q.where(OrderHistory.update_user_id == update_user_id)
    if from_status is not None:
        q = q.where(OrderHistory.from_status == from_status)
    if to_status is not None:
        q = q.where(OrderHistory.to_status == to_status)
    if date_start is not None:
        q = q.where(func.date(OrderHistory.created_at) >= date_start)
    if date_end is not None:
        q = q.where(func.date(OrderHistory.created_at) <= date_end)
    q = q.order_by(OrderHistory.created_at.desc(), OrderHistory.id.desc())
    return await apaginate(db, q, params)


@router.get("/{order_id}/histories", response_model=Page[OrderHistoryRead])
async def list_order_histories_by_order(
    order_id: UUID,
    page: int = 1,
    size: int = 10,
    db: AsyncSession = Depends(get_async_session),
    user: User = Depends(current_active_user),
):
    _ = await require_permission("orders", "read")(user, db)
    params = Params(page=page, size=size)
    q = (
        select(OrderHistory)
        .where(
            OrderHistory.is_delete == False,
            OrderHistory.order_id == order_id,
        )
        .order_by(OrderHistory.created_at.desc(), OrderHistory.id.desc())
    )
    return await apaginate(db, q, params)


def _user_display(u: User) -> str:
    email = getattr(u, "email", None)
    if email:
        return str(email)
    return str(u.id)


@router.get("/{order_id}/memos", response_model=list[OrderMemoRead])
async def list_order_memos(
    order_id: UUID,
    db: AsyncSession = Depends(get_async_session),
    user: User = Depends(current_active_user),
):
    _ = await require_permission("orders", "read")(user, db)
    exists = (
        await db.execute(select(Order.id).where(Order.id == order_id, Order.is_delete == False))
    ).scalar_one_or_none()
    if exists is None:
        raise HTTPException(status_code=404, detail="Order not found")

    mres = await db.execute(
        select(OrderMemo)
        .where(OrderMemo.order_id == order_id, OrderMemo.is_delete == False)
        .options(selectinload(OrderMemo.user))
        .order_by(OrderMemo.created_at.asc())
    )
    rows = mres.scalars().all()
    out: list[OrderMemoRead] = []
    for m in rows:
        u = m.user
        out.append(
            OrderMemoRead(
                id=m.id,
                order_id=m.order_id,
                user_id=m.user_id,
                user_display=_user_display(u) if u is not None else str(m.user_id),
                content=m.content,
                created_at=m.created_at,
            )
        )
    return out


@router.post("/{order_id}/memos", response_model=OrderMemoRead)
async def create_order_memo(
    order_id: UUID,
    data: OrderMemoCreate,
    db: AsyncSession = Depends(get_async_session),
    user: User = Depends(current_active_user),
):
    _ = await require_permission("orders", "update")(user, db)
    ores = await db.execute(
        select(Order).where(Order.id == order_id, Order.is_delete == False)
    )
    order = ores.scalars().first()
    if order is None:
        raise HTTPException(status_code=404, detail="Order not found")

    content = data.content.strip()
    if not content:
        raise HTTPException(status_code=400, detail="Memo content is empty")

    memo = OrderMemo(order_id=order_id, user_id=user.id, content=content)
    db.add(memo)
    await db.commit()
    mres = await db.execute(
        select(OrderMemo)
        .where(OrderMemo.id == memo.id)
        .options(selectinload(OrderMemo.user))
    )
    m = mres.scalars().first()
    if m is None:
        raise HTTPException(status_code=500, detail="Failed to load memo")
    u = m.user
    return OrderMemoRead(
        id=m.id,
        order_id=m.order_id,
        user_id=m.user_id,
        user_display=_user_display(u) if u is not None else str(m.user_id),
        content=m.content,
        created_at=m.created_at,
    )


@router.get("/{order_id}", response_model=OrderRead)
async def get_order(
    order_id: UUID,
    db: AsyncSession = Depends(get_async_session),
    user: User = Depends(current_active_user),
):
    _ = await require_permission("orders", "read")(user, db)
    result = await db.execute(
        select(Order)
        .filter(Order.id == order_id, Order.is_delete == False)
        .options(selectinload(Order.order_items).selectinload(OrderItem.product))
    )
    order = result.scalars().first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    return order


@router.post("/", response_model=OrderRead)
async def create_order(
    data: OrderCreate,
    db: AsyncSession = Depends(get_async_session),
    user: User = Depends(current_active_user),
):
    _ = await require_permission("orders", "create")(user, db)
    items = data.items
    total_quantity = sum(i.quantity for i in items)
    order = Order(
        receiver_id=data.receiver_id,
        channel_id=data.channel_id,
        mall_product_name=data.mall_product_name,
        raw=data.raw,
        price=data.price,
        quantity=total_quantity,
        memo=data.memo,
        order_date=data.order_date,
        status=data.status,
        invoice_number=data.invoice_number,
        update_user_id=user.id,
    )
    db.add(order)
    await db.flush()

    for item in items:
        oi = OrderItem(
            order_id=order.id,
            product_id=item.product_id,
            quantity=item.quantity,
        )
        db.add(oi)
    db.add(order)
    await db.commit()
    await db.refresh(order)

    await _try_add_order_history(
        db,
        OrderHistory(
            order_id=order.id,
            action_type=OrderHistoryActionType.CREATED,
            update_user_id=user.id,
            from_status=None,
            to_status=order.status,
            before_update=None,
            after_update=_order_snapshot(order),
            reason="create_order",
        ),
    )
    return order


@router.put("/{order_id}", response_model=OrderRead)
async def update_order(
    order_id: UUID,
    data: OrderUpdate,
    db: AsyncSession = Depends(get_async_session),
    user: User = Depends(current_active_user),
):
    _ = await require_permission("orders", "update")(user, db)
    result = await db.execute(
        select(Order)
        .filter(Order.id == order_id, Order.is_delete == False)
        .options(selectinload(Order.order_items))
    )
    order = result.scalars().first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    before = _order_snapshot(order)
    from_status = order.status
    updates = data.model_dump(exclude_unset=True)

    items_payload = updates.pop("items", None)

    for key, value in updates.items():
        setattr(order, key, value)

    if order.status == OrderStatus.SHIPPING and order.shipping_date is None:
        order.shipping_date = datetime.now(timezone.utc)

    if items_payload is not None:
        # 기존 order_items 제거 후 재생성
        for oi in list(order.order_items or []):
            await db.delete(oi)
        await db.flush()

        total_quantity = 0
        for item in items_payload:
            total_quantity += item["quantity"]
            oi = OrderItem(
                order_id=order.id,
                product_id=item["product_id"],
                quantity=item["quantity"],
            )
            db.add(oi)
        order.quantity = total_quantity
    order.update_user_id = user.id

    # 주문 취소 시 정산도 취소 상태로 전환
    if from_status != OrderStatus.CANCELLED and order.status == OrderStatus.CANCELLED:
        sres = await db.execute(select(Settlement).where(Settlement.order_id == order.id))
        st = sres.scalars().first()
        if st is not None and st.state != SettlementState.CANCELLED:
            from_state = st.state
            before_price = st.price
            st.state = SettlementState.CANCELLED
            st.update_user_id = user.id
            await _try_add_settlement_history(
                db,
                SettlementHistory(
                    settlement_id=st.id,
                    order_id=order.id,
                    action_type=SettlementHistoryActionType.CANCELLED_BY_ORDER,
                    update_user_id=user.id,
                    reason="주문 취소 연동",
                    from_state=from_state,
                    to_state=st.state,
                    before_price=before_price,
                    after_price=st.price,
                ),
            )

    await db.commit()
    await db.refresh(order)

    after = _order_snapshot(order)
    to_status = order.status
    status_changed = from_status != to_status
    if status_changed:
        if to_status == OrderStatus.CANCELLED:
            action = OrderHistoryActionType.CANCELLED
        else:
            action = OrderHistoryActionType.STATUS_CHANGED
    else:
        action = OrderHistoryActionType.UPDATED

    await _try_add_order_history(
        db,
        OrderHistory(
            order_id=order.id,
            action_type=action,
            update_user_id=user.id,
            from_status=from_status if status_changed else None,
            to_status=to_status if status_changed else None,
            before_update=before,
            after_update=after,
            reason="update_order",
        ),
    )
    return order


@router.delete("/{order_id}")
async def delete_order(
    order_id: UUID,
    db: AsyncSession = Depends(get_async_session),
    user: User = Depends(current_active_user),
):
    _ = await require_permission("orders", "delete")(user, db)
    result = await db.execute(select(Order).filter(Order.id == order_id, Order.is_delete == False))
    order = result.scalars().first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    order.is_delete = True
    await db.commit()
    await _try_add_order_history(
        db,
        OrderHistory(
            order_id=order.id,
            action_type=OrderHistoryActionType.DELETED,
            update_user_id=user.id,
            from_status=order.status,
            to_status=None,
            before_update=_order_snapshot(order),
            after_update=None,
            reason="delete_order",
        ),
    )
    return {"message": "Order successfully deleted"}


def _validate_order_items_match(
    order_items: list, order_id: UUID, shipments: list[PlaceOrderShipment]
) -> None:
    """주문 상품 합계와 발주 상품 합계가 일치하는지 검사."""
    from collections import Counter
    order_totals: Counter = Counter()
    for oi in order_items:
        order_totals[oi.product_id] += oi.quantity
    ship_totals: Counter = Counter()
    for sh in shipments:
        for it in sh.items:
            if it.quantity <= 0:
                raise HTTPException(
                    status_code=422,
                    detail=f"주문 {order_id}: 수량은 1 이상이어야 합니다.",
                )
            ship_totals[it.product_id] += it.quantity
    if order_totals != ship_totals:
        raise HTTPException(
            status_code=422,
            detail=f"주문 {order_id}: 주문 상품/수량 합계와 발주 할당이 일치하지 않습니다.",
        )


def _validate_consolidated_items_match(
    rep_order: Order,
    sub_orders: list[Order],
    shipments: list[PlaceOrderShipment],
) -> None:
    """합배송 그룹(rep + subs) 상품 합계와 발주 상품 합계가 일치하는지 검사."""
    from collections import Counter
    order_totals: Counter = Counter()
    for o in [rep_order, *sub_orders]:
        for oi in o.order_items or []:
            order_totals[oi.product_id] += oi.quantity
    ship_totals: Counter = Counter()
    for sh in shipments:
        for it in sh.items:
            if it.quantity <= 0:
                raise HTTPException(
                    status_code=422,
                    detail=f"주문 {rep_order.id}: 수량은 1 이상이어야 합니다.",
                )
            ship_totals[it.product_id] += it.quantity
    if order_totals != ship_totals:
        raise HTTPException(
            status_code=422,
            detail=(
                f"주문 {rep_order.id}(합배송): 합배송 그룹의 상품/수량 합계와 "
                "발주 할당이 일치하지 않습니다."
            ),
        )


@router.post("/place-order", response_model=PlaceOrderResponse)
async def place_order(
    payload: PlaceOrderRequest,
    db: AsyncSession = Depends(get_async_session),
    user: User = Depends(current_active_user),
):
    _ = await require_permission("orders", "update")(user, db)
    """ORDER 상태의 주문을 ORDER_PLACED로 변경하고, 지정한 대로 배송(shipment)을 생성한다."""
    use_shipments = (
        payload.order_shipments is not None
        and len(payload.order_shipments) > 0
    )
    if use_shipments:
        rep_order_ids: list[UUID] = [os.order_id for os in payload.order_shipments]
        sub_order_ids: list[UUID] = []
        seen_sub: set[UUID] = set()
        for os in payload.order_shipments:
            for sid in os.consolidated_sub_order_ids or []:
                if sid == os.order_id:
                    raise HTTPException(
                        status_code=422,
                        detail=f"대표 주문과 하위 주문이 같을 수 없습니다: {sid}",
                    )
                if sid in seen_sub:
                    raise HTTPException(
                        status_code=422,
                        detail=f"하위 주문이 여러 합배송 그룹에 중복되어 있습니다: {sid}",
                    )
                if sid in set(rep_order_ids):
                    raise HTTPException(
                        status_code=422,
                        detail=f"하위 주문이 다른 그룹의 대표 주문으로도 지정되었습니다: {sid}",
                    )
                seen_sub.add(sid)
                sub_order_ids.append(sid)
        order_ids = [*rep_order_ids, *sub_order_ids]
    else:
        order_ids = list(payload.order_ids or [])
    if not order_ids:
        return PlaceOrderResponse(updated_count=0, shipments_created=0)

    result = await db.execute(
        select(Order)
        .filter(
            Order.id.in_(order_ids),
            Order.is_delete == False,
            Order.status == OrderStatus.ORDER,
        )
        .options(selectinload(Order.order_items), selectinload(Order.receiver))
    )
    orders = result.scalars().unique().all()
    order_by_id = {o.id: o for o in orders}
    if len(orders) != len(order_ids):
        raise HTTPException(
            status_code=422,
            detail="일부 주문을 찾을 수 없거나 이미 발주/배송 상태입니다. ORDER 상태의 주문만 선택해 주세요.",
        )

    updated_count = 0
    shipments_created = 0

    if use_shipments:
        for os in payload.order_shipments:
            order = order_by_id.get(os.order_id)
            if not order:
                raise HTTPException(status_code=422, detail=f"주문을 찾을 수 없습니다: {os.order_id}")
            if not os.shipments:
                raise HTTPException(
                    status_code=422,
                    detail=f"주문 {order.id}에 최소 1건의 발주가 필요합니다.",
                )

            sub_orders: list[Order] = []
            for sid in os.consolidated_sub_order_ids or []:
                sub = order_by_id.get(sid)
                if not sub:
                    raise HTTPException(
                        status_code=422,
                        detail=f"하위 주문을 찾을 수 없습니다: {sid}",
                    )
                sub_orders.append(sub)

            if sub_orders:
                _validate_consolidated_items_match(order, sub_orders, os.shipments)
            else:
                _validate_order_items_match(
                    order.order_items or [],
                    order.id,
                    os.shipments,
                )

            before = _order_snapshot(order)
            from_status = order.status
            order.status = OrderStatus.ORDER_PLACED
            if order.order_placed_date is None:
                order.order_placed_date = datetime.now(timezone.utc)
            order.update_user_id = user.id
            updated_count += 1
            await _try_add_order_history(
                db,
                OrderHistory(
                    order_id=order.id,
                    action_type=OrderHistoryActionType.PLACED,
                    update_user_id=user.id,
                    from_status=from_status,
                    to_status=order.status,
                    before_update=before,
                    after_update=_order_snapshot(order),
                    reason="place_order",
                ),
            )

            for sub in sub_orders:
                sub_before = _order_snapshot(sub)
                sub_from = sub.status
                sub.status = OrderStatus.ORDER_PLACED
                sub.order_placed_date = order.order_placed_date
                sub.consolidated_to_order_id = order.id
                sub.update_user_id = user.id
                updated_count += 1
                await _try_add_order_history(
                    db,
                    OrderHistory(
                        order_id=sub.id,
                        action_type=OrderHistoryActionType.PLACED,
                        update_user_id=user.id,
                        from_status=sub_from,
                        to_status=sub.status,
                        before_update=sub_before,
                        after_update=_order_snapshot(sub),
                        reason="place_order(consolidated)",
                    ),
                )

            for sh in os.shipments:
                shipment = Shipment(
                    order_id=order.id,
                    invoice_number=order.invoice_number,
                    receiver_id=order.receiver_id,
                    order_placed_date=order.order_placed_date,
                    shipping_date=None,
                )
                db.add(shipment)
                await db.flush()
                for it in sh.items:
                    si = ShipmentItem(
                        shipment_id=shipment.id,
                        product_id=it.product_id,
                        quantity=it.quantity,
                    )
                    db.add(si)
                shipments_created += 1
    else:
        for order in orders:
            before = _order_snapshot(order)
            from_status = order.status
            order.status = OrderStatus.ORDER_PLACED
            if order.order_placed_date is None:
                order.order_placed_date = datetime.now(timezone.utc)
            order.update_user_id = user.id
            updated_count += 1
            await _try_add_order_history(
                db,
                OrderHistory(
                    order_id=order.id,
                    action_type=OrderHistoryActionType.PLACED,
                    update_user_id=user.id,
                    from_status=from_status,
                    to_status=order.status,
                    before_update=before,
                    after_update=_order_snapshot(order),
                    reason="place_order",
                ),
            )
            shipment = Shipment(
                order_id=order.id,
                invoice_number=order.invoice_number,
                receiver_id=order.receiver_id,
                order_placed_date=order.order_placed_date,
                shipping_date=None,
            )
            db.add(shipment)
            await db.flush()
            for oi in order.order_items or []:
                si = ShipmentItem(
                    shipment_id=shipment.id,
                    product_id=oi.product_id,
                    quantity=oi.quantity,
                )
                db.add(si)
            shipments_created += 1

    await db.commit()
    return PlaceOrderResponse(updated_count=updated_count, shipments_created=shipments_created)


@router.post("/excel/preview", response_model=ExcelOrderPreviewResponse)
async def preview_excel_orders(
    file: UploadFile = File(...),
    channel_id: UUID | None = Form(default=None),
    password: str | None = Form(default=None),
    db: AsyncSession = Depends(get_async_session),
    user: User = Depends(current_active_user),
):
    _ = await require_permission("orders", "create")(user, db)
    if not file.filename:
        raise HTTPException(status_code=422, detail="No file provided")

    content = await file.read()
    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as e:
        pw = str(password).strip() if password is not None else ""
        if not pw:
            # If no password provided, keep the original behavior.
            raise HTTPException(status_code=422, detail=f"Invalid excel file: {e}")
        decrypted = _decrypt_xlsx_bytes_if_needed(content=content, password=pw)
        try:
            wb = load_workbook(filename=BytesIO(decrypted), data_only=True)
        except Exception as e2:
            raise HTTPException(
                status_code=422, detail=f"Invalid excel file after decrypt: {e2}"
            )

    ws = wb.active

    selected_channel = None
    selected_channel_name: str | None = None
    if channel_id is not None:
        ch_res = await db.execute(
            select(Channel)
            .filter(Channel.id == channel_id, Channel.is_delete == False)
            .options(selectinload(Channel.current_mapping_version))
        )
        selected_channel = ch_res.scalars().first()
        if not selected_channel:
            raise HTTPException(status_code=422, detail=f"Unknown channel_id: {channel_id}")
        selected_channel_name = selected_channel.name

    preview_warnings: list[str] = []
    mapping_parsed: tuple[int, dict[str, str]] | None = None
    if selected_channel is not None and getattr(selected_channel, "current_mapping_version", None):
        parsed = _parse_channel_order_excel_mapping(selected_channel.current_mapping_version.mapping)
        if parsed is not None:
            hr, c_map, _src_h, w = parsed
            preview_warnings.extend(w)
            mapping_parsed = (hr, c_map)

    canonical_to_excel: dict[str, str] | None
    headers_full: list[str]
    if mapping_parsed is None:
        rows_iter = ws.iter_rows(values_only=True)
        try:
            first_row = next(rows_iter)
        except StopIteration:
            raise HTTPException(status_code=422, detail="Empty excel file")
        headers_full = [_cell_to_str(h) for h in first_row]
        header_to_idx = {h: i for i, h in enumerate(headers_full) if h}
        canonical_to_excel = None
    else:
        hr, c_map = mapping_parsed
        rows_iter, header_to_idx, headers_full = _excel_header_and_data_iter(ws, hr)
        canonical_to_excel = c_map if c_map else None

    required_headers = list(EXPECTED_HEADERS.keys())
    # 채널은 UI 선택값으로 대체 가능
    if selected_channel_name is not None and "채널" in required_headers:
        required_headers.remove("채널")

    if canonical_to_excel:
        missing = [
            h
            for h in required_headers
            if not canonical_to_excel.get(h) or canonical_to_excel.get(h) not in header_to_idx
        ]
    else:
        missing = [h for h in required_headers if h not in header_to_idx]
    if missing:
        preview_warnings.append(
            "엑셀에서 다음 필수 열을 찾지 못했습니다. 해당 값은 비어 있을 수 있습니다: "
            + ", ".join(missing)
        )

    # ACTIVE 상품 목록 (프론트에서 수동 매칭 셀렉트 용도)
    products_result = await db.execute(
        select(Product)
        .filter(Product.state == ProductState.ACTIVE, Product.is_delete == False)
    )
    active_products = products_result.scalars().all()
    product_by_id = {p.id: p for p in active_products}

    # 상품 별칭: (channel_id, alias) -> 여러 (product_id, quantity)
    # - 채널별 별칭을 우선 적용하고, 없으면 공용(channel_id IS NULL)을 fallback으로 사용한다.
    alias_result = await db.execute(
        select(ProductAliasItem)
        .join(ProductAliasDict, ProductAliasItem.alias_id == ProductAliasDict.id)
        .filter(
            ProductAliasItem.is_delete == False,
            ProductAliasDict.is_delete == False,
        )
        .options(selectinload(ProductAliasItem.alias))
    )
    alias_items = alias_result.scalars().all()
    alias_by_channel_and_name: dict[tuple[str | None, str], list[ProductAliasItem]] = {}
    for item in alias_items:
        alias_name = (item.alias.alias or "").strip()
        ch_key = str(getattr(item.alias, "channel_id", None)) if getattr(item.alias, "channel_id", None) else None
        alias_by_channel_and_name.setdefault((ch_key, alias_name), []).append(item)

    preview_rows: list[ExcelOrderPreviewRowOut] = []

    # Excel 데이터 행 인덱스는 1부터 시작 (헤더 제외한 프리뷰용)
    for excel_idx, excel_row in enumerate(rows_iter, start=1):
        errors: list[str] = []
        matched_product_id = None
        matched_product_label = None
        match_type = "none"
        alias_quantity: int | None = None
        matched_items: list[ExcelOrderUploadItem] = []
        commission: int | None = 0

        raw = _build_raw_row_from_excel(headers_full, excel_row)

        def get_cell(canonical: str):
            if canonical_to_excel:
                excel_h = canonical_to_excel.get(canonical)
                if not excel_h:
                    return None
            else:
                excel_h = canonical
            idx = header_to_idx.get(excel_h)
            if idx is None:
                return None
            return excel_row[idx]

        channel_name = selected_channel_name or _cell_to_str(get_cell("채널"))
        mall_product_name = _cell_to_str(get_cell("상품명"))

        if not channel_name:
            errors.append("채널이 비어있습니다.")
        if not mall_product_name:
            errors.append("상품명이 비어있습니다.")

        # 수량 형식 검증 (엑셀 원본 기준) + 값 파싱
        excel_quantity: int | None = None
        try:
            excel_quantity = _cell_to_int(get_cell("수량"))
        except Exception:
            errors.append("수량이 올바르지 않습니다.")
        try:
            _cell_to_decimal(get_cell("총 주문금액"))
        except Exception:
            errors.append("총 주문금액이 올바르지 않습니다.")

        channel = selected_channel
        if channel is None and channel_name:
            ch_res = await db.execute(
                select(Channel)
                .filter(Channel.name == channel_name, Channel.is_delete == False)
                .options(selectinload(Channel.current_mapping_version))
            )
            channel = ch_res.scalars().first()
            if not channel:
                errors.append(f"채널을 찾을 수 없습니다: {channel_name}")

        if channel and mall_product_name:
            name_key = mall_product_name.strip()

            # 1) 상품 별칭과 완전 일치하는 경우: alias 사용 (복수 매핑 지원)
            ch_key = str(channel.id) if channel is not None else None
            alias_list = (alias_by_channel_and_name.get((ch_key, name_key), []) or []) + (
                alias_by_channel_and_name.get((None, name_key), []) or []
            )
            if alias_list:
                # 기본 계수: 엑셀 수량이 있으면 그 값, 없으면 1
                factor = excel_quantity or 1
                commission = int(getattr(getattr(alias_list[0], "alias", None), "commission", 0) or 0)
                for item in alias_list:
                    qty = (item.quantity or 0) * factor
                    if qty <= 0:
                        continue
                    matched_items.append(
                        ExcelOrderUploadItem(
                            product_id=item.product_id,
                            quantity=qty,
                        )
                    )
                if matched_items:
                    # 프리뷰의 단일 필드는 첫 번째 항목 기준으로 유지 (기존 컬럼 호환)
                    first_item = alias_list[0]
                    matched_product_id = first_item.product_id
                    alias_quantity = excel_quantity
                    match_type = "alias"
            else:
                # 2) fallback: Product.name 과 exact(정규화) 매칭(편의 기능)
                if not matched_product_id:
                    for p in active_products:
                        if _norm(p.name) == _norm(mall_product_name):
                            matched_product_id = p.id
                            match_type = "product_name"
                            break

        if matched_product_id:
            p = product_by_id.get(matched_product_id)
            if p:
                matched_product_label = f"[{p.product_code}] {p.name}"

        preview_rows.append(
            ExcelOrderPreviewRowOut(
                row_index=excel_idx,
                channel=channel_name,
                product_name=mall_product_name,
                raw=raw,
                matched_product_id=matched_product_id,
                matched_product_label=matched_product_label,
                match_type=match_type,
                alias_quantity=alias_quantity,
                commission=commission,
                matched_items=matched_items,
                errors=errors,
            )
        )

    return ExcelOrderPreviewResponse(
        rows=preview_rows,
        active_products=[ProductRead.model_validate(p) for p in active_products],
        warnings=preview_warnings,
    )


@router.post("/excel/upload", response_model=ExcelOrderUploadResponse)
async def upload_excel_orders(
    payload: ExcelOrderUploadRequest,
    db: AsyncSession = Depends(get_async_session),
    user: User = Depends(current_active_user),
):
    _ = await require_permission("orders", "create")(user, db)
    created_orders = 0
    created_receivers = 0

    for row in payload.rows:
        # channel lookup (유효성 체크 용도)
        ch_res = await db.execute(
            select(Channel)
            .filter(Channel.name == row.channel, Channel.is_delete == False)
            .options(selectinload(Channel.current_mapping_version))
        )
        channel = ch_res.scalars().first()
        if not channel:
            raise HTTPException(status_code=422, detail=f"Unknown channel: {row.channel}")

        if getattr(channel, "current_mapping_version_id", None) is None:
            raise HTTPException(
                status_code=422,
                detail="선택한 채널에 주문 엑셀 매핑(버전)이 없습니다. 채널 설정에서 엑셀 매핑을 먼저 저장해 주세요.",
            )

        parsed_map = _parse_channel_order_excel_mapping(
            channel.current_mapping_version.mapping if channel.current_mapping_version is not None else None
        )
        excel_by_canonical = parsed_map[1] if parsed_map is not None else None

        # raw: 엑셀 원본 헤더 키 전체. 표준 필드 값은 채널 매핑(또는 레거시 한글 키)으로 조회.
        raw = row.raw or {}
        receiver_name = str(_get_raw_field(raw, "수취인명", excel_by_canonical) or "").strip()
        receiver_phone = str(_get_raw_field(raw, "수취인연락처", excel_by_canonical) or "").strip()
        zip_code = str(_get_raw_field(raw, "우편번호", excel_by_canonical) or "").strip()
        integrated_address = str(_get_raw_field(raw, "통합배송지", excel_by_canonical) or "").strip()
        sm = _get_raw_field(raw, "배송메세지", excel_by_canonical)
        shipping_message = str(sm).strip() if sm is not None else None
        mall_product_name = str(_get_raw_field(raw, "상품명", excel_by_canonical) or "").strip()

        if not row.items:
            raise HTTPException(status_code=422, detail="No items in row payload")

        # 프리뷰 단계에서 결정된 수량(별칭 또는 엑셀/수정 값)을 그대로 사용
        total_quantity = 0
        for item in row.items:
            if item.quantity <= 0:
                raise HTTPException(status_code=422, detail="Invalid quantity in payload")
            total_quantity += item.quantity

        try:
            total_price = Decimal(
                str(_get_raw_field(raw, "총 주문금액", excel_by_canonical) or "")
                .replace(",", "")
                .strip()
            )
        except Exception:
            raise HTTPException(status_code=422, detail="Invalid total_price in raw")

        # 수수료(원): 프리뷰에서 자동 입력된 값(사용자 수정 가능)을 우선 사용.
        # 값이 없으면 상품 별칭(채널별 우선, 없으면 공용) 매칭 결과로 계산해 스냅샷 저장.
        commission: int | None = getattr(row, "commission", None)
        if commission is None:
            commission = 0
            if mall_product_name:
                alias_res = await db.execute(
                    select(ProductAliasDict).where(
                        ProductAliasDict.is_delete == False,
                        ProductAliasDict.alias == mall_product_name,
                        or_(
                            ProductAliasDict.channel_id == channel.id,
                            ProductAliasDict.channel_id.is_(None),
                        ),
                    )
                )
                alias_dicts = alias_res.scalars().all()
                # prefer channel-specific
                chosen = next(
                    (a for a in alias_dicts if getattr(a, "channel_id", None) == channel.id),
                    None,
                )
                if chosen is None:
                    chosen = next(
                        (a for a in alias_dicts if getattr(a, "channel_id", None) is None),
                        None,
                    )
                commission = int(getattr(chosen, "commission", 0) or 0) if chosen is not None else 0
        else:
            try:
                commission = int(commission)
            except Exception:
                raise HTTPException(status_code=422, detail="Invalid commission in payload")
            if commission < 0:
                raise HTTPException(status_code=422, detail="Invalid commission in payload")

        receiver = Receiver(
            name=receiver_name,
            phone=receiver_phone,
            zip_code=zip_code,
            address=integrated_address,
            address_detail=None,
            email=None,
        )
        db.add(receiver)
        await db.flush()
        created_receivers += 1

        order = Order(
            channel_id=channel.id,
            channel_mapping_version_id=channel.current_mapping_version_id,
            receiver_id=receiver.id,
            mall_product_name=mall_product_name or None,
            raw=raw,
            price=total_price,
            commission=commission,
            quantity=total_quantity,
            memo=shipping_message or None,
            update_user_id=user.id,
        )
        db.add(order)
        await db.flush()

        for item in row.items:
            oi = OrderItem(
                order_id=order.id,
                product_id=item.product_id,
                quantity=item.quantity,
            )
            db.add(oi)

        created_orders += 1

    await db.commit()
    return ExcelOrderUploadResponse(
        created_orders=created_orders,
        created_receivers=created_receivers,
    )


@router.get("/export/raw-by-channel/preview")
async def preview_raw_export_by_channel(
    channel_id: UUID = Query(...),
    receiver_id: UUID | None = None,
    channel_name: str | None = Query(default=None),
    receiver_name: str | None = Query(default=None),
    receiver_phone: str | None = Query(default=None),
    receiver_zip_code: str | None = Query(default=None),
    receiver_address: str | None = Query(default=None),
    invoice_number: str | None = Query(default=None),
    product_query: str | None = Query(default=None),
    order_date_start: date | None = Query(default=None),
    order_date_end: date | None = Query(default=None),
    has_memos: bool | None = Query(default=None),
    db: AsyncSession = Depends(get_async_session),
    user: User = Depends(current_active_user),
):
    _ = await require_permission("orders", "read")(user, db)

    q = (
        select(Order)
        .filter(
            Order.is_delete == False,
            Order.channel_id == channel_id,
            Order.status == OrderStatus.SHIPPING,
            Order.raw.isnot(None),
        )
        .options(selectinload(Order.shipments))
    )
    q = _apply_order_filters(
        q,
        receiver_id=receiver_id,
        status=None,  # fixed to SHIPPING above
        channel_id=channel_id,
        channel_name=channel_name,
        receiver_name=receiver_name,
        receiver_phone=receiver_phone,
        receiver_zip_code=receiver_zip_code,
        receiver_address=receiver_address,
        invoice_number=invoice_number,
        product_query=product_query,
        order_date_start=order_date_start,
        order_date_end=order_date_end,
        has_memos=has_memos,
    )
    q = q.order_by(Order.updated_at.desc(), Order.id.desc()).limit(1)
    res = await db.execute(q)
    o = res.scalars().first()
    if not o or not isinstance(o.raw, dict):
        return {"found": False, "headers": [], "row": [], "version_id": None}

    version_id = getattr(o, "channel_mapping_version_id", None)
    mapping = await _load_mapping_by_version_id(db, version_id)
    raw_keys = set(o.raw.keys())
    raw_headers = _build_headers_for_version(mapping, raw_keys)
    headers = list(raw_headers) + list(_RAW_EXPORT_META_HEADERS)
    row = [_json_cell_for_excel(o.raw.get(h)) for h in raw_headers]
    row.extend(_raw_export_meta_cells(o))
    return {"found": True, "headers": headers, "row": row, "version_id": str(version_id) if version_id else None}


@router.get("/export/raw-by-channel")
async def export_raw_by_channel(
    channel_id: UUID = Query(...),
    mode: str | None = Query(default="auto", description="auto|xlsx|zip"),
    receiver_id: UUID | None = None,
    status: OrderStatus | None = None,
    channel_name: str | None = Query(default=None),
    receiver_name: str | None = Query(default=None),
    receiver_phone: str | None = Query(default=None),
    receiver_zip_code: str | None = Query(default=None),
    receiver_address: str | None = Query(default=None),
    invoice_number: str | None = Query(default=None),
    product_query: str | None = Query(default=None),
    order_date_start: date | None = Query(default=None),
    order_date_end: date | None = Query(default=None),
    has_memos: bool | None = Query(default=None),
    db: AsyncSession = Depends(get_async_session),
    user: User = Depends(current_active_user),
):
    _ = await require_permission("orders", "read")(user, db)

    q = (
        select(Order)
        .filter(
            Order.is_delete == False,
            Order.channel_id == channel_id,
            Order.raw.isnot(None),
        )
        .options(selectinload(Order.shipments))
    )
    q = _apply_order_filters(
        q,
        receiver_id=receiver_id,
        status=status,
        channel_id=channel_id,
        channel_name=channel_name,
        receiver_name=receiver_name,
        receiver_phone=receiver_phone,
        receiver_zip_code=receiver_zip_code,
        receiver_address=receiver_address,
        invoice_number=invoice_number,
        product_query=product_query,
        order_date_start=order_date_start,
        order_date_end=order_date_end,
        has_memos=has_memos,
    )
    q = q.order_by(Order.id.asc()).limit(EXPORT_MAX_ROWS + 1)
    res = await db.execute(q)
    orders = res.scalars().all()
    if len(orders) > EXPORT_MAX_ROWS:
        raise HTTPException(
            status_code=422,
            detail=f"다운로드 대상이 너무 많습니다. (최대 {EXPORT_MAX_ROWS}건) 검색 조건을 좁혀 주세요.",
        )
    if not orders:
        raise HTTPException(status_code=404, detail="다운로드할 주문이 없습니다.")

    # group by mapping version
    grouped: dict[str, list[Order]] = {}
    for o in orders:
        vid = getattr(o, "channel_mapping_version_id", None)
        key = str(vid) if vid else "__none__"
        grouped.setdefault(key, []).append(o)

    # load channel name for filename
    ch_res = await db.execute(select(Channel).filter(Channel.id == channel_id, Channel.is_delete == False))
    ch = ch_res.scalars().first()
    channel_name_safe = (getattr(ch, "name", None) or "channel").strip() or "channel"

    def build_xlsx_bytes(headers: list[str], rows: list[list[object]]) -> bytes:
        from openpyxl import Workbook

        wb = Workbook(write_only=True)
        ws = wb.create_sheet("orders")
        _append_excel_header(ws, headers)
        for r in rows:
            ws.append(r)
        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()

    # Decide output
    want_zip = False
    if (mode or "auto") == "zip":
        want_zip = True
    elif (mode or "auto") == "xlsx":
        want_zip = False
    else:
        want_zip = len(grouped.keys()) > 1

    if not want_zip and len(grouped.keys()) == 1:
        version_key = next(iter(grouped.keys()))
        version_id = None if version_key == "__none__" else UUID(version_key)
        mapping = await _load_mapping_by_version_id(db, version_id)
        raw_keys_union: set[str] = set()
        for o in grouped[version_key]:
            if isinstance(o.raw, dict):
                raw_keys_union.update(o.raw.keys())
        raw_headers = _build_headers_for_version(mapping, raw_keys_union)
        headers = list(raw_headers) + list(_RAW_EXPORT_META_HEADERS)
        rows = []
        for o in grouped[version_key]:
            raw = o.raw if isinstance(o.raw, dict) else {}
            row = [_json_cell_for_excel(raw.get(h)) for h in raw_headers]
            row.extend(_raw_export_meta_cells(o))
            rows.append(row)
        xlsx = build_xlsx_bytes(headers, rows)
        filename_utf8 = f"orders_raw_{channel_name_safe}.xlsx"
        filename_ascii = f"orders_raw_{str(channel_id)[:8]}.xlsx"
        disposition = (
            f'attachment; filename="{filename_ascii}"; '
            f"filename*=UTF-8''{quote(filename_utf8)}"
        )
        return StreamingResponse(
            BytesIO(xlsx),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": disposition},
        )

    # zip output (multi-version or forced)
    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for version_key, items in grouped.items():
            version_id = None if version_key == "__none__" else UUID(version_key)
            mapping = await _load_mapping_by_version_id(db, version_id)
            raw_keys_union: set[str] = set()
            for o in items:
                if isinstance(o.raw, dict):
                    raw_keys_union.update(o.raw.keys())
            raw_headers = _build_headers_for_version(mapping, raw_keys_union)
            headers = list(raw_headers) + list(_RAW_EXPORT_META_HEADERS)
            rows = []
            for o in items:
                raw = o.raw if isinstance(o.raw, dict) else {}
                row = [_json_cell_for_excel(raw.get(h)) for h in raw_headers]
                row.extend(_raw_export_meta_cells(o))
                rows.append(row)
            xlsx = build_xlsx_bytes(headers, rows)
            suffix = "legacy" if version_id is None else str(version_id)[:8]
            zf.writestr(f"orders_raw_{channel_name_safe}_v{suffix}.xlsx", xlsx)
    zip_buf.seek(0)
    filename_utf8 = f"orders_raw_{channel_name_safe}.zip"
    filename_ascii = f"orders_raw_{str(channel_id)[:8]}.zip"
    disposition = (
        f'attachment; filename="{filename_ascii}"; '
        f"filename*=UTF-8''{quote(filename_utf8)}"
    )
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": disposition},
    )
